from __future__ import annotations

import io
import html
import json
import os
import re
import sqlite3
import subprocess
import sys
import threading
import time
import unicodedata
import urllib.parse
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta, timezone
from difflib import SequenceMatcher
from email.utils import parsedate_to_datetime
from pathlib import Path
from urllib.parse import urljoin

import pandas as pd
import numpy as np
import requests
import streamlit as st
import altair as alt
import plotly.express as px
import streamlit.components.v1 as components
from st_aggrid import AgGrid, ColumnsAutoSizeMode, DataReturnMode, GridOptionsBuilder, JsCode
from streamlit.runtime.scriptrunner import RerunData, get_script_run_ctx
from streamlit.runtime import Runtime
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
SIMPLE_SITE = os.environ.get("FUBON_SIMPLE_SITE", "").strip() == "1"
APP_TITLE = "美股負面新聞風險中心"
APP_VERSION = "雙方法合併精簡版 1.0" if SIMPLE_SITE else "V4"
TAIPEI = timezone(timedelta(hours=8))
DATA_ROOT = Path(os.environ.get(
    "FUBON_V4_DATA_DIR",
    str(BASE_DIR / "data"),
))
OUTPUT_DIR = Path(os.environ.get(
    "FUBON_V4_OUTPUT_DIR",
    str(BASE_DIR / "每日資料"),
))
UPLOAD_DIR = DATA_ROOT / "uploads"
CONFIG_DIR = DATA_ROOT / "config"
DB_PATH = DATA_ROOT / "system.db"
RULE_PATH = CONFIG_DIR / "Parameter_Event.xlsx"
RULE_SEED_PATH = CONFIG_DIR / "Parameter_Event.xlsx"
COMPANY_INDEX_PATH = BASE_DIR / "美股指數成分股.xlsx"
EXPOSURE_PATH = CONFIG_DIR / "Exposure_Positions.xlsx"
SEC13F_DIR = DATA_ROOT / "13f"
SEC13F_PATH = SEC13F_DIR / "13F_latest.xlsx"
FINBERT_ONNX_DIR = DATA_ROOT / "models" / "finbert_onnx"
BACKUP_PATH = DATA_ROOT / "backups" / "美股負面新聞系統_V4_20260810_需求修改前.bak"

NEWS_COLUMNS = ["Ticker", "Company", "Published Time", "Title", "Source", "URL", "FinBERT"]
NEGATIVE_OUTPUT_COLUMNS = [
    "Published Time", "Ticker", "Company", "Title", "Title_ZH", "FinBERT",
    "Event_type", "Event_Code", "事件中文", "Level", "Action", "Keyword", "URL", "Source",
]

DOW_30 = [
    ("MMM", "3M"), ("AXP", "American Express"), ("AMGN", "Amgen"),
    ("AMZN", "Amazon"), ("AAPL", "Apple"), ("BA", "Boeing"),
    ("CAT", "Caterpillar"), ("CVX", "Chevron"), ("CSCO", "Cisco"),
    ("KO", "Coca-Cola"), ("DIS", "Disney"), ("GS", "Goldman Sachs"),
    ("HD", "Home Depot"), ("HON", "Honeywell"), ("IBM", "IBM"),
    ("JNJ", "Johnson & Johnson"), ("JPM", "JPMorgan Chase"),
    ("MCD", "McDonald's"), ("MRK", "Merck"), ("MSFT", "Microsoft"),
    ("NKE", "Nike"), ("NVDA", "NVIDIA"), ("PG", "Procter & Gamble"),
    ("CRM", "Salesforce"), ("SHW", "Sherwin-Williams"),
    ("TRV", "Travelers"), ("UNH", "UnitedHealth"), ("VZ", "Verizon"),
    ("V", "Visa"), ("WMT", "Walmart"),
]


def http_get(url: str, **kwargs) -> requests.Response:
    """一般連線若被失效的系統 Proxy 阻擋，自動改用直接連線重試。"""
    try:
        return requests.get(url, **kwargs)
    except requests.exceptions.ProxyError:
        direct_session = requests.Session()
        direct_session.trust_env = False
        return direct_session.get(url, **kwargs)


def http_post(url: str, **kwargs) -> requests.Response:
    """POST 版連線；翻譯批次不用把大量文字塞進網址，避免網址過長導致整批失敗。"""
    try:
        return requests.post(url, **kwargs)
    except requests.exceptions.ProxyError:
        direct_session = requests.Session()
        direct_session.trust_env = False
        return direct_session.post(url, **kwargs)


@st.cache_data(ttl=86400)
def load_sp500() -> pd.DataFrame:
    """每次快取到期後由公開成分表重新取得，不依賴舊專案名單。"""
    url = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
    response = http_get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
    response.raise_for_status()
    table = pd.read_html(io.StringIO(response.text), match="Symbol")[0]
    result = table[["Symbol", "Security"]].rename(columns={"Symbol": "Ticker", "Security": "Company"})
    result["Ticker"] = result["Ticker"].astype(str).str.replace(".", "-", regex=False).str.upper()
    result["Company"] = result["Company"].astype(str).str.strip()
    return result.drop_duplicates("Ticker").reset_index(drop=True)


def initialize() -> None:
    for folder in (DATA_ROOT, OUTPUT_DIR, UPLOAD_DIR, CONFIG_DIR):
        folder.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(DB_PATH) as connection:
        connection.executescript("""
            CREATE TABLE IF NOT EXISTS runs (
                id INTEGER PRIMARY KEY, started_at TEXT, finished_at TEXT, status TEXT,
                start_time TEXT, end_time TEXT, universe TEXT, rows INTEGER, output_path TEXT, message TEXT
            );
            CREATE TABLE IF NOT EXISTS manual_status (
                company_key TEXT PRIMARY KEY, ticker TEXT, company TEXT, status TEXT,
                owner TEXT, next_review TEXT, note TEXT, updated_at TEXT
            );
            CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT);
            CREATE TABLE IF NOT EXISTS finbert_cache (
                title_key TEXT PRIMARY KEY, title TEXT NOT NULL, score REAL NOT NULL, updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS translation_cache (
                text_key TEXT NOT NULL, target_language TEXT NOT NULL,
                source_text TEXT NOT NULL, translated_text TEXT NOT NULL, updated_at TEXT NOT NULL,
                PRIMARY KEY (text_key, target_language)
            );
        """)
    if not RULE_PATH.exists():
        if not RULE_SEED_PATH.is_file():
            raise FileNotFoundError(f"找不到外部關鍵字檔：{RULE_SEED_PATH}")
        RULE_PATH.write_bytes(RULE_SEED_PATH.read_bytes())


def last_success_end() -> datetime | None:
    with sqlite3.connect(DB_PATH) as connection:
        row = connection.execute(
            "SELECT end_time FROM runs WHERE status='success' ORDER BY id DESC LIMIT 1"
        ).fetchone()
    if not row or not row[0]:
        return None
    try:
        return datetime.fromisoformat(row[0]).astimezone(TAIPEI)
    except ValueError:
        return None


def latest_crawl_result(today_only: bool = False) -> dict | None:
    """從正式執行紀錄找回結果，不依賴目前瀏覽器的 session。"""
    with sqlite3.connect(DB_PATH) as connection:
        rows = connection.execute(
            """SELECT started_at,finished_at,start_time,end_time,universe,rows,output_path,message
               FROM runs WHERE status='success' AND (universe LIKE '方法一｜%' OR universe LIKE '方法二｜%' OR universe LIKE '方法一＋方法二｜%')
               ORDER BY id DESC"""
        ).fetchall()
    today = datetime.now(TAIPEI).date()
    for row in rows:
        try:
            end_at = datetime.fromisoformat(row[3]).astimezone(TAIPEI)
        except (TypeError, ValueError):
            continue
        if today_only and end_at.date() != today:
            continue
        recorded_path = Path(str(row[6]))
        # V4 會複製 V3 的歷史資料庫；舊紀錄可能保存 V3 的絕對路徑。
        # 一律依檔名重新定位到 V4 自己的 output，確保兩個版本完全獨立。
        output_path = OUTPUT_DIR / recorded_path.name
        if not output_path.is_file():
            continue
        method = str(row[4]).split("｜", 1)[0]
        stamp = end_at.strftime("%Y%m%d")
        event_candidate = OUTPUT_DIR / f"美股_{stamp}_負面新聞爬蟲.xlsx"
        event_path = event_candidate if event_candidate.is_file() else None
        finbert_path = None
        if method in ("方法二", "方法一＋方法二"):
            finbert_name = output_path.name.replace("MergeNews_", "MergeNews_FinBERT_").replace("Combined_News_", "Combined_FinBERT_News_").replace("All_News_", "FinBERT_News_")
            candidate = output_path.with_name(finbert_name)
            finbert_path = candidate if candidate.is_file() else None
        return {
            "started_at": row[0], "finished_at": row[1], "start_time": row[2], "end_time": row[3],
            "universe": row[4], "rows": int(row[5] or 0), "path": output_path,
            "event_path": event_path, "finbert_path": finbert_path, "message": row[7] or "", "method": method,
            "is_today": end_at.date() == today,
        }
    return None


def create_run(start: datetime, end: datetime, universe: str) -> int:
    with sqlite3.connect(DB_PATH) as connection:
        cursor = connection.execute(
            "INSERT INTO runs(started_at,status,start_time,end_time,universe,rows,message) VALUES(?,?,?,?,?,?,?)",
            (datetime.now(TAIPEI).isoformat(timespec="seconds"), "running", start.isoformat(), end.isoformat(), universe, 0, ""),
        )
        return int(cursor.lastrowid)


def finish_run(run_id: int, status: str, rows: int = 0, output_path: str = "", message: str = "") -> None:
    with sqlite3.connect(DB_PATH) as connection:
        connection.execute(
            "UPDATE runs SET finished_at=?,status=?,rows=?,output_path=?,message=? WHERE id=?",
            (datetime.now(TAIPEI).isoformat(timespec="seconds"), status, rows, output_path, message, run_id),
        )


class CrawlCancelled(Exception):
    """使用者主動停止抓取。"""


_ACTIVE_CRAWL_STOP_EVENT = None


def check_crawl_cancelled() -> None:
    """讓網路、翻譯與模型工作在細小步驟間立即回應停止要求。"""
    event = _ACTIVE_CRAWL_STOP_EVENT
    if event is not None and event.is_set():
        raise CrawlCancelled("使用者已停止抓取")


class BackgroundProgress:
    def __init__(self, job: dict):
        self.job = job

    def progress(self, value=0, text=""):
        if self.job["stop_event"].is_set():
            raise CrawlCancelled("使用者已停止抓取")
        self.job["progress"] = max(0.0, min(float(value or 0), 1.0))
        if text:
            self.job["progress_text"] = str(text)
        return self


@st.cache_resource
def crawl_job_registry() -> dict:
    return {}


def format_elapsed(seconds: float) -> str:
    total = max(0, int(seconds))
    hours, remainder = divmod(total, 3600)
    minutes, secs = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"


def begin_job_stage(job: dict, label: str) -> None:
    """切換主要執行階段，並保留前一階段的花費時間。"""
    now = time.monotonic()
    previous = job.get("current_stage")
    previous_started = job.get("current_stage_started")
    if previous and previous_started is not None:
        job.setdefault("stage_times", []).append({
            "label": previous,
            "seconds": max(0.0, now - float(previous_started)),
        })
    job["current_stage"] = label
    job["current_stage_started"] = now


def finish_job_stage(job: dict) -> None:
    now = time.monotonic()
    current = job.pop("current_stage", None)
    started = job.pop("current_stage_started", None)
    if current and started is not None:
        job.setdefault("stage_times", []).append({
            "label": current,
            "seconds": max(0.0, now - float(started)),
        })


def news_window_stats(frame: pd.DataFrame, start: datetime, end: datetime) -> dict:
    if frame.empty:
        return {"first_news": "—", "last_news": "—", "outside_window": 0}
    published = pd.to_datetime(frame["Published Time"], errors="coerce")
    start_value = start.astimezone(TAIPEI).replace(tzinfo=None)
    end_value = end.astimezone(TAIPEI).replace(tzinfo=None)
    valid = published.dropna()
    outside = int((~published.between(start_value, end_value, inclusive="both")).sum())
    return {
        "first_news": valid.min().strftime("%m/%d %H:%M") if not valid.empty else "—",
        "last_news": valid.max().strftime("%m/%d %H:%M") if not valid.empty else "—",
        "outside_window": outside,
    }


def run_crawl_job(job: dict, method_name: str, universe: str, companies: pd.DataFrame,
                  start_dt: datetime, end_dt: datetime) -> None:
    global _ACTIVE_CRAWL_STOP_EVENT
    run_id = job["run_id"]
    _ACTIVE_CRAWL_STOP_EVENT = job["stop_event"]
    progress = BackgroundProgress(job)
    active_method_key = ""
    english_title_failures = 0
    try:
        stamp = end_dt.strftime("%Y%m%d")
        if method_name == "方法一":
            begin_job_stage(job, "方法一｜多來源新聞擷取")
            active_method_key = "method1_monitor"
            job[active_method_key].update(status="running", stage="建立多來源新聞池")
            progress.progress(0, text="方法一正在建立新聞池")
            frame, all_errors, source_counts = fetch_method1_news(companies, start_dt, end_dt, progress)
            frame = prepare_english_titles(frame, progress)
            english_title_failures = int(frame.attrs.get("english_translation_failures", 0))
            job[active_method_key].update(
                status="success", stage="完成", rows=len(frame), errors=list(all_errors),
                source_counts=dict(source_counts), **news_window_stats(frame, start_dt, end_dt),
            )
            path = OUTPUT_DIR / f"RawNews_{stamp}.xlsx"
            begin_job_stage(job, "產生今日新聞檔案")
            source_summary = pd.DataFrame([{"來源": source, "新聞筆數": count} for source, count in source_counts.items()])
            progress.progress(0.94, text="正在產生新聞檔案")
            write_excel(path, {"新聞": frame, "來源摘要": source_summary})
            job["summary"] = f"方法一已完成：取得 {len(frame)} 則新聞"
        elif method_name == "方法二":
            begin_job_stage(job, "方法二｜Google News 擷取")
            active_method_key = "method2_monitor"
            job[active_method_key].update(status="running", stage="分批抓取 Google News")
            progress.progress(0, text="方法二正在分批爬取 Google News")
            raw_frame, all_errors, batch_count = fetch_method2_raw(companies, start_dt, end_dt, progress)
            job[active_method_key].update(stage="金融情緒分析 (FinBERT) 中", raw_rows=len(raw_frame), batches=batch_count, errors=list(all_errors))
            begin_job_stage(job, "金融情緒分析 (FinBERT)")
            frame, negative_frame = score_method2_finbert(raw_frame, progress)
            english_title_failures = int(frame.attrs.get("english_translation_failures", 0))
            job[active_method_key].update(
                status="success", stage="完成", rows=len(frame), negative_rows=len(negative_frame),
                batches=batch_count, errors=list(all_errors), **news_window_stats(frame, start_dt, end_dt),
            )
            path = OUTPUT_DIR / f"All_News_{stamp}.xlsx"
            finbert_path = OUTPUT_DIR / f"FinBERT_News_{stamp}.xlsx"
            begin_job_stage(job, "產生今日新聞檔案")
            progress.progress(0.94, text="正在產生新聞檔案")
            write_excel(path, {"All_News": frame})
            write_excel(finbert_path, {"FinBERT_News": negative_frame})
            job["summary"] = f"方法二已完成：{batch_count} 批、爬得 {len(frame)} 則；金融情緒分析 (FinBERT) ≤ 0 共 {len(negative_frame)} 則"
        else:
            begin_job_stage(job, "方法一｜多來源新聞擷取")
            active_method_key = "method1_monitor"
            job[active_method_key].update(status="running", stage="建立多來源新聞池")
            progress.progress(0, text="完整整合 1/4｜執行方法一")
            method1_frame, method1_errors, source_counts = fetch_method1_news(companies, start_dt, end_dt, progress)
            job[active_method_key].update(
                status="success", stage="完成", rows=len(method1_frame), errors=list(method1_errors),
                source_counts=dict(source_counts), **news_window_stats(method1_frame, start_dt, end_dt),
            )
            active_method_key = "method2_monitor"
            begin_job_stage(job, "方法二｜Google News 擷取")
            job[active_method_key].update(status="running", stage="分批抓取 Google News")
            progress.progress(0, text="完整整合 2/4｜執行方法二 Google News")
            method2_frame, method2_errors, batch_count = fetch_method2_raw(companies, start_dt, end_dt, progress)
            job[active_method_key].update(
                stage="金融情緒分析 (FinBERT) 中", raw_rows=len(method2_frame), batches=batch_count,
                errors=list(method2_errors),
            )
            begin_job_stage(job, "合併並刪除重複")
            progress.progress(0.55, text="完整整合 3/4｜合併並去除重複新聞")
            merged_frame, duplicate_log = merge_frames(method1_frame, method2_frame)
            begin_job_stage(job, "金融情緒分析 (FinBERT)")
            progress.progress(0.58, text="完整整合 4/4｜對全部整合新聞執行金融情緒分析 (FinBERT)")
            frame, negative_frame = score_method2_finbert(merged_frame, progress)
            english_title_failures = int(frame.attrs.get("english_translation_failures", 0))
            job[active_method_key].update(
                status="success", stage="完成", rows=len(method2_frame), merged_rows=len(frame),
                negative_rows=len(negative_frame), batches=batch_count, errors=list(method2_errors),
                **news_window_stats(method2_frame, start_dt, end_dt),
            )
            all_errors = method1_errors + method2_errors
            path = OUTPUT_DIR / f"MergeNews_{stamp}.xlsx"
            finbert_path = OUTPUT_DIR / f"MergeNews_FinBERT_{stamp}.xlsx"
            source_summary = pd.DataFrame(
                [{"來源": source, "新聞筆數": count} for source, count in source_counts.items()]
                + [{"來源": "Google News（方法二）", "新聞筆數": len(method2_frame)}]
            )
            begin_job_stage(job, "產生完整整合檔案")
            progress.progress(0.94, text="正在產生完整整合檔案")
            write_excel(path, {"完整整合新聞": frame, "重複紀錄": duplicate_log, "來源摘要": source_summary})
            write_excel(finbert_path, {"FinBERT小於等於0": negative_frame})
            job["summary"] = f"完整整合已完成：去重後保留 {len(frame)} 則；金融情緒分析 (FinBERT) ≤ 0 共 {len(negative_frame)} 則"
        begin_job_stage(job, "負面事件分類與中文翻譯")
        progress.progress(0.97, text="正在執行負面事件分類")
        rule_modified_ns = RULE_PATH.stat().st_mtime_ns
        event_frame, unknown_frame, irrelevant_frame = classify_news_sets(
            frame, load_rules(rule_modified_ns), load_core_sources(rule_modified_ns), progress
        )
        negative_translation_failures = int(event_frame.attrs.get("translation_failures", 0))
        unknown_translation_failures = int(unknown_frame.attrs.get("translation_failures", 0))
        irrelevant_translation_failures = int(irrelevant_frame.attrs.get("translation_failures", 0))
        translation_failures = english_title_failures + negative_translation_failures + unknown_translation_failures + irrelevant_translation_failures
        event_path = OUTPUT_DIR / f"美股_{stamp}_負面新聞爬蟲.xlsx"
        begin_job_stage(job, "產生負面新聞檔案")
        write_excel(event_path, {
            "負面新聞": event_frame,
            "待人工覆核": unknown_frame,
            "無關新聞": irrelevant_frame,
        })
        finish_run(run_id, "success", len(frame), str(path), "；".join(all_errors))
        job.update(status="success", progress=1.0, progress_text="抓取與分類全部完成",
                   rows=len(frame), event_rows=len(event_frame), unknown_rows=len(unknown_frame),
                   irrelevant_rows=len(irrelevant_frame), translation_failures=translation_failures,
                   english_title_failures=english_title_failures,
                   negative_translation_failures=negative_translation_failures,
                   unknown_translation_failures=unknown_translation_failures,
                   irrelevant_translation_failures=irrelevant_translation_failures,
                   path=str(path))
    except CrawlCancelled:
        if active_method_key and job.get(active_method_key, {}).get("status") == "running":
            job[active_method_key].update(status="stopped", stage="使用者停止")
        finish_run(run_id, "stopped", message="使用者主動停止")
        job.update(status="stopped", progress_text="已停止；未完成的檔案不會列為正式結果")
    except Exception as exc:
        if active_method_key:
            job[active_method_key].update(status="failed", stage="執行失敗", fatal_error=str(exc))
        finish_run(run_id, "failed", message=str(exc))
        job.update(status="failed", error=str(exc), progress_text="執行失敗")
    finally:
        _ACTIVE_CRAWL_STOP_EVENT = None
        finish_job_stage(job)
        job["finished_monotonic"] = time.monotonic()
        job["finished_at"] = datetime.now(TAIPEI).isoformat(timespec="seconds")
        # 背景工作只在結束瞬間通知原頁面完整更新一次；執行期間不輪詢、不刷新。
        script_context = job.get("script_context")
        try:
            runtime = Runtime.instance()
            session_info = runtime._session_mgr.get_session_info(job.get("session_id", ""))
            if session_info is not None:
                session = session_info.session
                session._event_loop.call_soon_threadsafe(session.request_rerun, None)
            elif script_context is not None and script_context.script_requests is not None:
                script_context.script_requests.request_rerun(RerunData())
        except Exception:
            pass


def split_terms(value: object) -> list[str]:
    return [item.strip() for item in re.split(r"[;；\n]+", str(value or "")) if item.strip()]


@st.cache_data(ttl=60)
def load_rules(modified_ns: int) -> list[dict]:
    del modified_ns
    workbook = load_workbook(RULE_PATH, read_only=True, data_only=True)
    sheet = workbook["Event"]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    rules = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, values))
        if not str(item.get("Event_Code") or "").strip():
            continue
        item["_keywords"] = split_terms(item.get("Keyword"))
        item["_exclusions"] = split_terms(item.get("Exclude_Keyword"))
        rules.append(item)
    workbook.close()
    return rules


@st.cache_data(ttl=60)
def load_core_sources(modified_ns: int) -> set[str]:
    del modified_ns
    workbook = load_workbook(RULE_PATH, read_only=True, data_only=True)
    if "主要網站" not in workbook.sheetnames:
        workbook.close()
        raise ValueError("目前關鍵字表缺少「主要網站」工作表")
    sheet = workbook["主要網站"]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    if "Source" not in headers:
        workbook.close()
        raise ValueError("「主要網站」工作表缺少 Source 欄位")
    source_index = headers.index("Source")
    sources = {
        str(values[source_index]).strip().casefold()
        for values in sheet.iter_rows(min_row=2, values_only=True)
        if source_index < len(values) and values[source_index] not in (None, "")
    }
    workbook.close()
    return sources


def normalize_news(frame: pd.DataFrame) -> pd.DataFrame:
    # 外部工具常在欄名尾端留下不可見空白；先統一清理再做欄位對應。
    frame = frame.copy()
    frame.columns = [str(column).replace("\u3000", " ").strip() for column in frame.columns]
    aliases = {
        "Ticker": ("Ticker", "股票代號", "標的代碼"),
        "Company": ("Company", "股票名稱", "標的名稱"),
        "Published Time": ("Published Time", "發布時間", "發布日期"),
        "Title": ("Title", "新聞主旨", "主旨"),
        "Source": ("Source", "新聞來源"),
        "URL": ("URL", "新聞網址", "新聞連結"),
        "FinBERT": ("FinBERT",),
    }
    result = pd.DataFrame(index=frame.index)
    for target, names in aliases.items():
        source = next((name for name in names if name in frame.columns), None)
        result[target] = frame[source] if source else ""
    for column in NEWS_COLUMNS:
        result[column] = result[column].fillna("").astype(str).str.strip()
    result["Ticker"] = result["Ticker"].str.replace(r"\.0$", "", regex=True).str.upper()
    result = result[(result["Title"] != "") & (result["URL"] != "")]
    return result.reset_index(drop=True)


TRACKING_QUERY_NAMES = {
    "utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content",
    "from", "source", "ref", "gclid", "fbclid", "mc_cid", "mc_eid",
}


def normalize_url(value: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    try:
        parts = urllib.parse.urlsplit(raw)
        if not parts.scheme or not parts.netloc:
            return raw.rstrip("/").casefold()
        query = [
            (key, val) for key, val in urllib.parse.parse_qsl(parts.query, keep_blank_values=True)
            if key.casefold() not in TRACKING_QUERY_NAMES and not key.casefold().startswith("utm_")
        ]
        return urllib.parse.urlunsplit((
            parts.scheme.casefold(), parts.netloc.casefold(), parts.path.rstrip("/"),
            urllib.parse.urlencode(query), "",
        ))
    except ValueError:
        return raw.rstrip("/").casefold()


def url_signatures(value: str) -> set[str]:
    """EXE 同款：解開 Google 轉址、URL encoding 與常見內嵌目標網址。"""
    raw = str(value or "").strip()
    if not raw:
        return set()
    decoded_versions = [raw]
    for _ in range(3):
        decoded = urllib.parse.unquote(decoded_versions[-1])
        if decoded == decoded_versions[-1]:
            break
        decoded_versions.append(decoded)
    candidates = set(decoded_versions)
    for text in decoded_versions:
        candidates.update(re.findall(r"https?://[^\s<>'\"]+", text, flags=re.I))
        try:
            for key, target in urllib.parse.parse_qsl(urllib.parse.urlsplit(text).query, keep_blank_values=False):
                if key.casefold() in {"q", "url", "target", "dest", "destination", "continue"}:
                    candidates.add(urllib.parse.unquote(target))
        except ValueError:
            pass
    return {signature for candidate in candidates if (signature := normalize_url(candidate))}


def exact_title(value: str) -> str:
    return re.sub(r"\s+", " ", unicodedata.normalize("NFKC", str(value or ""))).strip().casefold()


def title_for_similarity(value: str) -> str:
    text = exact_title(value)
    text = re.sub(r"\b(?:19|20)\d{2}[-/]\d{1,2}[-/]\d{1,2}(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?\b", " ", text)
    text = re.sub(r"\b(?:mon|tue|wed|thu|fri|sat|sun),?\s+\d{1,2}\s+(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+(?:19|20)\d{2}(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?(?:\s+gmt)?\b", " ", text, flags=re.I)
    text = re.sub(r"\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+\d{1,2},?\s+(?:19|20)\d{2}\b", " ", text, flags=re.I)
    text = re.sub(r"\s+(?:by|via)\s+[\w.&'’-]+(?:\s+[\w.&'’-]+){0,5}\s*$", "", text, flags=re.I)
    text = re.sub(r"\s+(?:[-–—|:]\s*)?(?:www\.)?[a-z0-9][a-z0-9.-]*\.(?:com|net|org|co|io|tw|ca|uk|news)(?:\.[a-z]{2})?\s*$", "", text, flags=re.I)
    return re.sub(r"[^0-9a-z\u3400-\u9fff]+", " ", text).strip()


def normalized_title(value: str) -> str:
    """給單一方法的快速去重使用，與 EXE 的標題清理一致。"""
    return title_for_similarity(value).replace(" ", "")


def numeric_facts(value: str) -> tuple[str, ...]:
    return tuple(sorted(re.findall(r"\d+(?:\.\d+)?%?", title_for_similarity(value))))


def title_similarity(left: str, right: str) -> float:
    left_key, right_key = title_for_similarity(left), title_for_similarity(right)
    if min(len(left_key), len(right_key)) < 12 or numeric_facts(left) != numeric_facts(right):
        return 0.0
    compact_score = SequenceMatcher(None, left_key.replace(" ", ""), right_key.replace(" ", "")).ratio()
    token_score = SequenceMatcher(None, " ".join(sorted(left_key.split())), " ".join(sorted(right_key.split()))).ratio()
    return max(compact_score, token_score)


def merge_company_key(record: dict) -> tuple[str, str] | None:
    ticker = re.sub(r"[\s_\-]+", "", unicodedata.normalize("NFKC", str(record.get("Ticker", ""))).strip().casefold())
    company = re.sub(r"[\s_\-]+", "", unicodedata.normalize("NFKC", str(record.get("Company", ""))).strip().casefold())
    return (ticker, company) if ticker and company else None


def merge_frames(first: pd.DataFrame, second: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """EXE 同款去重；但保留資料較完整的一筆，並固定輸出網站七欄。"""
    input_rows = (
        first.assign(_file="檔案一").to_dict("records")
        + second.assign(_file="檔案二").to_dict("records")
    )
    kept: list[dict] = []
    duplicates: list[dict] = []
    seen_urls: dict[tuple[tuple[str, str], str], int] = {}
    seen_exact_titles: dict[tuple[tuple[str, str], str], int] = {}
    retained_by_company: dict[tuple[str, str], list[int]] = {}

    def remember(record: dict, index: int) -> None:
        company_key = merge_company_key(record)
        if company_key is None:
            return
        for signature in url_signatures(record.get("URL", "")):
            seen_urls.setdefault((company_key, signature), index)
        title_key = exact_title(record.get("Title", ""))
        if title_key:
            seen_exact_titles.setdefault((company_key, title_key), index)
        indexes = retained_by_company.setdefault(company_key, [])
        if index not in indexes:
            indexes.append(index)

    for record in input_rows:
        company_key = merge_company_key(record)
        match_index: int | None = None
        reason = ""
        if company_key is not None:
            for signature in url_signatures(record.get("URL", "")):
                match_index = seen_urls.get((company_key, signature))
                if match_index is not None:
                    reason = "AB相同且網址相同（已忽略Google轉址前綴與追蹤參數）"
                    break
            if match_index is None:
                title_key = exact_title(record.get("Title", ""))
                if title_key:
                    match_index = seen_exact_titles.get((company_key, title_key))
                    if match_index is not None:
                        reason = "AB相同且主旨完全相同"
            if match_index is None:
                best_score, best_index = 0.0, None
                for candidate_index in retained_by_company.get(company_key, []):
                    score = title_similarity(kept[candidate_index].get("Title", ""), record.get("Title", ""))
                    if score > best_score:
                        best_score, best_index = score, candidate_index
                if best_index is not None and best_score >= 0.98:
                    match_index = best_index
                    reason = f"AB相同且主旨高度相似（{best_score:.0%}，數字資訊一致）"

        if match_index is None:
            kept.append(record)
            remember(record, len(kept) - 1)
            continue

        old = kept[match_index]
        if sum(bool(str(record.get(column, "")).strip()) for column in NEWS_COLUMNS) > sum(bool(str(old.get(column, "")).strip()) for column in NEWS_COLUMNS):
            kept[match_index] = record
        remember(record, match_index)
        duplicates.append({
            "原因": reason, "Ticker": record.get("Ticker", ""), "Company": record.get("Company", ""),
            "保留標題": kept[match_index].get("Title", ""), "重複標題": record.get("Title", ""),
            "保留網址": kept[match_index].get("URL", ""), "刪除網址": record.get("URL", ""),
            "來源檔": record.get("_file", ""),
        })

    result = pd.DataFrame(kept).drop(columns=["_file"], errors="ignore")
    if result.empty:
        result = pd.DataFrame(columns=NEWS_COLUMNS)
    return result.reindex(columns=NEWS_COLUMNS), pd.DataFrame(duplicates)


def translate_text(text: str, target_language: str) -> str:
    check_crawl_cancelled()
    text = str(text or "").strip()
    if not text:
        return ""
    for attempt in range(2):
        check_crawl_cancelled()
        try:
            request_data = {"client": "gtx", "sl": "auto", "tl": target_language, "dt": "t", "q": text}
            if len(text) > 700:
                response = http_post(
                    "https://translate.googleapis.com/translate_a/single", data=request_data,
                    headers={"User-Agent": "Mozilla/5.0"}, timeout=(3, 8),
                )
            else:
                response = http_get(
                    "https://translate.googleapis.com/translate_a/single", params=request_data,
                    headers={"User-Agent": "Mozilla/5.0"}, timeout=(3, 6),
                )
            response.raise_for_status()
            payload = response.json()
            translated = "".join(str(part[0]) for part in payload[0] if part and part[0]).strip()
            if translated:
                return translated
        except CrawlCancelled:
            raise
        except Exception:
            pass
        if attempt < 1:
            if _ACTIVE_CRAWL_STOP_EVENT is not None and _ACTIVE_CRAWL_STOP_EVENT.wait(1.0):
                raise CrawlCancelled("使用者已停止抓取")
    return ""


def translate_text_zh(text: str) -> str:
    return translate_text(text, "zh-TW")


def translate_text_en(text: str) -> str:
    return translate_text(text, "en")


def translation_text_key(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip().casefold()


def load_translation_cache(texts: list[str], target_language: str) -> dict[str, str]:
    keys = list(dict.fromkeys(translation_text_key(text) for text in texts if translation_text_key(text)))
    if not keys:
        return {}
    cached: dict[str, str] = {}
    with sqlite3.connect(DB_PATH) as connection:
        for start_index in range(0, len(keys), 700):
            chunk = keys[start_index:start_index + 700]
            placeholders = ",".join("?" for _ in chunk)
            rows = connection.execute(
                f"SELECT text_key,translated_text FROM translation_cache "
                f"WHERE target_language=? AND text_key IN ({placeholders})",
                [target_language, *chunk],
            ).fetchall()
            cached.update({str(key): str(translated) for key, translated in rows if translated})
    return cached


def save_translation_cache(items: list[tuple[str, str]], target_language: str) -> None:
    now_text = datetime.now(TAIPEI).isoformat(timespec="seconds")
    rows = []
    for source_text, translated_text in items:
        key = translation_text_key(source_text)
        translated = str(translated_text or "").strip()
        if key and translated:
            rows.append((key, target_language, str(source_text), translated, now_text))
    if not rows:
        return
    with sqlite3.connect(DB_PATH) as connection:
        connection.executemany(
            "INSERT OR REPLACE INTO translation_cache"
            "(text_key,target_language,source_text,translated_text,updated_at) VALUES(?,?,?,?,?)",
            rows,
        )


def translate_title_zh(title: str) -> str:
    return translate_text_zh(title)


TRANSLATION_SEPARATOR = "<<<FUBON_NEWS_SPLIT>>>"


def make_translation_batches(items: list, text_getter, max_items: int = 40, max_characters: int = 4500) -> list[list]:
    """同時限制篇數與總字數，避免少數長標題讓整個翻譯批次失敗。"""
    batches: list[list] = []
    current: list = []
    current_characters = 0
    separator_characters = len(TRANSLATION_SEPARATOR) + 2
    for item in items:
        text_length = len(str(text_getter(item) or "").strip())
        added_characters = text_length + (separator_characters if current else 0)
        if current and (len(current) >= max_items or current_characters + added_characters > max_characters):
            batches.append(current)
            current = []
            current_characters = 0
            added_characters = text_length
        current.append(item)
        current_characters += added_characters
    if current:
        batches.append(current)
    return batches


def translate_titles_resilient(titles: list[str], target_language: str) -> list[str]:
    """批次失敗時只拆分該批；不再把整批所有標題直接改為逐篇重跑。"""
    check_crawl_cancelled()
    if not titles:
        return []
    if len(titles) == 1:
        return [translate_text(titles[0], target_language)]
    translated = translate_text(f"\n{TRANSLATION_SEPARATOR}\n".join(titles), target_language)
    parts = [part.strip() for part in translated.split(TRANSLATION_SEPARATOR)] if translated else []
    if len(parts) == len(titles) and all(parts):
        return parts
    midpoint = len(titles) // 2
    return (
        translate_titles_resilient(titles[:midpoint], target_language)
        + translate_titles_resilient(titles[midpoint:], target_language)
    )


def prepare_english_titles(frame: pd.DataFrame, progress=None) -> pd.DataFrame:
    """中文原標題存入 Title_ZH，英文翻譯寫回 Title，供 FinBERT 與關鍵字規則使用。"""
    result = frame.copy()
    if "Title_ZH" not in result.columns:
        result["Title_ZH"] = ""
    if "Title" in result.columns:
        title_zh_values = result.pop("Title_ZH")
        result.insert(result.columns.get_loc("Title") + 1, "Title_ZH", title_zh_values)
    if result.empty:
        result.attrs["english_translation_failures"] = 0
        return result
    all_chinese_indexes = [
        index for index, title in result["Title"].fillna("").astype(str).items()
        if re.search(r"[\u3400-\u9fff]", title)
    ]
    if not all_chinese_indexes:
        result.attrs["english_translation_failures"] = 0
        return result
    original_titles = [str(result.at[index, "Title"] or "").strip() for index in all_chinese_indexes]
    cached_translations = load_translation_cache(original_titles, "en")
    chinese_indexes = []
    for index, original_title in zip(all_chinese_indexes, original_titles):
        result.at[index, "Title_ZH"] = original_title
        translated_title = cached_translations.get(translation_text_key(original_title), "")
        if translated_title and not re.search(r"[\u3400-\u9fff]", translated_title):
            result.at[index, "Title"] = translated_title
        else:
            chinese_indexes.append(index)
    if not chinese_indexes:
        result.attrs["english_translation_failures"] = 0
        return result
    index_groups: dict[str, list[int]] = {}
    for index in chinese_indexes:
        index_groups.setdefault(translation_text_key(result.at[index, "Title"]), []).append(index)
    representative_indexes = [indexes[0] for indexes in index_groups.values()]
    batches = make_translation_batches(
        representative_indexes, lambda index: result.at[index, "Title"]
    )

    def translate_batch(indexes: list[int]) -> list[str]:
        check_crawl_cancelled()
        titles = [str(result.at[index, "Title"] or "").strip() for index in indexes]
        return translate_titles_resilient(titles, "en")

    failures = 0
    done = 0
    new_cache_items: list[tuple[str, str]] = []
    # 六條連線可縮短大量標題等待時間；再提高容易觸發公開翻譯服務限流。
    with ThreadPoolExecutor(max_workers=6) as executor:
        futures = {executor.submit(translate_batch, indexes): indexes for indexes in batches}
        for future in as_completed(futures):
            check_crawl_cancelled()
            indexes = futures[future]
            try:
                parts = future.result()
            except CrawlCancelled:
                raise
            except Exception:
                parts = [""] * len(indexes)
            for index, translated_title in zip(indexes, parts):
                original_title = str(result.at[index, "Title"] or "").strip()
                duplicate_indexes = index_groups.get(translation_text_key(original_title), [index])
                for duplicate_index in duplicate_indexes:
                    result.at[duplicate_index, "Title_ZH"] = str(result.at[duplicate_index, "Title"] or "").strip()
                if translated_title and not re.search(r"[\u3400-\u9fff]", translated_title):
                    for duplicate_index in duplicate_indexes:
                        result.at[duplicate_index, "Title"] = translated_title
                    new_cache_items.append((original_title, translated_title))
                else:
                    failures += len(duplicate_indexes)
            done += sum(len(index_groups.get(translation_text_key(result.at[index, "Title_ZH"] or result.at[index, "Title"]), [index])) for index in indexes)
            if progress is not None:
                progress.progress(0.55, text=f"中文標題翻譯成英文｜{min(done, len(chinese_indexes))}/{len(chinese_indexes)}")
    save_translation_cache(new_cache_items, "en")
    result.attrs["english_translation_failures"] = failures
    return result


def translate_output_rows(rows: list[dict], label: str, progress=None,
                          progress_start: float = 0.97, progress_end: float = 0.995) -> int:
    """批次翻譯輸出列；批次無法正確分割時再逐篇重試。"""
    rows_to_translate = [row for row in rows if not str(row.get("Title_ZH", "") or "").strip()]
    if not rows_to_translate:
        return 0
    cached_translations = load_translation_cache(
        [str(row.get("Title", "") or "") for row in rows_to_translate], "zh-TW"
    )
    uncached_rows = []
    for row in rows_to_translate:
        translated = cached_translations.get(translation_text_key(row.get("Title", "")), "")
        if translated:
            row["Title_ZH"] = translated
        else:
            uncached_rows.append(row)
    rows_to_translate = uncached_rows
    if not rows_to_translate:
        return 0
    failures = 0
    row_groups: dict[str, list[dict]] = {}
    for row in rows_to_translate:
        row_groups.setdefault(translation_text_key(row.get("Title", "")), []).append(row)
    representative_rows = [rows[0] for rows in row_groups.values()]
    batches = make_translation_batches(
        representative_rows, lambda row: row.get("Title", "")
    )

    def translate_batch(batch: list[dict]) -> list[str]:
        check_crawl_cancelled()
        titles = [str(row.get("Title", "") or "").strip() for row in batch]
        return translate_titles_resilient(titles, "zh-TW")

    done = 0
    new_cache_items: list[tuple[str, str]] = []
    with ThreadPoolExecutor(max_workers=6) as executor:
        futures = {executor.submit(translate_batch, batch): batch for batch in batches}
        for future in as_completed(futures):
            check_crawl_cancelled()
            batch = futures[future]
            try:
                parts = future.result()
            except CrawlCancelled:
                raise
            except Exception:
                parts = [""] * len(batch)
            completed_rows = 0
            for row, title_zh in zip(batch, parts):
                duplicate_rows = row_groups.get(translation_text_key(row.get("Title", "")), [row])
                for duplicate_row in duplicate_rows:
                    duplicate_row["Title_ZH"] = title_zh
                completed_rows += len(duplicate_rows)
                if title_zh:
                    new_cache_items.append((str(row.get("Title", "") or ""), title_zh))
                else:
                    failures += len(duplicate_rows)
            done += completed_rows
            if progress is not None:
                value = progress_start + (progress_end - progress_start) * done / len(rows_to_translate)
                progress.progress(value, text=f"翻譯{label}標題｜{done}/{len(rows_to_translate)}")
    save_translation_cache(new_cache_items, "zh-TW")
    return failures


def event_output(record: dict, rule: dict, keyword: str = "", title_zh: str | None = None) -> dict:
    return {
        "Published Time": record.get("Published Time", ""),
        "Ticker": record.get("Ticker", ""),
        "Company": record.get("Company", ""),
        "Title": record.get("Title", ""),
        "Title_ZH": record.get("Title_ZH", "") if title_zh is None else title_zh,
        "FinBERT": record.get("FinBERT", ""),
        "Event_type": rule.get("Event_type", ""),
        "Event_Code": rule.get("Event_Code", ""),
        "事件中文": rule.get("事件中文", ""),
        "Level": rule.get("Level", ""),
        "Action": rule.get("Action", ""),
        "Keyword": keyword,
        "URL": record.get("URL", ""),
        "Source": record.get("Source", ""),
    }


def classify_news_sets(frame: pd.DataFrame, rules: list[dict], core_sources: set[str], progress=None) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    negative_rows: list[dict] = []
    unknown_rows, irrelevant_rows = [], []
    unknown_rule = next((rule for rule in rules if str(rule.get("Event_Code")) == "UNKNOWN"), None)
    irrelevant_rule = next((rule for rule in rules if str(rule.get("Event_Code")) == "NOT_RELEVANT"), None)
    non_core_rule = next((rule for rule in rules if str(rule.get("Event_Code")) == "NON_CORE_SOURCE"), None)
    if unknown_rule is None or irrelevant_rule is None or non_core_rule is None:
        raise ValueError("目前關鍵字表必須包含 UNKNOWN、NOT_RELEVANT 與 NON_CORE_SOURCE 事件")
    for record in frame.to_dict("records"):
        check_crawl_cancelled()
        title = record["Title"].lower()
        matches = []
        for rule in rules:
            if not rule["_keywords"]:
                continue
            if any(term.lower() in title for term in rule["_exclusions"]):
                continue
            hits = [term for term in rule["_keywords"] if term.lower() in title]
            if hits:
                matches.append((rule, hits))
        if not matches:
            source = str(record.get("Source") or "").strip().casefold()
            if source not in core_sources:
                irrelevant_rows.append(event_output(record, non_core_rule))
            else:
                try:
                    finbert = float(record.get("FinBERT"))
                except (TypeError, ValueError):
                    finbert = None
                fallback = irrelevant_rule if finbert is not None and finbert > 0 else unknown_rule
                target = irrelevant_rows if fallback is irrelevant_rule else unknown_rows
                target.append(event_output(record, fallback))
            continue
        # 所有判定值皆取自目前 Parameter_Event.xlsx；不再另建新聞分數。
        matches.sort(key=lambda pair: (-int(pair[0].get("Primary_Priority") or 0), str(pair[0].get("Event_ID") or "")))
        primary, hits = matches[0]
        output = event_output(record, primary, "；".join(hits))
        code = str(primary.get("Event_Code") or "")
        event_type = str(primary.get("Event_type") or "").upper()
        if code == "UNKNOWN":
            unknown_rows.append(output)
        elif event_type == "SYSTEM":
            irrelevant_rows.append(output)
        else:
            negative_rows.append(output)
    negative_translation_failures = translate_output_rows(negative_rows, "負面新聞", progress, 0.970, 0.980)
    unknown_translation_failures = translate_output_rows(unknown_rows, "待人工覆核", progress, 0.980, 0.990)
    irrelevant_translation_failures = translate_output_rows(irrelevant_rows, "無關新聞", progress, 0.990, 0.998)

    def make_frame(rows: list[dict]) -> pd.DataFrame:
        result = pd.DataFrame(rows, columns=NEGATIVE_OUTPUT_COLUMNS)
        if not result.empty:
            result["Level"] = pd.to_numeric(result["Level"], errors="coerce")
            result = result.sort_values(["Level", "Published Time"], ascending=[False, False]).reset_index(drop=True)
        return result

    negative_frame = make_frame(negative_rows)
    # 正式負面新聞檔只保留規則定義的 Level 3～5；低於 Level 3 的資料不進入後續頁面。
    if not negative_frame.empty:
        negative_frame = negative_frame[negative_frame["Level"].between(3, 5, inclusive="both")].reset_index(drop=True)
    unknown_frame = make_frame(unknown_rows)
    irrelevant_frame = make_frame(irrelevant_rows)
    negative_frame.attrs["translation_failures"] = negative_translation_failures
    unknown_frame.attrs["translation_failures"] = unknown_translation_failures
    irrelevant_frame.attrs["translation_failures"] = irrelevant_translation_failures
    return negative_frame, unknown_frame, irrelevant_frame


def classify(frame: pd.DataFrame, rules: list[dict]) -> pd.DataFrame:
    """相容既有呼叫：只回傳正式負面新聞。"""
    modified_ns = RULE_PATH.stat().st_mtime_ns
    negative, _, _ = classify_news_sets(frame, rules, load_core_sources(modified_ns))
    return negative


def write_excel(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name[:31], index=False)
            sheet = writer.sheets[name[:31]]
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            is_negative_output = list(frame.columns) == NEGATIVE_OUTPUT_COLUMNS
            for index, cell in enumerate(sheet[1], 1):
                if is_negative_output:
                    cell.fill = PatternFill("solid", fgColor="FFF200" if 6 <= index <= 12 else "E2F0D9")
                    cell.font = Font(color="000000", bold=True)
                else:
                    cell.fill = PatternFill("solid", fgColor="1F4E78")
                    cell.font = Font(color="FFFFFF", bold=True)
            negative_widths = {
                "Published Time": 21, "Ticker": 12, "Company": 24, "Title": 60,
                "Title_ZH": 45, "FinBERT": 12, "Event_type": 18, "Event_Code": 27,
                "事件中文": 22, "Level": 10, "Action": 45, "Keyword": 30, "URL": 65, "Source": 18,
            }
            for index, column in enumerate(frame.columns, 1):
                width = negative_widths.get(column, 22) if is_negative_output else (80 if column in {"URL", "Definition"} else 60 if column == "Title" else 22)
                sheet.column_dimensions[get_column_letter(index)].width = width
            sheet.row_dimensions[1].height = 24
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)


METHOD1_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/126 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9,zh-TW;q=0.8",
}
METHOD1_POOL_LIMIT = 300
METHOD1_EXCHANGES = {"AAPL", "AMZN", "CSCO", "GOOGL", "MSFT", "NVDA", "AMGN", "HON"}
METHOD1_THREAD_LOCAL = threading.local()


def method1_http_session() -> requests.Session:
    session = getattr(METHOD1_THREAD_LOCAL, "session", None)
    if session is None:
        session = requests.Session()
        session.trust_env = False
        session.headers.update(METHOD1_HEADERS)
        METHOD1_THREAD_LOCAL.session = session
    return session


def parse_method1_time(value: object, reference: datetime | None = None) -> datetime | None:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if not text:
        return None
    now = reference or datetime.now(TAIPEI)
    relative = re.search(r"(\d+)\s+(minute|minutes|hour|hours|day|days)\s+ago", text, re.I)
    if relative:
        amount = int(relative.group(1))
        unit = relative.group(2).lower()
        delta = timedelta(minutes=amount) if unit.startswith("minute") else timedelta(hours=amount) if unit.startswith("hour") else timedelta(days=amount)
        return now - delta
    try:
        return parsedate_to_datetime(text).astimezone(TAIPEI)
    except Exception:
        pass
    cleaned = re.sub(r"\b(?:ET|EST|EDT|CT|CST|CDT)\b", "", text, flags=re.I).strip()
    parsed = pd.to_datetime(cleaned, errors="coerce")
    if pd.isna(parsed):
        return None
    result = parsed.to_pydatetime()
    if result.tzinfo is None:
        result = result.replace(tzinfo=TAIPEI)
    return result.astimezone(TAIPEI)


def method1_in_window(published: datetime | None, start: datetime, end: datetime) -> bool:
    return published is not None and start <= published.astimezone(TAIPEI) <= end


def filter_news_window(frame: pd.DataFrame, start: datetime, end: datetime) -> pd.DataFrame:
    """輸出前再次依台北時間嚴格檢查新聞區間，防止任何來源漏過前段篩選。"""
    if frame.empty:
        return frame.copy()
    result = frame.copy()
    published = pd.to_datetime(result["Published Time"], errors="coerce")
    if getattr(published.dt, "tz", None) is not None:
        published = published.dt.tz_convert(TAIPEI).dt.tz_localize(None)
    start_value = start.astimezone(TAIPEI).replace(tzinfo=None)
    end_value = end.astimezone(TAIPEI).replace(tzinfo=None)
    return result[published.notna() & published.between(start_value, end_value, inclusive="both")].copy()


def method1_company_terms(ticker: str, company: str) -> list[str]:
    name = re.sub(r"\s+", " ", company).strip()
    simplified = re.sub(
        r"\s+(?:incorporated|inc\.?|corporation|corp\.?|company|co\.?|limited|ltd\.?|plc|group|holdings?)$",
        "", name, flags=re.I,
    ).strip()
    terms = [name]
    if simplified and simplified.lower() != name.lower() and len(simplified) >= 4:
        terms.append(simplified)
    if ticker == "JNJ":
        terms.extend(["Johnson & Johnson", "J&J", "嬌生"])
    return list(dict.fromkeys(term for term in terms if len(term) >= 4))


def method1_match_company(title: str, summary: str, companies: pd.DataFrame) -> tuple[str, str] | None:
    text = re.sub(r"\s+", " ", f"{title} {summary}").strip()
    candidates: list[tuple[int, str, str]] = []
    for row in companies.itertuples(index=False):
        ticker, company = str(row.Ticker).upper(), str(row.Company)
        structured = re.search(rf"(?:\${re.escape(ticker)}\b|\({re.escape(ticker)}\)|(?:NASDAQ|NYSE|AMEX)\s*:\s*{re.escape(ticker)}\b)", text, re.I)
        matched_terms = [term for term in method1_company_terms(ticker, company) if re.search(rf"(?<![A-Za-z0-9]){re.escape(term)}(?![A-Za-z0-9])", text, re.I)]
        if not structured and matched_terms and company_is_information_source(text, ticker, company):
            continue
        if structured or matched_terms:
            strength = 1000 if structured else max(len(term) for term in matched_terms)
            candidates.append((strength, ticker, company))
    if not candidates:
        return None
    candidates.sort(key=lambda item: (-item[0], -len(item[2]), item[1]))
    return candidates[0][1], candidates[0][2]


def method1_rss(source: str, url: str, start: datetime, end: datetime) -> list[dict]:
    check_crawl_cancelled()
    response = http_get(url, headers=METHOD1_HEADERS, timeout=(3, 6))
    response.raise_for_status()
    root = ET.fromstring(response.content)
    rows = []
    for node in root.findall(".//item") + root.findall(".//{*}entry"):
        check_crawl_cancelled()
        title = (node.findtext("title") or node.findtext("{*}title") or "").strip()
        link = (node.findtext("link") or "").strip()
        if not link:
            link_node = node.find("{*}link")
            link = (link_node.get("href", "") if link_node is not None else "").strip()
        description = node.findtext("description") or node.findtext("{*}summary") or ""
        published_text = node.findtext("pubDate") or node.findtext("{*}published") or node.findtext("{*}updated") or ""
        published = parse_method1_time(published_text)
        if title and link and method1_in_window(published, start, end):
            rows.append({"Title": title, "URL": link, "Source": source, "Published": published, "Summary": BeautifulSoup(description, "html.parser").get_text(" ", strip=True)})
    return rows


def method1_udn(start: datetime, end: datetime) -> list[dict]:
    rows = []
    for url in ("https://money.udn.com/rssfeed/news/1001/5591", "https://money.udn.com/rssfeed/news/1001/5590"):
        rows.extend(method1_rss("經濟日報", url, start, end))
    return rows[:METHOD1_POOL_LIMIT]


def method1_cnbc(start: datetime, end: datetime) -> list[dict]:
    rows = []
    for url in ("https://www.cnbc.com/id/100003114/device/rss/rss.html", "https://www.cnbc.com/id/10001147/device/rss/rss.html"):
        rows.extend(method1_rss("CNBC", url, start, end))
    return rows[:METHOD1_POOL_LIMIT]


def method1_cnyes(start: datetime, end: datetime) -> list[dict]:
    rows = []
    for category in ("us_stock", "wd_stock", "headline"):
        check_crawl_cancelled()
        url = f"https://api.cnyes.com/media/api/v1/newslist/category/{category}"
        for page in range(1, 7):
            check_crawl_cancelled()
            response = http_get(url, params={"limit": 50, "page": page}, headers=METHOD1_HEADERS, timeout=(3, 6))
            if response.status_code == 422:
                break
            response.raise_for_status()
            payload = response.json().get("items") or response.json().get("data", {}).get("items") or []
            records = payload.get("data", []) if isinstance(payload, dict) else payload
            if not records:
                break
            stop_category = False
            for record in records:
                check_crawl_cancelled()
                timestamp = record.get("publishAt") or record.get("createdAt") or record.get("created_at")
                try:
                    stamp = int(timestamp)
                    if stamp > 10_000_000_000:
                        stamp //= 1000
                    published = datetime.fromtimestamp(stamp, TAIPEI)
                except Exception:
                    published = None
                if published and published < start:
                    stop_category = True
                    continue
                if not method1_in_window(published, start, end):
                    continue
                raw_title = str(record.get("title") or record.get("name") or "")
                title = BeautifulSoup(raw_title, "html.parser").get_text(" ", strip=True) if "<" in raw_title and ">" in raw_title else raw_title.strip()
                news_id = record.get("newsId") or record.get("id")
                link = record.get("url") or (f"https://news.cnyes.com/news/id/{news_id}" if news_id else "")
                if title and link:
                    rows.append({"Title": title, "URL": link, "Source": "鉅亨網", "Published": published, "Summary": ""})
            if stop_category or len(rows) >= METHOD1_POOL_LIMIT:
                break
    return rows[:METHOD1_POOL_LIMIT]


def method1_tradingview(ticker: str, company: str, start: datetime, end: datetime,
                        session: requests.Session | None = None) -> list[dict]:
    primary = "NASDAQ" if ticker in METHOD1_EXCHANGES else "NYSE"
    client = session or requests.Session()
    client.trust_env = False
    response = None
    used_symbol = ""
    for exchange in (primary, "NYSE" if primary == "NASDAQ" else "NASDAQ"):
        check_crawl_cancelled()
        used_symbol = f"{exchange}:{ticker}"
        response = client.get(
            "https://news-mediator.tradingview.com/public/news-flow/v2/news",
            params=[("filter", "lang:zh-Hant"), ("filter", f"symbol:{used_symbol}"), ("client", "landing"), ("streaming", "false"), ("user_prostatus", "non_pro")],
            headers={**METHOD1_HEADERS, "Referer": f"https://tw.tradingview.com/symbols/{used_symbol.replace(':', '-')}/news/"},
            timeout=(3, 6),
        )
        if response.status_code != 422:
            response.raise_for_status()
            break
    if response is None or response.status_code == 422:
        return []
    rows = []
    for record in response.json().get("items", []):
        check_crawl_cancelled()
        try:
            published = datetime.fromtimestamp(int(record.get("published")), TAIPEI)
        except Exception:
            continue
        title = str(record.get("title") or "").strip()
        link = str(record.get("link") or "").strip()
        if link.startswith("/"):
            link = urljoin("https://tw.tradingview.com", link)
        if not link:
            link = f"https://tw.tradingview.com/symbols/{used_symbol.replace(':', '-')}/news/"
        if title and method1_in_window(published, start, end):
            rows.append({"Ticker": ticker, "Company": company, "Published Time": published.strftime("%Y-%m-%d %H:%M:%S"), "Title": title, "Source": "TradingView", "URL": link, "FinBERT": ""})
    return rows


def fetch_method1_news(companies: pd.DataFrame, start: datetime, end: datetime, progress=None) -> tuple[pd.DataFrame, list[str], dict[str, int]]:
    """方法一：三個新聞池，加 TradingView 逐公司補抓。"""
    errors: list[str] = []
    source_counts: dict[str, int] = {name: 0 for name in ("經濟日報", "鉅亨網", "CNBC", "TradingView")}
    rows: list[dict] = []
    pool_fetchers = (("經濟日報", method1_udn), ("鉅亨網", method1_cnyes), ("CNBC", method1_cnbc))
    pool_items: list[dict] = []
    pool_started = time.monotonic()
    # 三個新聞池互不相依，平行抓取可避免其中一個來源較慢時拖住其餘來源。
    with ThreadPoolExecutor(max_workers=len(pool_fetchers)) as pool_executor:
        pool_futures = {
            pool_executor.submit(fetcher, start, end): source
            for source, fetcher in pool_fetchers
        }
        for future in as_completed(pool_futures):
            check_crawl_cancelled()
            source = pool_futures[future]
            try:
                pool_items.extend(future.result())
            except CrawlCancelled:
                raise
            except Exception as exc:
                # 個別來源偶發逾時屬正常網路落差；保留內部紀錄，但不影響其他來源完成。
                errors.append(f"{source} 新聞池：{exc}")
            if isinstance(progress, BackgroundProgress):
                progress.job["method1_monitor"].update(
                    stage=f"新聞池處理中｜{source}", errors=list(errors), source_counts=dict(source_counts)
                )
    for item in pool_items:
        check_crawl_cancelled()
        matched = method1_match_company(item["Title"], item.get("Summary", ""), companies)
        if not matched:
            continue
        ticker, company = matched
        rows.append({"Ticker": ticker, "Company": company, "Published Time": item["Published"].strftime("%Y-%m-%d %H:%M:%S"), "Title": item["Title"], "Source": item["Source"], "URL": item["URL"], "FinBERT": ""})
        source_counts[item["Source"]] += 1

    def fetch_company(record: dict) -> tuple[list[dict], list[str]]:
        check_crawl_cancelled()
        ticker, company = record["Ticker"], record["Company"]
        company_rows, company_errors = [], []
        session = method1_http_session()
        for source, fetcher in (("TradingView", method1_tradingview),):
            check_crawl_cancelled()
            try:
                company_rows.extend(fetcher(ticker, company, start, end, session=session))
            except CrawlCancelled:
                raise
            except Exception as exc:
                company_errors.append(f"{source} {ticker}：{exc}")
        return company_rows, company_errors

    records = companies[["Ticker", "Company"]].to_dict("records")
    tradingview_started = time.monotonic()
    if isinstance(progress, BackgroundProgress):
        progress.job["method1_monitor"].setdefault("timings", {})["新聞池"] = time.monotonic() - pool_started
    # TradingView 是方法一最主要的耗時階段；每個執行緒會重用自己的連線。
    # 16 條並行可把 S&P 500 的等待時間壓低，同時避免過高並行造成限流。
    executor = ThreadPoolExecutor(max_workers=16)
    futures = {executor.submit(fetch_company, record): record for record in records}
    try:
        for completed, future in enumerate(as_completed(futures), 1):
            if isinstance(progress, BackgroundProgress) and progress.job["stop_event"].is_set():
                raise CrawlCancelled("使用者已停止抓取")
            company_rows, company_errors = future.result()
            rows.extend(company_rows)
            errors.extend(company_errors)
            for row in company_rows:
                source_counts[row["Source"]] += 1
            if progress is not None:
                progress.progress(completed / len(records), text=f"方法一逐公司補抓｜{completed}/{len(records)}")
            if isinstance(progress, BackgroundProgress):
                progress.job["method1_monitor"].update(
                    stage=f"逐公司補抓｜{completed}/{len(records)}", rows=len(rows),
                    errors=list(errors), source_counts=dict(source_counts),
                )
    finally:
        stopping = isinstance(progress, BackgroundProgress) and progress.job["stop_event"].is_set()
        if stopping:
            for future in futures:
                future.cancel()
        executor.shutdown(wait=not stopping, cancel_futures=stopping)
    if isinstance(progress, BackgroundProgress):
        progress.job["method1_monitor"].setdefault("timings", {})["TradingView 逐公司補抓"] = time.monotonic() - tradingview_started
    frame = normalize_news(pd.DataFrame(rows, columns=NEWS_COLUMNS))
    frame = filter_news_window(frame, start, end)
    if frame.empty:
        return frame, errors, source_counts
    frame["_url_key"] = frame["URL"].map(normalize_url)
    frame["_title_key"] = frame["Title"].map(normalized_title)
    frame = frame.drop_duplicates(["Company", "_url_key"]).drop_duplicates(["Company", "_title_key"]).drop(columns=["_url_key", "_title_key"])
    frame = frame.sort_values("Published Time", ascending=False).reset_index(drop=True)
    return frame[NEWS_COLUMNS], errors, source_counts


METHOD2_AMBIGUOUS_TICKERS = {
    "A", "AI", "ALL", "AN", "APP", "ARE", "AT", "BALL", "BE", "BY", "C", "CAT", "COST",
    "D", "DE", "DOW", "F", "FAST", "FOR", "HAS", "ICE", "IT", "KEY", "LOW", "NOW",
    "MAR", "NWS", "O", "ON", "OR", "SO", "T", "TECH", "TEL", "V", "WELL",
}


def company_is_information_source(title: str, ticker: str, company: str) -> bool:
    """排除公司只是研究發布者或評等機構、而非新聞事件當事人的標題。"""
    text = re.sub(r"\s+", " ", str(title or "")).strip()
    actor_terms = list(dict.fromkeys(method1_company_terms(ticker, company) + [str(company).strip(), str(ticker).strip()]))
    for term in (value for value in actor_terms if len(value) >= 2):
        escaped = re.escape(term)
        if re.search(
            rf"(?<![A-Za-z0-9]){escaped}(?:'s)?\s+(?:study|survey|report|index|analysts?)\b",
            text, re.I,
        ):
            return True
        if re.search(
            rf"(?<![A-Za-z0-9]){escaped}\s+(?:downgrades?|upgrades?|initiates?|rates?|cuts|raises|lowers)\b",
            text, re.I,
        ):
            return True
    return False


def method2_matches(title: str, companies: pd.DataFrame) -> list[tuple[str, str]]:
    """比對 Google News 標題；同一則新聞可對應多家公司。"""
    text = str(title).strip()
    lowered = text.lower()
    matches: list[tuple[str, str]] = []
    for row in companies.itertuples(index=False):
        ticker, company = str(row.Ticker).strip().upper(), str(row.Company).strip()
        structured_ticker = bool(ticker and re.search(
            rf"(?:\${re.escape(ticker)}\b|\({re.escape(ticker)}\)|(?:NASDAQ|NYSE|AMEX)\s*:\s*{re.escape(ticker)}\b)",
            text, re.I,
        ))
        plain_ticker = bool(
            ticker and len(ticker) >= 3 and ticker not in METHOD2_AMBIGUOUS_TICKERS
            and re.search(rf"(?<![A-Za-z0-9]){re.escape(ticker)}(?![A-Za-z0-9])", text)
        )
        company_hit = any(
            re.search(rf"(?<![A-Za-z0-9]){re.escape(term)}(?![A-Za-z0-9])", text, re.I)
            for term in method1_company_terms(ticker, company)
        )
        ticker_hit = structured_ticker or plain_ticker
        if not structured_ticker and (company_hit or plain_ticker) and company_is_information_source(text, ticker, company):
            company_hit = False
            ticker_hit = False
        if ticker_hit or company_hit:
            matches.append((ticker, company))
    return matches


def fetch_method2_raw(companies: pd.DataFrame, start: datetime, end: datetime, progress=None) -> tuple[pd.DataFrame, list[str], int]:
    """方法二第一階段：每五家公司一批、依日查詢，並以有限並行縮短網路等待。"""
    records = companies[["Ticker", "Company"]].drop_duplicates("Ticker").to_dict("records")
    batches = [records[index:index + 5] for index in range(0, len(records), 5)]
    rows: list[dict] = []
    errors: list[str] = []
    query_days = list(pd.date_range(start.astimezone(TAIPEI).date(), end.astimezone(TAIPEI).date(), freq="D").date)
    query_tasks = [(batch, query_day) for batch in batches for query_day in query_days]
    total_queries = max(len(query_tasks), 1)
    completed_queries = 0

    def fetch_query(batch: list[dict], query_day: date) -> tuple[list[dict], str | None]:
        check_crawl_cancelled()
        terms: list[str] = []
        batch_frame = pd.DataFrame(batch)
        for item in batch:
            ticker = str(item["Ticker"]).strip().upper()
            if len(ticker) >= 3 and ticker not in METHOD2_AMBIGUOUS_TICKERS:
                terms.append(f'"{ticker}"')
            terms.append(f'"{item["Company"]}"')
        after_day = query_day.isoformat()
        before_day = (query_day + timedelta(days=1)).isoformat()
        query = f'intitle:({" OR ".join(terms)}) after:{after_day} before:{before_day}'
        rss_url = "https://news.google.com/rss/search?" + urllib.parse.urlencode(
            {"q": query, "hl": "en-US", "gl": "US", "ceid": "US:en"}
        )
        query_rows: list[dict] = []
        try:
            response = http_get(rss_url, headers=METHOD1_HEADERS, timeout=(3, 6))
            response.raise_for_status()
            root = ET.fromstring(response.content)
            for item in root.findall(".//item"):
                check_crawl_cancelled()
                full_title = (item.findtext("title") or "").strip()
                title, source = full_title, "未知媒體"
                if " - " in full_title:
                    title, source = full_title.rsplit(" - ", 1)
                source = (item.findtext("source") or source).strip()
                published = parse_method1_time(item.findtext("pubDate"))
                if not method1_in_window(published, start, end):
                    continue
                for ticker, company in method2_matches(title, batch_frame):
                    query_rows.append({
                        "Ticker": ticker, "Company": company,
                        "Published Time": published.astimezone(TAIPEI).strftime("%Y-%m-%d %H:%M:%S"),
                        "Title": title, "Source": source,
                        "URL": (item.findtext("link") or "").strip(), "FinBERT": "",
                    })
            return query_rows, None
        except CrawlCancelled:
            raise
        except Exception as exc:
            tickers = ", ".join(str(item["Ticker"]) for item in batch)
            return [], f"Google News {after_day}｜{tickers}：{exc}"

    executor = ThreadPoolExecutor(max_workers=6)
    stopping = False
    try:
        futures = {executor.submit(fetch_query, batch, query_day): (batch, query_day) for batch, query_day in query_tasks}
        for future in as_completed(futures):
            check_crawl_cancelled()
            query_rows, error = future.result()
            rows.extend(query_rows)
            if error:
                errors.append(error)
            completed_queries += 1
            if progress is not None:
                progress.progress(completed_queries / total_queries * 0.55, text=f"方法二分日查詢｜{completed_queries}/{total_queries}")
            if isinstance(progress, BackgroundProgress):
                progress.job["method2_monitor"].update(
                    stage=f"Google News 分日查詢｜{completed_queries}/{total_queries}", raw_rows=len(rows),
                    batches=completed_queries, errors=list(errors),
                )
    except CrawlCancelled:
        stopping = True
        for future in futures:
            future.cancel()
        raise
    finally:
        executor.shutdown(wait=not stopping, cancel_futures=stopping)
    result = pd.DataFrame(rows, columns=NEWS_COLUMNS)
    result = filter_news_window(result, start, end)
    if not result.empty:
        result["_url"] = result["URL"].map(normalize_url)
        result["_title"] = result["Title"].map(normalized_title)
        result = result.drop_duplicates(["Ticker", "_url"]).drop_duplicates(["Ticker", "_title"]).drop(columns=["_url", "_title"])
        result = result.sort_values("Published Time", ascending=False).reset_index(drop=True)
    return result, errors, total_queries


_FINBERT_ENGINE = None
_FINBERT_ENGINE_LOCK = threading.Lock()


def _load_finbert_pytorch_model():
    try:
        import torch
        from transformers import AutoModelForSequenceClassification, AutoTokenizer
    except ImportError as exc:
        raise RuntimeError("方法二需要 transformers 與 torch，請先執行 pip install -r requirements.txt") from exc
    model_name = "ProsusAI/finbert"
    device = "cuda" if torch.cuda.is_available() else "cpu"
    if device == "cpu":
        # 本機為 8 個邏輯核心；保留少數核心給網站與作業系統，實測 6 執行緒最快。
        # 留一個邏輯核心給 Streamlit 介面與網路工作，其餘供本機推論使用。
        cpu_threads = min(8, max(1, int(os.cpu_count() or 4) - 1))
        torch.set_num_threads(cpu_threads)
    try:
        tokenizer = AutoTokenizer.from_pretrained(model_name, local_files_only=True)
        model = AutoModelForSequenceClassification.from_pretrained(model_name, local_files_only=True).to(device)
    except OSError:
        tokenizer = AutoTokenizer.from_pretrained(model_name)
        model = AutoModelForSequenceClassification.from_pretrained(model_name).to(device)
    model.eval()
    return torch, tokenizer, model, device


def load_finbert_model():
    """優先使用 Intel OpenVINO；不可用時自動退回原本的 PyTorch。"""
    global _FINBERT_ENGINE
    if _FINBERT_ENGINE is not None:
        return _FINBERT_ENGINE
    with _FINBERT_ENGINE_LOCK:
        if _FINBERT_ENGINE is not None:
            return _FINBERT_ENGINE
        try:
            import openvino as ov
            from transformers import AutoTokenizer

            model_file = FINBERT_ONNX_DIR / "model.onnx"
            if not model_file.is_file():
                from optimum.onnxruntime import ORTModelForSequenceClassification

                FINBERT_ONNX_DIR.mkdir(parents=True, exist_ok=True)
                model_name = "ProsusAI/finbert"
                try:
                    tokenizer = AutoTokenizer.from_pretrained(model_name, local_files_only=True)
                    export_model = ORTModelForSequenceClassification.from_pretrained(
                        model_name, export=True, local_files_only=True
                    )
                except OSError:
                    tokenizer = AutoTokenizer.from_pretrained(model_name)
                    export_model = ORTModelForSequenceClassification.from_pretrained(model_name, export=True)
                export_model.save_pretrained(FINBERT_ONNX_DIR)
                tokenizer.save_pretrained(FINBERT_ONNX_DIR)
            tokenizer = AutoTokenizer.from_pretrained(FINBERT_ONNX_DIR, local_files_only=True)
            core = ov.Core()
            ov_model = core.read_model(model_file)
            # OpenVINO 是目前主要引擎；使用多數核心但保留一個給網站介面。
            cpu_threads = min(8, max(1, int(os.cpu_count() or 4) - 1))
            compiled_model = core.compile_model(
                ov_model,
                "CPU",
                {"PERFORMANCE_HINT": "LATENCY", "INFERENCE_NUM_THREADS": str(cpu_threads)},
            )
            config = json.loads((FINBERT_ONNX_DIR / "config.json").read_text(encoding="utf-8"))
            labels = {int(key): str(value).lower() for key, value in config["id2label"].items()}
            _FINBERT_ENGINE = ("openvino", tokenizer, compiled_model, labels)
        except Exception:
            torch, tokenizer, model, device = _load_finbert_pytorch_model()
            labels = {int(key): str(value).lower() for key, value in model.config.id2label.items()}
            _FINBERT_ENGINE = ("pytorch", tokenizer, (torch, model, device), labels)
        return _FINBERT_ENGINE


def score_method2_finbert(frame: pd.DataFrame, progress=None) -> tuple[pd.DataFrame, pd.DataFrame]:
    """金融情緒分析：重複標題沿用本機快取，新標題才交給模型批次評分。"""
    scored = prepare_english_titles(frame, progress)
    english_translation_failures = int(scored.attrs.get("english_translation_failures", 0))
    if scored.empty:
        scored["FinBERT"] = pd.Series(dtype=float)
        scored.attrs["english_translation_failures"] = english_translation_failures
        return scored, scored.copy()
    titles = scored["Title"].fillna("").astype(str).tolist()
    title_keys = [re.sub(r"\s+", " ", title).strip().casefold() for title in titles]
    cached_scores: dict[str, float] = {}
    unique_keys = list(dict.fromkeys(key for key in title_keys if key))
    with sqlite3.connect(DB_PATH) as connection:
        for start_index in range(0, len(unique_keys), 800):
            chunk = unique_keys[start_index:start_index + 800]
            placeholders = ",".join("?" for _ in chunk)
            cached_scores.update({
                str(key): float(value)
                for key, value in connection.execute(
                    f"SELECT title_key,score FROM finbert_cache WHERE title_key IN ({placeholders})", chunk
                ).fetchall()
            })
    uncached_items: dict[str, tuple[int, str]] = {}
    for index, (key, title) in enumerate(zip(title_keys, titles)):
        if key and key not in cached_scores and key not in uncached_items:
            uncached_items[key] = (index, title)
    ordered_titles = sorted(uncached_items.items(), key=lambda item: len(item[1][1].split()))
    scores: list[float] = [cached_scores.get(key, 0.0) for key in title_keys]
    if not ordered_titles:
        scored["FinBERT"] = scores
        scored = scored.sort_values("FinBERT").reset_index(drop=True)
        scored.attrs["english_translation_failures"] = english_translation_failures
        filtered = scored[(scored["FinBERT"] >= -1) & (scored["FinBERT"] <= 0)].reset_index(drop=True)
        filtered.attrs["english_translation_failures"] = english_translation_failures
        return scored, filtered
    engine, tokenizer, model_runtime, labels = load_finbert_model()
    # 本機 CPU 用較大批次降低模型呼叫開銷；依標題長度分組可減少 padding 的無效運算。
    batch_size = 128 if engine == "openvino" else 32
    positive_index = next((key for key, value in labels.items() if value == "positive"), 0)
    negative_index = next((key for key, value in labels.items() if value == "negative"), 1)
    if engine == "openvino":
        for index in range(0, len(ordered_titles), batch_size):
            check_crawl_cancelled()
            batch_items = ordered_titles[index:index + batch_size]
            batch_keys = [item[0] for item in batch_items]
            original_indexes = [item[1][0] for item in batch_items]
            batch = [item[1][1] for item in batch_items]
            inputs = tokenizer(batch, padding=True, truncation=True, max_length=64, return_tensors="np")
            logits = model_runtime(dict(inputs))[model_runtime.output(0)]
            shifted = logits - logits.max(axis=1, keepdims=True)
            probabilities = np.exp(shifted)
            probabilities /= probabilities.sum(axis=1, keepdims=True)
            for key, original_index, values in zip(batch_keys, original_indexes, probabilities):
                score = round(float(values[positive_index] - values[negative_index]), 3)
                cached_scores[key] = score
                scores[original_index] = score
            if progress is not None:
                done = min(index + batch_size, len(ordered_titles))
                progress.progress(0.55 + 0.45 * done / len(ordered_titles), text=f"金融情緒分析 (FinBERT)｜新標題 {done}/{len(ordered_titles)}；快取 {len(titles) - len(ordered_titles):,}")
    else:
        torch, model, device = model_runtime
        with torch.inference_mode():
            for index in range(0, len(ordered_titles), batch_size):
                check_crawl_cancelled()
                batch_items = ordered_titles[index:index + batch_size]
                batch_keys = [item[0] for item in batch_items]
                original_indexes = [item[1][0] for item in batch_items]
                batch = [item[1][1] for item in batch_items]
                inputs = tokenizer(batch, padding=True, truncation=True, max_length=64, return_tensors="pt").to(device)
                probabilities = torch.nn.functional.softmax(model(**inputs).logits, dim=-1).cpu().numpy()
                for key, original_index, values in zip(batch_keys, original_indexes, probabilities):
                    score = round(float(values[positive_index] - values[negative_index]), 3)
                    cached_scores[key] = score
                    scores[original_index] = score
                if progress is not None:
                    done = min(index + batch_size, len(ordered_titles))
                    progress.progress(0.55 + 0.45 * done / len(ordered_titles), text=f"金融情緒分析 (FinBERT)｜新標題 {done}/{len(ordered_titles)}；快取 {len(titles) - len(ordered_titles):,}")
    now_text = datetime.now(TAIPEI).isoformat(timespec="seconds")
    with sqlite3.connect(DB_PATH) as connection:
        connection.executemany(
            "INSERT OR REPLACE INTO finbert_cache(title_key,title,score,updated_at) VALUES(?,?,?,?)",
            [(key, title, cached_scores[key], now_text) for key, (_, title) in uncached_items.items() if key in cached_scores],
        )
    scores = [cached_scores.get(key, score) for key, score in zip(title_keys, scores)]
    scored["FinBERT"] = scores
    scored = scored.sort_values("FinBERT").reset_index(drop=True)
    scored.attrs["english_translation_failures"] = english_translation_failures
    filtered = scored[(scored["FinBERT"] >= -1) & (scored["FinBERT"] <= 0)].reset_index(drop=True)
    filtered.attrs["english_translation_failures"] = english_translation_failures
    return scored, filtered


def load_company_upload(upload, sheet_name=0) -> pd.DataFrame:
    frame = pd.read_excel(io.BytesIO(upload.getvalue()), sheet_name=sheet_name, dtype=str)
    columns = {str(column).strip().lower(): column for column in frame.columns}
    ticker_col = next((columns[key] for key in ("ticker", "symbol", "股票代號") if key in columns), None)
    company_col = next((columns[key] for key in ("company", "name", "股票名稱") if key in columns), None)
    if not ticker_col or not company_col:
        raise ValueError("名單需要 Ticker/Symbol 與 Company/Name 欄位")
    result = frame[[ticker_col, company_col]].rename(columns={ticker_col: "Ticker", company_col: "Company"}).dropna()
    result["Ticker"] = result["Ticker"].astype(str).str.strip().str.upper()
    result["Company"] = result["Company"].astype(str).str.strip()
    return result[(result["Ticker"] != "") & (result["Company"] != "")].drop_duplicates("Ticker")


def load_saved_companies(sheet_name: str) -> pd.DataFrame:
    if not COMPANY_INDEX_PATH.is_file():
        raise ValueError("尚未建立美股指數成分股名單，請先上傳新版 Excel")
    payload = io.BytesIO(COMPANY_INDEX_PATH.read_bytes())
    return load_company_upload(payload, sheet_name=sheet_name)


@st.cache_data(ttl=3600)
def stock_prices(ticker: str, start_text: str, end_text: str) -> pd.DataFrame:
    start = int(datetime.fromisoformat(start_text).replace(tzinfo=timezone.utc).timestamp())
    end = int((datetime.fromisoformat(end_text) + timedelta(days=1)).replace(tzinfo=timezone.utc).timestamp())
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{urllib.parse.quote(ticker)}?period1={start}&period2={end}&interval=1d&events=history"
    headers = {"User-Agent": "Mozilla/5.0"}
    response = http_get(url, timeout=15, headers=headers)
    response.raise_for_status()
    chart = response.json().get("chart", {})
    results = chart.get("result") or []
    if not results:
        message = (chart.get("error") or {}).get("description") or "Yahoo Finance 未回傳股價資料"
        raise ValueError(message)
    result = results[0]
    timestamps = result.get("timestamp", [])
    quotes = result["indicators"]["quote"][0]
    rows = []
    for index, stamp in enumerate(timestamps):
        close = quotes["close"][index]
        if close is not None:
            rows.append({"Date": datetime.fromtimestamp(stamp, timezone.utc).date(), "Close": float(close)})
    if not rows:
        raise ValueError("選定期間沒有可用的收盤價資料")
    return pd.DataFrame(rows)


def output_files() -> list[Path]:
    return sorted(OUTPUT_DIR.glob("美股_*_負面新聞爬蟲.xlsx"), key=lambda path: (re.search(r"20\d{6}", path.name).group(0), path.stat().st_mtime), reverse=True)


def seed_finbert_cache() -> None:
    """首次啟用快取時，直接沿用既有完整整合檔的評分，避免下一次又全部重算。"""
    with sqlite3.connect(DB_PATH) as connection:
        if int(connection.execute("SELECT COUNT(*) FROM finbert_cache").fetchone()[0]) > 0:
            return
    candidates = sorted(OUTPUT_DIR.glob("MergeNews_20*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
    for candidate in candidates:
        try:
            historical = pd.read_excel(candidate, usecols=lambda column: column in ("Title", "FinBERT"))
            historical["FinBERT"] = pd.to_numeric(historical.get("FinBERT"), errors="coerce")
            historical = historical.dropna(subset=["Title", "FinBERT"])
            if historical.empty:
                continue
            now_text = datetime.now(TAIPEI).isoformat(timespec="seconds")
            rows = []
            for title, score in historical[["Title", "FinBERT"]].itertuples(index=False, name=None):
                key = re.sub(r"\s+", " ", str(title)).strip().casefold()
                if key:
                    rows.append((key, str(title), float(score), now_text))
            with sqlite3.connect(DB_PATH) as connection:
                connection.executemany(
                    "INSERT OR IGNORE INTO finbert_cache(title_key,title,score,updated_at) VALUES(?,?,?,?)", rows
                )
            return
        except Exception:
            continue


def seed_translation_cache() -> None:
    """沿用既有輸出中的中英文標題，避免升級後第一次又翻譯全部歷史新聞。"""
    candidates = output_files()[:3]
    seed_signature = "|".join(
        f"{candidate.name}:{candidate.stat().st_mtime_ns}:{candidate.stat().st_size}"
        for candidate in candidates
        if candidate.is_file()
    )
    with sqlite3.connect(DB_PATH) as connection:
        seeded = connection.execute(
            "SELECT value FROM settings WHERE key='translation_cache_seeded'"
        ).fetchone()
    if seeded and seeded[0] == seed_signature:
        return
    zh_items: list[tuple[str, str]] = []
    en_items: list[tuple[str, str]] = []
    for candidate in candidates:
        try:
            sheets = pd.read_excel(
                candidate,
                sheet_name=None,
                usecols=lambda column: column in ("Title", "Title_ZH"),
            )
            for frame in sheets.values():
                if not {"Title", "Title_ZH"}.issubset(frame.columns):
                    continue
                for title, title_zh in frame[["Title", "Title_ZH"]].fillna("").itertuples(index=False, name=None):
                    title = str(title).strip()
                    title_zh = str(title_zh).strip()
                    if not title or not title_zh:
                        continue
                    zh_items.append((title, title_zh))
                    if re.search(r"[\u3400-\u9fff]", title_zh) and not re.search(r"[\u3400-\u9fff]", title):
                        en_items.append((title_zh, title))
        except Exception:
            continue
    save_translation_cache(zh_items, "zh-TW")
    save_translation_cache(en_items, "en")
    if zh_items or en_items:
        with sqlite3.connect(DB_PATH) as connection:
            connection.execute(
                "INSERT OR REPLACE INTO settings(key,value) VALUES('translation_cache_seeded',?)",
                (seed_signature,),
            )


def load_negative_files() -> pd.DataFrame:
    frames = []
    for path in output_files():
        frame = pd.read_excel(path, dtype=str)
        levels = pd.to_numeric(frame.get("Level"), errors="coerce")
        frame = frame[levels.between(3, 5, inclusive="both")].copy()
        match = re.search(r"20\d{6}", path.name)
        frame["資料日期"] = match.group(0) if match else ""
        frames.append(frame)
    return pd.concat(frames, ignore_index=True, sort=False) if frames else pd.DataFrame()


def assign_event_batches(frame: pd.DataFrame) -> pd.DataFrame:
    """同公司、同事件種類超過設定的間隔天數後，下一則即建立新事件。"""
    if frame.empty:
        return frame.copy()
    with sqlite3.connect(DB_PATH) as connection:
        saved_gap_days = connection.execute(
            "SELECT value FROM settings WHERE key='event_group_gap_days'"
        ).fetchone()
    try:
        gap_days = max(1, min(3650, int(saved_gap_days[0]))) if saved_gap_days else 14
    except (TypeError, ValueError):
        gap_days = 14
    data = frame.copy()
    data["公司鍵"] = data["Ticker"].fillna("").where(data["Ticker"].fillna("") != "", data["Company"].fillna("").str.lower())
    data["事件時間值"] = pd.to_datetime(data["Published Time"], errors="coerce")
    fallback_dates = pd.to_datetime(data.get("資料日期", ""), format="%Y%m%d", errors="coerce")
    data["事件時間值"] = data["事件時間值"].fillna(fallback_dates)
    data["事件日期"] = data["事件時間值"].dt.strftime("%Y-%m-%d")
    data["Event_ID"] = ""
    for (company_key, event_code), indexes in data.groupby(["公司鍵", "Event_Code"], dropna=False).groups.items():
        ordered_indexes = sorted(indexes, key=lambda index: (data.at[index, "事件時間值"] if pd.notna(data.at[index, "事件時間值"]) else pd.Timestamp.max, index))
        safe_company = re.sub(r"[^A-Za-z0-9]+", "", str(company_key).upper()) or "COMPANY"
        safe_event = re.sub(r"[^A-Za-z0-9_]+", "_", str(event_code).upper()) or "EVENT"
        event_number = 0
        previous_time = pd.NaT
        for index in ordered_indexes:
            event_time = data.at[index, "事件時間值"]
            if event_number == 0:
                event_number = 1
            elif pd.isna(event_time) or pd.isna(previous_time) or (event_time.normalize() - previous_time.normalize()).days > gap_days:
                event_number += 1
            data.at[index, "Event_ID"] = f"{safe_company}-{safe_event}-{event_number:03d}"
            if pd.notna(event_time):
                previous_time = event_time
    return data


def validate_negative_frame(frame: pd.DataFrame) -> pd.DataFrame:
    required = {"Published Time", "Ticker", "Company", "Title", "Level", "Action", "Event_Code", "Event_type"}
    missing = sorted(required - set(frame.columns))
    if missing:
        raise ValueError(f"缺少必要欄位：{', '.join(missing)}")
    result = frame.copy()
    result["Level"] = pd.to_numeric(result["Level"], errors="coerce")
    if result["Level"].isna().any():
        raise ValueError("Level 欄含有無法辨識的值")
    result["Level"] = result["Level"].astype(int)
    return result[result["Level"].between(3, 5, inclusive="both")].reset_index(drop=True)


def save_corrected_negative_workbook(target: Path, negative_frame: pd.DataFrame,
                                     other_sheets: dict[str, pd.DataFrame] | None = None) -> None:
    sheets = {"負面新聞": validate_negative_frame(negative_frame)}
    for sheet_name, sheet_frame in (other_sheets or {}).items():
        if sheet_name != "負面新聞":
            sheets[sheet_name] = sheet_frame
    temporary = target.with_name(f".{target.stem}_editing.xlsx")
    write_excel(temporary, sheets)
    temporary.replace(target)


def load_negative_workbook(path_or_bytes) -> tuple[pd.DataFrame, dict[str, pd.DataFrame]]:
    excel = pd.ExcelFile(path_or_bytes)
    negative_sheet = "負面新聞" if "負面新聞" in excel.sheet_names else excel.sheet_names[0]
    negative = validate_negative_frame(pd.read_excel(excel, sheet_name=negative_sheet))
    other_sheets = {
        sheet_name: pd.read_excel(excel, sheet_name=sheet_name)
        for sheet_name in excel.sheet_names if sheet_name != negative_sheet
    }
    return negative, other_sheets


BUSINESS_LINE_ALL = "全部（加總）"
BUSINESS_LINE_ORDER = ["財管", "經紀", "自營"]
BUSINESS_LINE_ALIASES = {
    "財管": "財管", "財富管理": "財管", "wm": "財管", "wealth": "財管", "wealth management": "財管",
    "經紀": "經紀", "證券經紀": "經紀", "brokerage": "經紀", "broker": "經紀",
    "自營": "自營", "自營交易": "自營", "proprietary": "自營", "prop": "自營",
}


def parse_exposure_frame(frame: pd.DataFrame) -> pd.DataFrame:
    """接受長表（業務別＋曝險金額）及寬表（財管／經紀／自營各一欄）。"""
    if frame.empty:
        return pd.DataFrame(columns=["Ticker", "Company", "業務別", "曝險金額"])
    columns = {str(column).strip().lower(): column for column in frame.columns}
    ticker_col = next((columns[key] for key in ("ticker", "symbol", "股票代號", "代號") if key in columns), None)
    company_col = next((columns[key] for key in ("company", "公司", "公司名稱") if key in columns), None)
    if ticker_col is None:
        raise ValueError("部位清單需要 Ticker、Symbol、股票代號或代號欄位")
    exposure_col = next(
        (columns[key] for key in ("exposure", "amount", "曝險金額", "曝險", "部位金額", "部位") if key in columns),
        None,
    )
    business_col = next(
        (columns[key] for key in ("業務別", "業務", "部門", "business", "businessline", "business_line", "desk") if key in columns),
        None,
    )
    company_values = frame[company_col].fillna("").astype(str).str.strip() if company_col else pd.Series("", index=frame.index)
    if exposure_col is not None:
        result = pd.DataFrame({
            "Ticker": frame[ticker_col], "Company": company_values,
            "業務別": frame[business_col] if business_col else BUSINESS_LINE_ALL,
            "曝險金額": frame[exposure_col],
        })
    else:
        wide_columns = []
        for raw_name, original in columns.items():
            normalized = BUSINESS_LINE_ALIASES.get(raw_name)
            if normalized:
                wide_columns.append((original, normalized))
        if not wide_columns:
            raise ValueError("部位清單需要曝險金額欄，或財管／經紀／自營任一金額欄")
        parts = []
        for original, business_name in wide_columns:
            parts.append(pd.DataFrame({
                "Ticker": frame[ticker_col], "Company": company_values,
                "業務別": business_name, "曝險金額": frame[original],
            }))
        result = pd.concat(parts, ignore_index=True)
    result["Ticker"] = result["Ticker"].fillna("").astype(str).str.strip().str.upper().str.replace(".", "-", regex=False)
    result["Company"] = result["Company"].fillna("").astype(str).str.strip()
    raw_business = result["業務別"].fillna("").astype(str).str.strip()
    result["業務別"] = raw_business.map(lambda value: BUSINESS_LINE_ALIASES.get(value.lower(), value or "未分類"))
    result["曝險金額"] = pd.to_numeric(
        result["曝險金額"].astype(str).str.replace(",", "", regex=False).str.replace("$", "", regex=False),
        errors="coerce",
    ).fillna(0.0)
    result = result[(result["Ticker"] != "") & (result["曝險金額"] != 0)].copy()
    return result.groupby(["Ticker", "Company", "業務別"], as_index=False)["曝險金額"].sum()


@st.cache_data(ttl=60)
def load_exposure_file(path_text: str, modified_ns: int) -> pd.DataFrame:
    del modified_ns
    return parse_exposure_frame(pd.read_excel(path_text, dtype=str))


def exposure_business_options(exposure: pd.DataFrame) -> list[str]:
    if exposure.empty:
        return [BUSINESS_LINE_ALL]
    present = exposure["業務別"].dropna().astype(str).unique().tolist()
    ordered = [name for name in BUSINESS_LINE_ORDER if name in present]
    ordered.extend(sorted(name for name in present if name not in ordered and name != BUSINESS_LINE_ALL))
    return [BUSINESS_LINE_ALL] + ordered


def exposure_map_for_business(exposure: pd.DataFrame, business_name: str) -> dict[str, float]:
    if exposure.empty:
        return {}
    selected = exposure if business_name == BUSINESS_LINE_ALL else exposure[exposure["業務別"] == business_name]
    return selected.groupby("Ticker")["曝險金額"].sum().to_dict()


def parse_sec13f_workbook(path_or_bytes) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """讀取舊 SEC13F 專案的成果檔；同時相容其部分中文欄名已亂碼的範例檔。"""
    excel = pd.ExcelFile(path_or_bytes)
    required = {"Ticker Summary", "13F Detail"}
    missing = sorted(required - set(excel.sheet_names))
    if missing:
        raise ValueError(f"13F 檔案缺少工作表：{'、'.join(missing)}")
    raw_summary = pd.read_excel(excel, sheet_name="Ticker Summary")
    detail = pd.read_excel(excel, sheet_name="13F Detail")
    run_log = pd.read_excel(excel, sheet_name="Run Log") if "Run Log" in excel.sheet_names else pd.DataFrame()
    if raw_summary.shape[1] < 20:
        raise ValueError("Ticker Summary 欄位不足，請使用 SEC13F 專案產出的完整檔案")
    # 舊專案欄位順序固定；使用位置重新命名可避開舊檔中文編碼損壞。
    summary = raw_summary.iloc[:, :20].copy()
    summary.columns = [
        "Quarter", "Quarter End", "Ticker", "Company", "追蹤Universe", "本季納入機構數",
        "本季持有機構數", "前季持有機構數", "Current Shares", "Previous Shares", "機構持股QoQ",
        "增持家數", "減持家數", "新建倉", "清倉", "持股不變家數", "淨增持家數",
        "持股量方向", "機構持股方向", "13F訊號",
    ]
    summary["Ticker"] = summary["Ticker"].fillna("").astype(str).str.strip().str.upper().str.replace(".", "-", regex=False)
    summary["Company"] = summary["Company"].fillna("").astype(str).str.strip()
    numeric_columns = [
        "追蹤Universe", "本季納入機構數", "本季持有機構數", "前季持有機構數",
        "Current Shares", "Previous Shares", "機構持股QoQ", "增持家數", "減持家數",
        "新建倉", "清倉", "持股不變家數", "淨增持家數",
    ]
    for column in numeric_columns:
        summary[column] = pd.to_numeric(summary[column], errors="coerce")
    summary["持股量方向"] = np.select(
        [summary["機構持股QoQ"] > 0, summary["機構持股QoQ"] < 0],
        ["持股增加", "持股下降"], default="持股不變",
    )
    summary["機構持股方向"] = np.select(
        [summary["增持家數"] > summary["減持家數"], summary["增持家數"] < summary["減持家數"]],
        ["增持家數較多", "減持家數較多"], default="增減持家數相同",
    )
    summary["13F顯示"] = summary.apply(
        lambda row: (
            f"{'↑' if row['機構持股QoQ'] > 0 else '↓' if row['機構持股QoQ'] < 0 else '→'} "
            f"{row['機構持股QoQ']:.1%}｜{int(row['增持家數'] or 0)}增/{int(row['減持家數'] or 0)}減"
        ) if pd.notna(row["機構持股QoQ"]) else "— 尚無前季比較",
        axis=1,
    )
    if "Ticker" not in detail.columns or "Institution" not in detail.columns:
        raise ValueError("13F Detail 缺少 Ticker 或 Institution 欄位")
    detail["Ticker"] = detail["Ticker"].fillna("").astype(str).str.strip().str.upper().str.replace(".", "-", regex=False)
    return summary[summary["Ticker"] != ""].reset_index(drop=True), detail, run_log


@st.cache_data(ttl=60)
def load_sec13f_file(path_text: str, modified_ns: int) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    del modified_ns
    return parse_sec13f_workbook(path_text)


def latest_sec13f_path() -> Path | None:
    if SEC13F_PATH.is_file():
        return SEC13F_PATH
    candidates = sorted(SEC13F_DIR.glob("*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True) if SEC13F_DIR.is_dir() else []
    return candidates[0] if candidates else None


def sec13f_lookup() -> tuple[dict[str, str], str]:
    path = latest_sec13f_path()
    if path is None:
        return {}, ""
    try:
        summary, _, _ = load_sec13f_file(str(path), path.stat().st_mtime_ns)
        quarter = str(summary["Quarter"].dropna().iloc[0]) if summary["Quarter"].notna().any() else ""
        return summary.set_index("Ticker")["13F顯示"].to_dict(), quarter
    except Exception:
        return {}, ""


def prepare_position_news(frame: pd.DataFrame) -> pd.DataFrame:
    """以發布日統計並跨檔去除同一篇新聞，避免重疊抓取區間重複計數。"""
    if frame.empty:
        return frame.copy()
    result = frame.copy()
    result["Ticker"] = result.get("Ticker", "").fillna("").astype(str).str.upper().str.strip()
    result["Company"] = result.get("Company", "").fillna("").astype(str).str.strip()
    result["Level"] = pd.to_numeric(result.get("Level"), errors="coerce")
    result["發布日期"] = pd.to_datetime(result.get("Published Time"), errors="coerce").dt.normalize()
    fallback = pd.to_datetime(result.get("資料日期", ""), format="%Y%m%d", errors="coerce")
    result["發布日期"] = result["發布日期"].fillna(fallback)
    url_key = result.get("URL", "").fillna("").astype(str).str.strip().str.lower()
    title_key = result.get("Title", "").fillna("").astype(str).str.lower().str.replace(r"\s+", " ", regex=True).str.strip()
    # 同一篇新聞可能同時涉及多家公司；去重必須保留公司維度，不能只用網址把其他公司一起刪掉。
    result["新聞鍵"] = result["Ticker"] + "|" + np.where(url_key.ne(""), url_key, title_key)
    return result.dropna(subset=["發布日期"]).drop_duplicates("新聞鍵", keep="last").reset_index(drop=True)


def render_sec13f_panel(negative_news: pd.DataFrame, exposure: pd.DataFrame) -> None:
    st.divider()
    st.markdown("### 13F 機構持股動向")
    st.markdown(
        "<div class='method-card'><div class='method-title'>13F｜機構季度持股動向</div>"
        "<div class='method-desc'>比較追蹤機構本季與前季持股，並與負面新聞及曝險部位交叉查看。"
        "13F 並非即時持股，只作輔助判讀，不會改變負面新聞 Level。</div></div>",
        unsafe_allow_html=True,
    )
    path = latest_sec13f_path()
    if path is None:
        st.info("目前尚無 13F 成果檔。可到第六頁「設定與說明」上傳 SEC13F 專案產出的完整 Excel。")
        return
    try:
        summary, detail, run_log = load_sec13f_file(str(path), path.stat().st_mtime_ns)
    except Exception as exc:
        st.warning(f"目前 13F 檔案無法讀取：{exc}")
        return
    if summary.empty:
        st.info("目前 13F 成果檔沒有股票資料。")
        return
    quarter = str(summary["Quarter"].dropna().iloc[0]) if summary["Quarter"].notna().any() else "未知"
    quarter_end = pd.to_datetime(summary["Quarter End"], errors="coerce").max()
    filing_dates = pd.to_datetime(detail.get("Filing Date"), errors="coerce").dropna()
    filing_range = f"{filing_dates.min():%Y-%m-%d} ～ {filing_dates.max():%Y-%m-%d}" if not filing_dates.empty else "未知"
    successful_managers = int((run_log.get("Status", pd.Series(dtype=str)).astype(str).str.upper() == "OK").sum()) if not run_log.empty else int(summary["本季納入機構數"].max() or 0)
    included_managers = int(summary["本季納入機構數"].max() or 0)
    st.caption(
        f"資料季度 {quarter}｜截止日 {quarter_end.strftime('%Y-%m-%d') if pd.notna(quarter_end) else '未知'}｜"
        f"涵蓋 {summary['Ticker'].nunique():,} 檔股票｜正常完成 {successful_managers:,} 家機構"
        + (f"／納入彙總 {included_managers:,} 家" if included_managers else "")
        + f"｜SEC 申報日期 {filing_range}｜網站套用 {datetime.fromtimestamp(path.stat().st_mtime):%Y-%m-%d %H:%M}"
    )
    st.info(
        f"本期 13F 成果檔涵蓋 {summary['Ticker'].nunique():,} 檔股票；未出現在本頁不代表機構沒有持股，"
        "也可能只是未列入本期股票範圍。"
    )

    news = negative_news.copy()
    if not news.empty:
        news["Ticker"] = news.get("Ticker", "").fillna("").astype(str).str.upper().str.strip()
        news["Level"] = pd.to_numeric(news.get("Level"), errors="coerce")
        news_summary = news.groupby("Ticker", as_index=False).agg(最高新聞Level=("Level", "max"), 負面新聞則數=("Ticker", "size"))
        summary = summary.merge(news_summary, on="Ticker", how="left")
    else:
        summary["最高新聞Level"] = np.nan
        summary["負面新聞則數"] = 0
    summary["最高新聞Level"] = pd.to_numeric(summary["最高新聞Level"], errors="coerce")
    summary["負面新聞則數"] = pd.to_numeric(summary["負面新聞則數"], errors="coerce").fillna(0).astype(int)
    exposure_map = exposure.groupby("Ticker")["曝險金額"].sum().to_dict() if not exposure.empty else {}
    summary["曝險金額"] = summary["Ticker"].map(exposure_map).fillna(0.0)
    summary["輔助提示"] = np.select(
        [
            summary["最高新聞Level"].ge(4) & summary["機構持股QoQ"].lt(0),
            summary["最高新聞Level"].ge(4) & summary["機構持股QoQ"].gt(0),
            summary["最高新聞Level"].isna() & summary["機構持股QoQ"].lt(0),
        ],
        ["重大新聞＋機構持股下降", "重大新聞＋機構持股增加", "無重大新聞＋機構持股下降"],
        default="—",
    )
    summary["方向判讀"] = np.select(
        [
            summary["機構持股QoQ"].gt(0) & summary["減持家數"].gt(summary["增持家數"]),
            summary["機構持股QoQ"].lt(0) & summary["增持家數"].gt(summary["減持家數"]),
        ],
        ["⚠️ 股數增加、減持家數較多", "⚠️ 股數下降、增持家數較多"],
        default="—",
    )
    st.markdown("#### 優先注意摘要")
    st.caption("先看機構合計持股是否下降，再確認是否同時出現重大新聞或現有曝險部位。")
    action_metrics = st.columns(4)
    action_metrics[0].metric("機構持股下降", f"{int(summary['機構持股QoQ'].lt(0).sum()):,} 檔")
    action_metrics[1].metric("減持家數較多", f"{int((summary['減持家數'] > summary['增持家數']).sum()):,} 檔")
    action_metrics[2].metric("重大新聞＋持股下降", f"{int((summary['最高新聞Level'].ge(4) & summary['機構持股QoQ'].lt(0)).sum()):,} 檔")
    action_metrics[3].metric("曝險部位＋持股下降", f"{int((summary['曝險金額'].ne(0) & summary['機構持股QoQ'].lt(0)).sum()):,} 檔")
    st.caption(
        "「機構持股下降」看所有追蹤機構的合計股數；「減持家數較多」比較增持與減持的機構家數。"
        "兩者可能不同，例如少數大型機構增持的股數可能高於多家小型機構的減持總量。"
    )

    filters = st.columns([2, 1.5])
    search = filters[0].text_input("搜尋 Ticker 或公司", key="sec13f_search")
    priority_view = filters[1].selectbox(
        "優先查看",
        ["全部股票", "重大新聞＋機構減持", "曝險部位＋機構減持", "機構持股增加", "機構持股下降", "持股方向分歧"],
        key="sec13f_priority_view",
    )
    shown = summary.copy()
    if search:
        shown = shown[shown[["Ticker", "Company"]].fillna("").astype(str).apply(lambda column: column.str.contains(search, case=False, regex=False)).any(axis=1)]
    if priority_view == "重大新聞＋機構減持":
        shown = shown[shown["最高新聞Level"].ge(4) & shown["機構持股QoQ"].lt(0)]
    elif priority_view == "曝險部位＋機構減持":
        shown = shown[shown["曝險金額"].ne(0) & shown["機構持股QoQ"].lt(0)]
    elif priority_view == "機構持股增加":
        shown = shown[shown["機構持股QoQ"].gt(0)]
    elif priority_view == "機構持股下降":
        shown = shown[shown["機構持股QoQ"].lt(0)]
    elif priority_view == "持股方向分歧":
        shown = shown[shown["方向判讀"].ne("—")]
    shown = shown.sort_values(["最高新聞Level", "負面新聞則數", "機構持股QoQ"], ascending=[False, False, True], na_position="last")

    chart_heading, chart_control = st.columns([3, 1])
    chart_heading.markdown("#### 機構持股季增減排名")
    chart_scope = chart_control.selectbox("圖表範圍", ["變化最大 20 檔", "全部股票"], key="sec13f_chart_scope")
    st.caption("藍色向右代表機構合計持股增加，紅色向左代表下降；依變動幅度排列，游標移到橫條可查看機構家數與建倉狀況。")
    ranked = shown.dropna(subset=["機構持股QoQ"]).copy()
    ranked["變動幅度"] = ranked["機構持股QoQ"].abs()
    if chart_scope == "變化最大 20 檔":
        ranked = ranked.nlargest(20, "變動幅度")
    ranked = ranked.sort_values("機構持股QoQ")
    ranked["方向"] = np.where(ranked["機構持股QoQ"] >= 0, "持股增加", "持股下降")
    ranked["變動標籤"] = ranked["機構持股QoQ"].map(lambda value: f"{value:+.1%}")
    holding_counts = ranked["本季持有機構數"].fillna(0).astype(float)
    holding_min = float(holding_counts.min()) if not holding_counts.empty else 0.0
    holding_max = float(holding_counts.max()) if not holding_counts.empty else 0.0
    holding_span = max(holding_max - holding_min, 1.0)
    coverage_strength = ((holding_counts - holding_min) / holding_span).clip(0, 1)

    def soft_direction_color(direction_value: str, strength: float) -> str:
        # 淺色仍需清楚可辨；深淺代表持有機構覆蓋數，不重複表達橫條的季增減幅度。
        light = np.array([191, 219, 254]) if direction_value == "持股增加" else np.array([254, 202, 202])
        dark = np.array([59, 130, 246]) if direction_value == "持股增加" else np.array([220, 88, 88])
        ratio = 0.22 + 0.58 * float(strength)
        rgb = np.rint(light * (1 - ratio) + dark * ratio).astype(int)
        return f"rgb({rgb[0]},{rgb[1]},{rgb[2]})"

    ranked["圖表顏色"] = [
        soft_direction_color(direction_value, strength)
        for direction_value, strength in zip(ranked["方向"], coverage_strength)
    ]
    if ranked.empty:
        st.info("目前篩選條件沒有可顯示的股票。")
    else:
        figure = px.bar(
            ranked, x="機構持股QoQ", y="Ticker", orientation="h", text="變動標籤",
            hover_name="Company",
            hover_data={"機構持股QoQ": ":.1%", "增持家數": True, "減持家數": True, "新建倉": True,
                        "清倉": True, "本季持有機構數": True, "變動幅度": False, "方向": True,
                        "圖表顏色": False},
        )
        figure.update_traces(
            textposition="outside", cliponaxis=False, marker_color=ranked["圖表顏色"].tolist(),
            marker_line_color="white", marker_line_width=.6,
        )
        figure.add_vline(x=0, line_color="#64748B", line_width=1.2)
        figure.update_layout(
            height=max(410, 30 * len(ranked) + 95), margin={"t": 15, "b": 20, "l": 20, "r": 65},
            plot_bgcolor="white", paper_bgcolor="white", xaxis_title="機構合計持股季增減", yaxis_title=None,
            showlegend=False,
        )
        figure.update_xaxes(showgrid=True, gridcolor="#E2E8F0", tickformat=".0%", zeroline=False)
        figure.update_yaxes(showgrid=False, categoryorder="array", categoryarray=ranked["Ticker"].tolist())
        st.plotly_chart(figure, use_container_width=True, config={"displaylogo": False})
        st.caption("色彩說明：藍色為持股增加、紅色為持股下降；同一色系越深，代表本季持有該股票的追蹤機構越多。")

    st.markdown("#### 股票持股與風險對照")
    st.caption(
        "把 13F 持股方向、機構家數、負面新聞與曝險部位放在同一列。"
        "出現「方向分歧」時，代表合計股數與增減持機構家數方向不同，需查看下方機構明細確認。"
    )
    table_columns = ["Ticker", "Company", "13F顯示", "方向判讀", "輔助提示", "最高新聞Level", "曝險金額", "增持家數", "減持家數", "新建倉", "清倉", "負面新聞則數"]
    st.dataframe(
        shown[table_columns], use_container_width=True, hide_index=True,
        column_config={
            "Company": st.column_config.TextColumn("公司"), "13F顯示": st.column_config.TextColumn("13F 機構動向", width="medium"),
            "方向判讀": st.column_config.TextColumn("持股方向判讀", width="medium"),
            "輔助提示": st.column_config.TextColumn("新聞與籌碼提示", width="medium"),
            "最高新聞Level": st.column_config.NumberColumn("最高 Level", format="%.0f"),
            "負面新聞則數": st.column_config.NumberColumn("負面新聞", format="%d 則"),
            "曝險金額": st.column_config.NumberColumn("曝險金額", format="%.0f"),
            "增持家數": st.column_config.NumberColumn("增持機構", format="%d 家"),
            "減持家數": st.column_config.NumberColumn("減持機構", format="%d 家"),
        },
    )
    ticker_options = shown["Ticker"].dropna().astype(str).tolist()
    if ticker_options:
        st.markdown("#### 單一股票機構明細")
        st.caption("選擇股票後，顯示持股增減幅度最大的 10 家機構；其餘資料可在完整明細中展開查看。")
        detail_filter_key = re.sub(r"[^A-Za-z0-9]+", "_", f"{priority_view}_{search}")
        selected_ticker = st.selectbox("查看機構明細", ticker_options, key=f"sec13f_detail_ticker_{detail_filter_key}")
        selected_summary = summary[summary["Ticker"] == selected_ticker].iloc[0]
        st.markdown(f"#### {selected_ticker}｜{selected_summary['Company']}｜{quarter}")
        selected_detail = detail[detail["Ticker"] == selected_ticker].copy()
        if selected_detail.empty:
            st.info("此股票沒有可顯示的機構明細。")
        else:
            status_map = {"INCREASED": "增持", "DECREASED": "減持", "NEW": "新建倉", "CLOSED": "清倉", "UNCHANGED": "持股不變"}
            selected_detail["狀態"] = selected_detail.get("Status", "").astype(str).str.upper().map(status_map).fillna(selected_detail.get("Status", ""))
            selected_detail["Change Shares"] = pd.to_numeric(selected_detail.get("Change Shares"), errors="coerce").fillna(0)
            changed_managers = selected_detail[selected_detail["Change Shares"].ne(0)].copy()
            if not changed_managers.empty:
                changed_managers["變動幅度"] = changed_managers["Change Shares"].abs()
                changed_managers = changed_managers.nlargest(10, "變動幅度").sort_values("Change Shares")
                changed_managers["方向"] = np.where(changed_managers["Change Shares"] >= 0, "增持／新建倉", "減持／清倉")
                changed_managers["股數標籤"] = changed_managers["Change Shares"].map(lambda value: f"{value:+,.0f}")
                manager_figure = px.bar(
                    changed_managers, x="Change Shares", y="Institution", orientation="h", color="方向", text="股數標籤",
                    hover_data={"Shares": ":,.0f", "Previous Shares": ":,.0f", "Change %": ":.1%", "狀態": True,
                                "變動幅度": False, "方向": False},
                    color_discrete_map={"增持／新建倉": "#7CA9E8", "減持／清倉": "#E89A9A"},
                )
                manager_figure.update_traces(textposition="outside", cliponaxis=False, marker_line_color="white", marker_line_width=.6)
                manager_figure.add_vline(x=0, line_color="#64748B", line_width=1)
                manager_figure.update_layout(
                    height=max(350, 35 * len(changed_managers) + 80), margin={"t": 15, "b": 20, "l": 20, "r": 80},
                    plot_bgcolor="white", paper_bgcolor="white", xaxis_title="本季相較前季增減股數", yaxis_title=None,
                    legend_title=None, legend={"orientation": "h", "yanchor": "bottom", "y": 1.01, "xanchor": "right", "x": 1},
                )
                manager_figure.update_xaxes(showgrid=True, gridcolor="#E2E8F0", tickformat=",.0f", zeroline=False)
                manager_figure.update_yaxes(showgrid=False)
                st.markdown("##### 機構持股增減前 10 名")
                st.caption("橫條向右為增持或新建倉，向左為減持或清倉；長度代表相較前季的增減股數，並非交易金額。")
                st.plotly_chart(manager_figure, use_container_width=True, config={"displaylogo": False})
            with st.expander("查看完整機構持股明細", expanded=False):
                detail_columns = ["Institution", "狀態", "Shares", "Previous Shares", "Change Shares", "Change %", "Filing Date"]
                st.dataframe(selected_detail[[column for column in detail_columns if column in selected_detail.columns]], use_container_width=True, hide_index=True,
                             column_config={"Institution": st.column_config.TextColumn("機構", width="large"), "Shares": st.column_config.NumberColumn("本季持股", format="%,.0f"),
                                            "Previous Shares": st.column_config.NumberColumn("前季持股", format="%,.0f"), "Change Shares": st.column_config.NumberColumn("增減股數", format="%+,.0f"),
                                            "Change %": st.column_config.NumberColumn("增減幅", format="%.1%%"), "Filing Date": st.column_config.DateColumn("申報日期")})
    st.caption("下載檔包含股票彙總、13F 機構明細與執行紀錄；如有新季度資料，請到第六頁人工上傳並套用。")
    st.download_button("下載目前 13F 完整 Excel", path.read_bytes(), path.name, use_container_width=True, key="download_sec13f")


def save_status(ticker: str, company: str, status: str, owner: str, next_review: str, note: str) -> None:
    key = ticker or company.lower()
    with sqlite3.connect(DB_PATH) as connection:
        connection.execute(
            "INSERT INTO manual_status VALUES(?,?,?,?,?,?,?,?) ON CONFLICT(company_key) DO UPDATE SET status=excluded.status,owner=excluded.owner,next_review=excluded.next_review,note=excluded.note,updated_at=excluded.updated_at",
            (key, ticker, company, status, owner, next_review, note, datetime.now(TAIPEI).isoformat(timespec="seconds")),
        )


PAGE_COPY = {
    "今日任務": ("今日新聞擷取", "選擇公司範圍與時間，一次取得可交付的新聞 Excel。"),
    "編輯頁面": ("負面新聞編輯", "可直接修正分類結果；儲存後，儀表板與持續追蹤會立即採用最新版。"),
    "風險儀表板": ("負面新聞風險儀表板", "先掌握高風險事件，再查看公司與新聞明細。"),
    "查看部位": ("部位與 13F 機構動向", "先查看季度機構持股變化，再把實際曝險與負面新聞放在一起交叉分析。"),
    "持續追蹤": ("持續追蹤工作台", "集中查看追蹤事件、事件明細與事件後股價。"),
    "設定與說明": ("設定與系統說明", "管理事件規則版本與程式版本恢復。"),
}


def page_heading(page_name: str) -> None:
    title, description = PAGE_COPY[page_name]
    st.markdown(f"<div class='page-kicker'>{page_name}</div><div class='page-title'>{title}</div><div class='page-description'>{description}</div>", unsafe_allow_html=True)


initialize()
seed_finbert_cache()
seed_translation_cache()
st.set_page_config(page_title=APP_TITLE, page_icon="🛡️", layout="wide")
st.markdown("""
<style>
  .stApp { background:#F7F9FC; }
  .block-container { max-width:1500px; padding-top:1.4rem; padding-bottom:4rem; }
  .app-hero { background:linear-gradient(120deg,#102A43,#1D4ED8); color:white; border-radius:16px; padding:20px 26px; margin-bottom:14px; box-shadow:0 10px 30px rgba(15,42,67,.12); }
  .app-hero-title { font-size:25px; font-weight:750; letter-spacing:.02em; }
  .app-hero-sub { opacity:.78; font-size:13px; margin-top:4px; }
  .page-kicker { color:#2563EB; font-size:13px; font-weight:700; letter-spacing:.08em; margin-top:24px; }
  .page-title { color:#172B4D; font-size:31px; line-height:1.25; font-weight:780; margin-top:4px; }
  .page-description { color:#64748B; font-size:16px; margin:7px 0 20px; }
  div[data-testid='stRadio'] > div { background:white; border:1px solid #E2E8F0; border-radius:13px; padding:10px; box-shadow:0 3px 12px rgba(15,23,42,.04); }
  div[data-testid='stRadio'] div[role='radiogroup'] { display:flex; gap:8px; width:100%; }
  div[data-testid='stRadio'] label[data-baseweb='radio'] { flex:1; min-height:50px; display:flex; align-items:center; justify-content:center; border-radius:9px; padding:8px 10px; color:#64748B; font-weight:700; cursor:pointer; transition:all .16s ease; }
  div[data-testid='stRadio'] label[data-baseweb='radio']:hover { background:#F1F5F9; color:#1D4ED8; }
  div[data-testid='stRadio'] label[data-baseweb='radio']:has(input:checked) { background:#EAF2FF; color:#1D4ED8; box-shadow:inset 0 0 0 1px #BFDBFE; }
  div[data-testid='stRadio'] label[data-baseweb='radio'] > div:first-child { display:none; }
  .method-card { background:white; border:1px solid #DBEAFE; border-left:6px solid #2563EB; border-radius:12px; padding:16px 18px; margin:4px 0 20px; }
  .method-title { color:#1E3A8A; font-weight:750; font-size:17px; }
  .method-desc { color:#64748B; font-size:14px; margin-top:5px; }
  .status-title { color:#172B4D; font-size:27px; line-height:1.15; font-weight:750; margin:0 0 4px; }
  .previous-result-grid { display:grid; grid-template-columns:repeat(5,minmax(0,1fr)); gap:16px; margin:4px 0 16px; }
  .previous-result-card { background:white; border:1px solid #E2E8F0; border-radius:12px; padding:16px 18px; min-height:108px; box-shadow:0 2px 8px rgba(15,23,42,.035); }
  .previous-result-label { color:#334155; font-size:15px; margin-bottom:10px; }
  .previous-result-value { color:#1E293B; font-size:24px; line-height:1.3; font-weight:520; white-space:normal; overflow-wrap:anywhere; }
  .previous-result-value.is-period { display:flex; flex-wrap:wrap; column-gap:6px; font-size:24px; font-weight:520; overflow-wrap:normal; }
  .previous-result-value.is-period span { white-space:nowrap; }
  .previous-result-value.is-datetime { font-size:24px; font-weight:520; white-space:normal; overflow-wrap:normal; word-break:keep-all; }
  .settings-info-grid { display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:16px; margin:4px 0 16px; }
  .settings-info-card { background:white; border:1px solid #E2E8F0; border-radius:12px; padding:15px 18px; min-height:96px; box-shadow:0 2px 8px rgba(15,23,42,.035); }
  .settings-info-label { color:#475569; font-size:14px; margin-bottom:8px; }
  .settings-info-value { color:#1E293B; font-size:18px; line-height:1.35; font-weight:520; overflow-wrap:break-word; word-break:normal; }
  .settings-info-value.is-number { font-size:26px; font-weight:650; }
  .tracking-metric-grid { display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:16px; margin:4px 0 16px; }
  .tracking-metric-card { background:white; border:1px solid #E2E8F0; border-radius:12px; padding:15px 18px; min-height:96px; box-shadow:0 2px 8px rgba(15,23,42,.035); }
  .tracking-metric-label { color:#334155; font-size:15px; margin-bottom:8px; }
  .tracking-metric-value { color:#1E293B; font-size:25px; line-height:1.3; font-weight:520; white-space:normal; }
  .tracking-metric-value.is-date { font-size:20px; line-height:1.35; white-space:nowrap; }
  .status-legend { display:flex; flex-wrap:wrap; gap:7px 18px; background:white; border:1px solid #E2E8F0; border-radius:10px; padding:10px 14px; margin:3px 0 14px; color:#475569; font-size:13px; line-height:1.45; }
  .status-legend span { white-space:nowrap; }
  div[data-testid='stMetric'] { background:white; border:1px solid #E2E8F0; border-radius:12px; padding:14px 16px; box-shadow:0 2px 8px rgba(15,23,42,.035); }
  div[data-testid='stFileUploader'] { background:white; border-radius:12px; }
  div[data-testid='stDataFrame'] { border:1px solid #E2E8F0; border-radius:12px; overflow:hidden; }
  div.stButton > button[kind='primary'] { min-height:46px; border-radius:10px; font-weight:700; }
  div.stDownloadButton > button { min-height:43px; border-radius:10px; font-weight:650; }
  h3 { color:#172B4D !important; margin-top:1.4rem !important; }
  @media (max-width:1000px) { .previous-result-grid { grid-template-columns:repeat(2,minmax(0,1fr)); } .settings-info-grid { grid-template-columns:1fr; } .tracking-metric-grid { grid-template-columns:repeat(2,minmax(0,1fr)); } }
  @media (max-width:800px) { div[data-testid='stRadio'] div[role='radiogroup'] { display:block; } div[data-testid='stRadio'] label[data-baseweb='radio'] { justify-content:flex-start; } .page-title { font-size:25px; } .previous-result-grid { grid-template-columns:1fr; } .tracking-metric-grid { grid-template-columns:1fr; } }
</style>
""", unsafe_allow_html=True)
st.markdown(f"<div class='app-hero'><div class='app-hero-title'>🛡️ {APP_TITLE}</div></div>", unsafe_allow_html=True)

if not SIMPLE_SITE and st.session_state.pop("_go_dashboard", False):
    st.session_state["page"] = "風險儀表板"
if SIMPLE_SITE:
    page = "今日任務"
    page_heading(page)
    st.markdown("<div class='method-card'><div class='method-title'>今日任務｜雙方法合併與負面新聞整理</div><div class='method-desc'>系統會依序抓取多來源新聞與 Google News，合併兩份結果、刪除重複新聞、執行金融情緒分析 (FinBERT)，最後依事件規則整理負面新聞。</div></div>", unsafe_allow_html=True)
else:
    nav_icons = {
        "今日任務": "① 今日任務",
        "編輯頁面": "② 編輯頁面",
        "風險儀表板": "③ 風險儀表板",
        "查看部位": "④ 查看部位",
        "持續追蹤": "⑤ 持續追蹤",
        "設定與說明": "⑥ 設定與說明",
    }
    page = st.radio("頁面", list(nav_icons), horizontal=True, label_visibility="collapsed", key="page", format_func=lambda value: nav_icons[value])

if page == "編輯頁面":
    page_heading(page)
    uploaded_correction = st.file_uploader(
        "上傳已修正的負面新聞 Excel",
        type=["xlsx"],
        help="可上傳系統匯出的三工作表檔案；套用後，後續頁面會讀取這份最新版。",
        key="corrected_negative_upload",
    )
    if st.button("檢查並套用上傳檔案", type="primary", use_container_width=True, disabled=uploaded_correction is None):
        try:
            uploaded_negative, uploaded_other_sheets = load_negative_workbook(io.BytesIO(uploaded_correction.getvalue()))
            stamp_match = re.search(r"20\d{6}", uploaded_correction.name)
            stamp = stamp_match.group(0) if stamp_match else datetime.now(TAIPEI).strftime("%Y%m%d")
            upload_target = OUTPUT_DIR / f"美股_{stamp}_負面新聞爬蟲.xlsx"
            save_corrected_negative_workbook(upload_target, uploaded_negative, uploaded_other_sheets)
            st.session_state.pop("negative_editor", None)
            st.success(f"已套用修正版：{upload_target.name}，共 {len(uploaded_negative):,} 筆。")
            st.rerun()
        except Exception as exc:
            st.error(f"修正版無法套用：{exc}")
    edit_files = output_files()
    if not edit_files:
        st.info("目前沒有可編輯的負面新聞。請先完成今日新聞擷取，或在上方上傳修正版檔案。")
    else:
        edit_file = st.selectbox("選擇要編輯的結果", edit_files, format_func=lambda path: path.name, key="negative_edit_file")
        edit_data, edit_other_sheets = load_negative_workbook(edit_file)
        st.caption("可直接點選任一儲存格修改；完成後按下方「儲存編輯結果」。未按儲存不會改動正式資料。")
        editor_source = edit_data.copy()
        editor_source["Published Time"] = editor_source["Published Time"].astype(str)
        if "Action" in editor_source.columns and "Title_ZH" in editor_source.columns:
            action_values = editor_source.pop("Action")
            editor_source.insert(editor_source.columns.get_loc("Title_ZH") + 1, "Action", action_values)
        editor_levels = editor_source.pop("Level")
        level_insert_at = editor_source.columns.get_loc("Company") + 1
        editor_source.insert(level_insert_at, "Level", editor_levels)
        grid_builder = GridOptionsBuilder.from_dataframe(editor_source)
        grid_builder.configure_default_column(editable=True, sortable=True, filter=True, resizable=True, minWidth=90)
        grid_builder.configure_column("Published Time", header_name="發布時間", width=170)
        grid_builder.configure_column("Ticker", width=90)
        grid_builder.configure_column("Company", width=160)
        grid_builder.configure_column(
            "Level", width=82, editable=True,
            cellEditor="agSelectCellEditor", cellEditorParams={"values": [3, 4, 5]},
        )
        grid_builder.configure_column("Title", header_name="英文標題", width=430)
        grid_builder.configure_column("Title_ZH", header_name="中文標題", width=430)
        grid_builder.configure_column("Action", header_name="處理建議", width=300)
        grid_builder.configure_column("FinBERT", header_name="金融情緒分析 (FinBERT)", width=175, type=["numericColumn"])
        grid_builder.configure_column("URL", header_name="新聞網址", width=320)
        grid_builder.configure_grid_options(
            rowHeight=42,
            headerHeight=42,
            getRowStyle=JsCode("""
                function(params) {
                    const level = Number(params.data.Level || 0);
                    if (level === 5) {
                        return {backgroundColor: '#FDECEC', color: '#1E293B'};
                    }
                    if (level === 4) {
                        return {backgroundColor: '#FFF4D6', color: '#1E293B'};
                    }
                    return {backgroundColor: '#FFFFFF', color: '#1E293B'};
                }
            """),
        )
        grid_response = AgGrid(
            editor_source,
            gridOptions=grid_builder.build(),
            height=540,
            use_container_width=True,
            fit_columns_on_grid_load=False,
            columns_auto_size_mode=ColumnsAutoSizeMode.NO_AUTOSIZE,
            data_return_mode=DataReturnMode.AS_INPUT,
            update_on=["cellValueChanged"],
            allow_unsafe_jscode=True,
            theme="streamlit",
            custom_css={
                ".ag-root-wrapper": {
                    "border": "1px solid #CBD5E1 !important",
                    "border-radius": "8px !important",
                },
                ".ag-header-cell": {
                    "border-right": "1px solid #CBD5E1 !important",
                    "border-bottom": "1px solid #CBD5E1 !important",
                },
                ".ag-cell": {
                    "border-right": "1px solid #D8DEE8 !important",
                    "border-bottom": "1px solid #D8DEE8 !important",
                },
                ".ag-row": {
                    "border-bottom": "none !important",
                },
            },
            key=f"negative_editor_{edit_file.name}_{edit_file.stat().st_mtime_ns}",
        )
        edited_data = pd.DataFrame(grid_response["data"])
        save_column, download_column = st.columns(2)
        if save_column.button("儲存編輯結果", type="primary", use_container_width=True):
            try:
                save_corrected_negative_workbook(
                    edit_file,
                    pd.DataFrame(edited_data),
                    edit_other_sheets,
                )
                st.success(f"已儲存 {edit_file.name}；其他頁面將使用這份最新版。")
                st.rerun()
            except Exception as exc:
                st.error(f"無法儲存：{exc}")
        download_column.download_button("下載目前正式版本", edit_file.read_bytes(), edit_file.name, use_container_width=True)

if page == "查看部位":
    page_heading(page)
    position_subpage = st.radio(
        "查看內容",
        ["部位曝險", "13F 機構動向"],
        horizontal=True,
        key="position_subpage",
    )
    if position_subpage == "部位曝險":
        with st.expander("上傳曝險部位清單（選填）", expanded=not EXPOSURE_PATH.exists()):
            st.caption("支援 Ticker＋曝險金額＋業務別的長表，也支援以財管、經紀、自營分欄的寬表；上傳新版後會沿用至下次更換。")
            exposure_upload = st.file_uploader("選擇部位 Excel", type=["xlsx"], key="v4_exposure_upload")
            if st.button("檢查並套用部位清單", type="primary", use_container_width=True,
                         disabled=exposure_upload is None, key="v4_apply_exposure"):
                try:
                    checked_exposure = parse_exposure_frame(pd.read_excel(io.BytesIO(exposure_upload.getvalue()), dtype=str))
                    if checked_exposure.empty:
                        raise ValueError("檔案中沒有非零的有效部位")
                    EXPOSURE_PATH.write_bytes(exposure_upload.getvalue())
                    load_exposure_file.clear()
                    st.success(f"已套用 {exposure_upload.name}：{checked_exposure['Ticker'].nunique():,} 家公司、{len(checked_exposure):,} 筆部位。")
                    st.rerun()
                except Exception as exc:
                    st.error(f"部位清單無法套用：{exc}")

    exposure_df = pd.DataFrame(columns=["Ticker", "Company", "業務別", "曝險金額"])
    if EXPOSURE_PATH.is_file():
        try:
            exposure_df = load_exposure_file(str(EXPOSURE_PATH), EXPOSURE_PATH.stat().st_mtime_ns)
            if position_subpage == "部位曝險":
                st.caption(
                    f"目前部位版本：{EXPOSURE_PATH.name}｜{exposure_df['Ticker'].nunique():,} 家公司｜"
                    f"更新時間 {datetime.fromtimestamp(EXPOSURE_PATH.stat().st_mtime).strftime('%Y-%m-%d %H:%M')}"
                )
        except Exception as exc:
            st.warning(f"目前部位檔無法讀取：{exc}")

    historical_position_news = prepare_position_news(load_negative_files())
    latest_position_news = pd.DataFrame()
    latest_negative_files = output_files()
    if latest_negative_files:
        try:
            latest_position_news = validate_negative_frame(pd.read_excel(latest_negative_files[0], sheet_name="負面新聞"))
            latest_stamp = re.search(r"20\d{6}", latest_negative_files[0].name)
            latest_position_news["資料日期"] = latest_stamp.group(0) if latest_stamp else ""
            latest_position_news = prepare_position_news(latest_position_news)
        except Exception as exc:
            st.warning(f"最新負面新聞檔無法讀取：{exc}")

    if position_subpage == "13F 機構動向":
        render_sec13f_panel(historical_position_news, exposure_df)
        st.stop()

    st.markdown("### 部位曝險與新聞趨勢")
    if historical_position_news.empty:
        st.info("目前沒有可供分析的 Level 3～5 負面新聞資料。")
    else:
        business_options = exposure_business_options(exposure_df)
        business_filter = st.selectbox(
            "依業務別查看曝險",
            business_options,
            key="v4_exposure_business_filter",
            help="選擇全部時，同一家公司在各業務別的曝險會加總。",
        )
        if not exposure_df.empty and len(business_options) == 1:
            st.caption("目前部位檔沒有業務別資料，因此只能顯示全部加總；若要切換財管、經紀或自營，請在部位檔加入「業務別」欄，或分別建立同名金額欄。")
        exposure_map = exposure_map_for_business(exposure_df, business_filter)
        position_news = latest_position_news.copy()
        history_for_trend = historical_position_news.copy()
        if exposure_map:
            position_news = position_news[position_news["Ticker"].isin(exposure_map)].copy()
            history_for_trend = history_for_trend[history_for_trend["Ticker"].isin(exposure_map)].copy()
        position_news["曝險金額"] = position_news["Ticker"].map(exposure_map).fillna(0.0)

        if exposure_map:
            company_risk = (
                position_news[position_news["曝險金額"] > 0]
                .groupby(["Ticker", "Company"], as_index=False)
                .agg(曝險金額=("曝險金額", "max"), 負面新聞則數=("新聞鍵", "nunique"), 最高Level=("Level", "max"))
            )
            company_risk["風險嚴重度"] = company_risk["最高Level"] * company_risk["負面新聞則數"]
            if company_risk.empty:
                st.info("目前的負面新聞沒有命中這份部位清單。")
            else:
                metric_cols = st.columns(3)
                metric_cols[0].metric("命中部位公司", f"{company_risk['Ticker'].nunique():,} 家")
                metric_cols[1].metric("命中曝險金額", f"{company_risk['曝險金額'].sum():,.0f}")
                metric_cols[2].metric("相關負面新聞", f"{company_risk['負面新聞則數'].sum():,.0f} 則")
                st.markdown("### 曝險風險象限氣泡圖")
                exposure_median = float(company_risk["曝險金額"].median())
                severity_median = float(company_risk["風險嚴重度"].median())
                figure_exposure = px.scatter(
                    company_risk,
                    x="曝險金額",
                    y="風險嚴重度",
                    size="負面新聞則數",
                    color="最高Level",
                    text="Ticker",
                    hover_name="Company",
                    hover_data={
                        "Ticker": True,
                        "曝險金額": ":,.0f",
                        "風險嚴重度": True,
                        "負面新聞則數": True,
                        "最高Level": True,
                    },
                    color_continuous_scale=["#BFDBFE", "#3B82F6", "#1E3A8A"],
                    range_color=[3, 5],
                    size_max=36,
                )
                figure_exposure.update_traces(
                    textposition="top center",
                    marker={"opacity": 0.82, "line": {"color": "white", "width": 1.2}},
                )
                figure_exposure.add_vline(
                    x=exposure_median, line_dash="dash", line_color="#94A3B8", line_width=2
                )
                figure_exposure.add_hline(
                    y=severity_median, line_dash="dash", line_color="#94A3B8", line_width=2
                )
                figure_exposure.update_layout(
                    xaxis_title="曝險金額",
                    yaxis_title="風險嚴重度（最高 Level × 新聞則數）",
                    height=430,
                    margin={"t": 28, "b": 20, "l": 20, "r": 20},
                    coloraxis_colorbar={"title": "最高 Level", "tickvals": [3, 4, 5]},
                    plot_bgcolor="white",
                    paper_bgcolor="white",
                )
                figure_exposure.update_xaxes(showgrid=True, gridcolor="#E2E8F0", rangemode="tozero")
                figure_exposure.update_yaxes(showgrid=True, gridcolor="#E2E8F0", rangemode="tozero")
                st.plotly_chart(figure_exposure, use_container_width=True, config={"displaylogo": False})
                st.caption("越靠右代表曝險越高，越靠上代表事件越嚴重；氣泡越大表示負面新聞越多，虛線是目前公司的中位數。")
        else:
            st.info("上傳部位清單後，這裡會顯示曝險風險象限氣泡圖；下方新聞頻率分析仍可直接使用。")

        st.markdown("### 新聞頻率異常與走勢")
        daily_company = (
            history_for_trend.groupby(["發布日期", "Ticker", "Company"], as_index=False)
            .agg(則數=("新聞鍵", "nunique"))
        )
        available_days = sorted(daily_company["發布日期"].dropna().unique())
        if len(available_days) < 2:
            st.caption("至少需要兩個發布日的負面新聞，才能比較新聞頻率變化。")
        else:
            latest_day = pd.Timestamp(available_days[-1])
            baseline_days = [pd.Timestamp(value) for value in available_days[-7:-1]]
            latest_counts = daily_company[daily_company["發布日期"] == latest_day].set_index(["Ticker", "Company"])["則數"]
            baseline_data = daily_company[daily_company["發布日期"].isin(baseline_days)]
            baseline_totals = baseline_data.groupby(["Ticker", "Company"])["則數"].sum()
            baseline_divisor = max(len(baseline_days), 1)
            anomaly_rows = []
            for (ticker, company), latest_count in latest_counts.items():
                baseline_average = float(baseline_totals.get((ticker, company), 0.0)) / baseline_divisor
                is_new_spike = baseline_average == 0 and latest_count >= 3
                is_ratio_spike = baseline_average > 0 and latest_count >= 2 and latest_count >= baseline_average * 2
                if is_new_spike or is_ratio_spike:
                    anomaly_rows.append({
                        "Ticker": ticker, "Company": company,
                        f"{latest_day.strftime('%Y-%m-%d')} 則數": int(latest_count),
                        f"過去 {baseline_divisor} 日平均": round(baseline_average, 1),
                        "倍數": "新出現" if baseline_average == 0 else f"{latest_count / baseline_average:.1f}x",
                    })
            if anomaly_rows:
                st.warning(
                    f"偵測到 {len(anomaly_rows)} 家公司負面新聞則數異常增加（"
                    f"{latest_day.strftime('%Y-%m-%d')} 與過去 {baseline_divisor} 個發布日平均相比）"
                )
                st.dataframe(pd.DataFrame(anomaly_rows), use_container_width=True, hide_index=True,
                             height=min(330, 42 + 35 * len(anomaly_rows)))
            else:
                st.success(f"{latest_day.strftime('%Y-%m-%d')} 未偵測到明顯的負面新聞頻率異常。")

            company_labels = (
                history_for_trend[["Ticker", "Company"]].drop_duplicates()
                .assign(選項=lambda data: data["Ticker"] + "｜" + data["Company"])
                .sort_values("選項")["選項"].tolist()
            )
            trend_pick = st.selectbox("查看頻率趨勢", ["全部持有部位" if exposure_map else "全市場"] + company_labels,
                                      key="v4_position_trend_pick")
            if "｜" in trend_pick:
                trend_ticker = trend_pick.split("｜", 1)[0]
                trend_source = daily_company[daily_company["Ticker"] == trend_ticker]
            else:
                trend_source = daily_company
            trend_data = trend_source.groupby("發布日期", as_index=False)["則數"].sum().sort_values("發布日期")
            trend_chart = alt.Chart(trend_data).mark_line(point=alt.OverlayMarkDef(size=70), color="#1677D2", strokeWidth=2.5).encode(
                x=alt.X("發布日期:T", title=None, axis=alt.Axis(format="%m/%d", labelAngle=0)),
                y=alt.Y("則數:Q", title="負面新聞則數", scale=alt.Scale(zero=True)),
                tooltip=[alt.Tooltip("發布日期:T", title="日期", format="%Y-%m-%d"), alt.Tooltip("則數:Q", title="負面新聞則數")],
            ).properties(height=320)
            st.altair_chart(trend_chart, use_container_width=True)

if page == "今日任務":
    st.markdown("### 1. 選擇公司範圍")
    company_universe_options = ["Dow Jones 30", "S&P 500", "自行上傳最新版"]
    universe = st.radio("本次要使用的公司名單", company_universe_options, index=1, horizontal=True)
    company_upload = None
    if universe == "自行上傳最新版":
        st.caption("上傳名單時，同一個 Excel 需包含「Dow Jones」與「S&P 500」兩張工作表。")
        company_upload = st.file_uploader("上傳最新版美股指數成分股 Excel", type=["xlsx"], key="company_list_upload")
        if st.button("檢查並套用新版", use_container_width=True, disabled=company_upload is None):
            try:
                checked_dow = load_company_upload(company_upload, sheet_name="Dow Jones")
                checked_sp500 = load_company_upload(company_upload, sheet_name="S&P 500")
                if checked_dow.empty or checked_sp500.empty:
                    raise ValueError("Dow Jones 或 S&P 500 工作表沒有可用資料")
                backup_dir = DATA_ROOT / "backups"
                backup_dir.mkdir(parents=True, exist_ok=True)
                if COMPANY_INDEX_PATH.is_file():
                    backup_stamp = datetime.now(TAIPEI).strftime("%Y%m%d_%H%M%S")
                    (backup_dir / f"美股指數成分股_{backup_stamp}.xlsx").write_bytes(COMPANY_INDEX_PATH.read_bytes())
                COMPANY_INDEX_PATH.write_bytes(company_upload.getvalue())
                applied_at = datetime.now(TAIPEI).strftime("%Y-%m-%d %H:%M:%S")
                with sqlite3.connect(DB_PATH) as connection:
                    connection.execute("INSERT OR REPLACE INTO settings(key,value) VALUES('company_index_original_name',?)", (company_upload.name,))
                    connection.execute("INSERT OR REPLACE INTO settings(key,value) VALUES('company_index_updated_at',?)", (applied_at,))
                st.success(f"新版名單已套用：Dow Jones {len(checked_dow):,} 家、S&P 500 {len(checked_sp500):,} 家。之後會持續沿用到下次更新。")
            except Exception as exc:
                st.error(f"新版公司名單無法套用：{exc}")
    else:
        with sqlite3.connect(DB_PATH) as connection:
            saved_company_name = connection.execute("SELECT value FROM settings WHERE key='company_index_original_name'").fetchone()
        if COMPANY_INDEX_PATH.is_file():
            current_company_name = saved_company_name[0] if saved_company_name else COMPANY_INDEX_PATH.name
            sheet_label = "Dow Jones 工作表" if universe == "Dow Jones 30" else "S&P 500 工作表"
            st.success(f"目前使用：{current_company_name}｜{sheet_label}（更新同一份檔案後會自動套用）")
        elif universe == "Dow Jones 30":
            st.info("目前使用程式內建的 Dow Jones 30 名單；上傳新版後將改用上傳版本。")
        else:
            st.info("目前使用線上更新的 S&P 500 名單；上傳新版後將改用上傳版本。")
    now = datetime.now(TAIPEI).replace(hour=9, minute=0, second=0, microsecond=0)
    previous_end = last_success_end()
    default_start = (previous_end or (now - timedelta(days=3))).replace(hour=9, minute=0, second=0, microsecond=0)
    if previous_end:
        st.caption(f"前次成功結束日期：{previous_end:%Y-%m-%d}；本次時間預設為 09:00。")
    st.markdown("### 2. 設定新聞期間")
    start_col, end_col = st.columns(2)
    start_date = start_col.date_input("開始日期", value=default_start.date())
    start_hour = start_col.selectbox("開始時間", list(range(24)), index=default_start.hour, format_func=lambda hour: f"{hour:02d}:00")
    end_date = end_col.date_input("結束日期", value=now.date())
    end_hour = end_col.selectbox("結束時間", list(range(24)), index=now.hour, format_func=lambda hour: f"{hour:02d}:00")
    start_dt = datetime.combine(start_date, datetime.min.time(), TAIPEI) + timedelta(hours=start_hour)
    end_dt = datetime.combine(end_date, datetime.min.time(), TAIPEI) + timedelta(hours=end_hour)
    if start_dt > end_dt:
        st.error("開始時間不可晚於結束時間")
    st.caption(f"預計擷取：{start_dt:%Y-%m-%d %H:%M} ～ {end_dt:%Y-%m-%d %H:%M}（台北時間）")
    st.markdown("### 3. 今日新聞擷取")
    crawl_method = st.radio("選擇擷取方法", ["方法一｜多來源快速擷取", "方法二｜Google News＋金融情緒分析 (FinBERT)", "方法一＋方法二｜完整整合"], index=2, horizontal=True)
    if crawl_method == "方法一｜多來源快速擷取":
        st.markdown("<div class='method-card'><div class='method-title'>方法一｜多來源快速擷取</div><div class='method-desc'>彙整經濟日報、鉅亨網與 CNBC，再由 TradingView 逐公司補抓。</div></div>", unsafe_allow_html=True)
    elif crawl_method == "方法二｜Google News＋金融情緒分析 (FinBERT)":
        st.markdown("<div class='method-card'><div class='method-title'>方法二｜Google News＋金融情緒分析 (FinBERT)</div><div class='method-desc'>分批查詢英文 Google News RSS，完成後進行金融情緒分析。</div></div>", unsafe_allow_html=True)
    else:
        st.markdown("<div class='method-card'><div class='method-title'>方法一＋方法二｜完整整合</div><div class='method-desc'>依序完成兩種爬蟲、合併去重，再統一執行金融情緒分析 (FinBERT)。</div></div>", unsafe_allow_html=True)
    method_name = {"方法一｜多來源快速擷取": "方法一", "方法二｜Google News＋金融情緒分析 (FinBERT)": "方法二", "方法一＋方法二｜完整整合": "方法一＋方法二"}[crawl_method]
    registry = crawl_job_registry()
    current_job = registry.get("current")
    is_running = bool(current_job and current_job.get("status") in ("running", "stopping"))
    if st.button(
        f"開始執行{method_name}", type="primary",
        disabled=start_dt > end_dt or is_running or universe == "自行上傳最新版",
        use_container_width=True,
    ):
        try:
            if universe.startswith("Dow"):
                companies = load_saved_companies("Dow Jones") if COMPANY_INDEX_PATH.is_file() else pd.DataFrame(DOW_30, columns=["Ticker", "Company"])
            elif universe.startswith("S&P"):
                companies = load_saved_companies("S&P 500") if COMPANY_INDEX_PATH.is_file() else load_sp500()
            else:
                raise ValueError("請先套用新版名單，再選擇 Dow Jones 30 或 S&P 500 開始執行")
            run_id = create_run(start_dt, end_dt, f"{method_name}｜{universe}")
            script_context = get_script_run_ctx()
            use_method1 = method_name in ("方法一", "方法一＋方法二")
            use_method2 = method_name in ("方法二", "方法一＋方法二")
            job = {
                "run_id": run_id, "status": "running", "method": method_name,
                "started_monotonic": time.monotonic(), "started_at": datetime.now(TAIPEI).isoformat(timespec="seconds"),
                "started_epoch_ms": int(time.time() * 1000),
                "progress": 0.0, "progress_text": "準備開始", "stop_event": threading.Event(),
                "stage_times": [], "current_stage": None, "current_stage_started": None,
                "script_context": script_context,
                "session_id": script_context.session_id if script_context is not None else "",
                "method1_monitor": {
                    "status": "pending" if use_method1 else "skipped", "stage": "等待執行" if use_method1 else "本次未執行",
                    "rows": 0, "errors": [], "source_counts": {},
                },
                "method2_monitor": {
                    "status": "pending" if use_method2 else "skipped", "stage": "等待執行" if use_method2 else "本次未執行",
                    "rows": 0, "raw_rows": 0, "negative_rows": 0, "batches": 0, "errors": [],
                },
            }
            registry["current"] = job
            worker = threading.Thread(
                target=run_crawl_job,
                args=(job, method_name, universe, companies.copy(), start_dt, end_dt),
                name=f"news-crawl-{run_id}", daemon=True,
            )
            job["thread"] = worker
            worker.start()
            st.rerun()
        except Exception as exc:
            st.error(f"無法開始抓取：{exc}")

    @st.fragment(run_every=3 if is_running else None)
    def render_crawl_status():
        job = crawl_job_registry().get("current")
        if not job:
            return False
        end_tick = job.get("finished_monotonic", time.monotonic())
        elapsed_text = format_elapsed(end_tick - job["started_monotonic"])
        status = job.get("status", "running")
        with st.container(border=True):
            left, right = st.columns([3, 1])
            with left:
                st.markdown("<div class='status-title'>執行狀態</div>", unsafe_allow_html=True)
                st.progress(job.get("progress", 0), text=job.get("progress_text", "執行中"))
            if status in ("running", "stopping"):
                started_epoch_ms = int(job.get("started_epoch_ms", time.time() * 1000))
                with right:
                    components.html(
                        f"""
                        <div style="font-family:Arial,sans-serif;border:1px solid #e2e8f0;border-radius:12px;
                                    padding:10px 16px;background:white;height:78px;box-sizing:border-box;">
                          <div style="font-size:14px;color:#334155;margin-bottom:3px;">抓新聞所用時間</div>
                          <div id="crawl-elapsed" style="font-size:26px;color:#1e293b;line-height:1.08;">{elapsed_text}</div>
                        </div>
                        <script>
                          const started = {started_epoch_ms};
                          const pad = n => String(n).padStart(2, '0');
                          const update = () => {{
                            const total = Math.max(0, Math.floor((Date.now() - started) / 1000));
                            const h = Math.floor(total / 3600);
                            const m = Math.floor((total % 3600) / 60);
                            const s = total % 60;
                            document.getElementById('crawl-elapsed').textContent = `${{pad(h)}}:${{pad(m)}}:${{pad(s)}}`;
                          }};
                          update(); setInterval(update, 1000);
                        </script>
                        """,
                        height=82,
                    )
            else:
                right.metric("抓新聞所用時間", elapsed_text)

            stage_rows = [
                {"執行階段": item["label"], "所用時間": format_elapsed(item["seconds"]), "狀態": "已完成"}
                for item in job.get("stage_times", [])
            ]
            current_stage = job.get("current_stage")
            current_stage_started = job.get("current_stage_started")
            if current_stage and current_stage_started is not None:
                stage_rows.append({
                    "執行階段": current_stage,
                    "所用時間": format_elapsed(time.monotonic() - float(current_stage_started)),
                    "狀態": "執行中" if status == "running" else "停止中",
                })
            if stage_rows and not SIMPLE_SITE:
                st.markdown("#### 各階段所用時間")
                st.caption("顯示本次任務目前執行到哪個階段，以及各階段實際耗時，可用來找出速度異常的步驟。")
                st.dataframe(pd.DataFrame(stage_rows), hide_index=True, use_container_width=True)

            if job.get("method") == "方法一＋方法二" and not SIMPLE_SITE:
                st.markdown("#### 方法別監控")
                st.caption("分別顯示方法一與方法二的執行狀態與新聞筆數，方便確認兩種方法是否都有正常完成。")
                monitor_columns = st.columns(2)
                status_labels = {
                    "pending": "⏳ 等待中", "running": "🔄 執行中", "success": "✅ 已完成",
                    "failed": "❌ 執行異常", "stopped": "⏹️ 已停止", "skipped": "— 本次未執行",
                }

                def render_method_monitor(container, title: str, monitor: dict, is_method1: bool) -> None:
                    shown_status = monitor.get("status", "pending")
                    status_text = status_labels.get(shown_status, shown_status)
                    with container:
                        with st.container(border=True):
                            st.markdown(f"**{title}｜{status_text}**")
                            st.caption(str(monitor.get("stage") or "等待更新"))
                            st.metric("新聞筆數", f"{int(monitor.get('rows') or monitor.get('raw_rows') or 0):,}")
                            if monitor.get("first_news") and monitor.get("first_news") != "—":
                                st.caption(
                                    f"實際發布區間：{monitor['first_news']} ～ {monitor.get('last_news', '—')}｜"
                                    f"區間外：{int(monitor.get('outside_window') or 0):,} 筆"
                                )
                            if is_method1:
                                counts = {source: count for source, count in (monitor.get("source_counts") or {}).items() if source != "MoneyDJ"}
                                if counts:
                                    source_text = "｜".join(f"{source} {int(count):,}" for source, count in sorted(counts.items()))
                                    st.caption(f"來源筆數：{source_text}")
                                else:
                                    st.caption("來源筆數：尚未產生")
                                timings = monitor.get("timings") or {}
                                if timings:
                                    timing_text = "｜".join(
                                        f"{label} {format_elapsed(float(seconds))}"
                                        for label, seconds in timings.items()
                                    )
                                    st.caption(f"方法一階段耗時：{timing_text}")
                            else:
                                st.caption(
                                    f"Google News 批次：{int(monitor.get('batches') or 0):,}｜"
                                    f"金融情緒分析 (FinBERT) ≤ 0：{int(monitor.get('negative_rows') or 0):,}"
                                )
                            if monitor.get("fatal_error"):
                                st.error(str(monitor["fatal_error"]))

                render_method_monitor(monitor_columns[0], "方法一｜多來源", job.get("method1_monitor", {}), True)
                render_method_monitor(monitor_columns[1], "方法二｜Google News＋金融情緒分析 (FinBERT)", job.get("method2_monitor", {}), False)

            button_labels = {
                "running": "停止抓取", "stopping": "停止中…", "success": "✅ 已完成",
                "stopped": "已停止", "failed": "執行失敗",
            }
            stop_clicked = False
            if status != "success":
                stop_col, refresh_col = st.columns(2)
                stop_clicked = stop_col.button(
                    button_labels.get(status, "目前不可用"),
                    type="secondary", use_container_width=True, key="stop_crawl",
                    disabled=status != "running",
                )
                refresh_col.button(
                    "自動更新中…" if status in ("running", "stopping") else "狀態已更新",
                    use_container_width=True,
                    key="refresh_crawl_status",
                    disabled=True,
                )
            if stop_clicked:
                job["status"] = "stopping"
                job["progress_text"] = "正在停止目前工作…"
                job["stop_event"].set()
                status = "stopping"
            status_messages = {
                "running": "抓取正在背景執行，可繼續停留在本頁查看進度。",
                "stopping": "已收到停止要求；尚未開始的工作已取消，目前單一網路請求結束後即停止（通常數秒內）。",
                "success": (
                    f"{str(job.get('summary', '抓取完成')).replace('；FinBERT ≤', '；金融情緒分析 (FinBERT) ≤')}；負面新聞 {job.get('event_rows', 0)} 則；"
                    f"待人工覆核 {job.get('unknown_rows', 0)} 則；無關新聞 {job.get('irrelevant_rows', 0)} 則。"
                    f"標題翻譯失敗 {job.get('translation_failures', 0)} 則。"
                    f"總耗時 {elapsed_text}"
                ),
                "stopped": f"抓取已停止。總耗時 {elapsed_text}；未完成的結果不會覆蓋前次成功紀錄。",
                "failed": f"抓取失敗：{job.get('error', '未知錯誤')}（耗時 {elapsed_text}）",
            }
            status_message = status_messages.get(status, "正在更新執行狀態…")
            if status == "success":
                st.success(f"✅ 全部完成｜{status_message}")
                if int(job.get("translation_failures", 0)) > 0:
                    st.caption("翻譯失敗通常是翻譯服務暫時逾時、限流，或批次回傳格式不完整。系統已自動重試；仍失敗的新聞會保留原始標題，不會刪除，並可到「編輯頁面」補正。")
            elif status == "failed":
                st.error(status_message)
            elif status in ("stopping", "stopped"):
                st.warning(status_message)
            else:
                st.info(status_message)
        if status in ("success", "failed", "stopped") and not job.get("completion_page_synced"):
            job["completion_page_synced"] = True
            st.rerun(scope="app")
        return False

    render_crawl_status()
    saved_crawl = latest_crawl_result()
    if saved_crawl:
        current_job = crawl_job_registry().get("current")
        is_current_result = bool(
            current_job
            and current_job.get("status") == "success"
            and current_job.get("path")
            and Path(current_job["path"]) == saved_crawl["path"]
        )
        st.markdown("### 本次抓取結果" if is_current_result else "### 最近一次抓取結果")
        try:
            saved_start = datetime.fromisoformat(saved_crawl["start_time"]).astimezone(TAIPEI)
            saved_end = datetime.fromisoformat(saved_crawl["end_time"]).astimezone(TAIPEI)
            started = datetime.fromisoformat(saved_crawl["started_at"]).astimezone(TAIPEI)
            finished = datetime.fromisoformat(saved_crawl["finished_at"]).astimezone(TAIPEI)
            period_start_text = f"{saved_start:%m/%d %H:%M}"
            period_end_text = f"{saved_end:%m/%d %H:%M}"
            finished_text = f"{finished:%Y-%m-%d %H:%M}"
            elapsed_text = format_elapsed((finished - started).total_seconds())
        except (TypeError, ValueError):
            period_start_text, period_end_text = "時間紀錄", "無法解析"
            finished_text, elapsed_text = str(saved_crawl["finished_at"] or ""), "—"
        st.markdown(
            f"""
            <div class="previous-result-grid">
              <div class="previous-result-card"><div class="previous-result-label">擷取方法</div><div class="previous-result-value">{saved_crawl['method']}</div></div>
              <div class="previous-result-card"><div class="previous-result-label">新聞筆數</div><div class="previous-result-value">{saved_crawl['rows']:,} 筆</div></div>
              <div class="previous-result-card"><div class="previous-result-label">新聞期間</div><div class="previous-result-value is-period"><span>{period_start_text} ～</span><span>{period_end_text}</span></div></div>
              <div class="previous-result-card"><div class="previous-result-label">抓新聞所用時間</div><div class="previous-result-value">{elapsed_text}</div></div>
              <div class="previous-result-card"><div class="previous-result-label">完成時間</div><div class="previous-result-value is-datetime">{finished_text}</div></div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        file_prefix = "今日" if saved_crawl["is_today"] else "前次"
        download_count = 1 + int(saved_crawl["event_path"] is not None) + int(saved_crawl["finbert_path"] is not None)
        download_columns = st.columns(download_count)
        main_label = "完整整合新聞" if saved_crawl["method"] == "方法一＋方法二" else "新聞"
        if saved_crawl["method"] == "方法一＋方法二":
            main_download_name = saved_crawl["path"].name
        else:
            try:
                download_stamp = datetime.fromisoformat(saved_crawl["end_time"]).astimezone(TAIPEI).strftime("%Y%m%d")
            except (TypeError, ValueError):
                download_stamp = datetime.now(TAIPEI).strftime("%Y%m%d")
            main_download_name = f"新聞爬蟲_{download_stamp}.xlsx"
        download_columns[0].download_button(f"下載{file_prefix}{main_label}", saved_crawl["path"].read_bytes(), main_download_name, use_container_width=True, key="saved_crawl_all")
        column_index = 1
        if saved_crawl["event_path"]:
            download_columns[column_index].download_button(f"下載{file_prefix}負面新聞", saved_crawl["event_path"].read_bytes(), saved_crawl["event_path"].name, use_container_width=True, key="saved_crawl_event")
            column_index += 1
        if saved_crawl["finbert_path"]:
            download_columns[column_index].download_button(f"下載{file_prefix}金融情緒分析 (FinBERT) ≤ 0", saved_crawl["finbert_path"].read_bytes(), saved_crawl["finbert_path"].name, use_container_width=True, key="saved_crawl_finbert")
        if not SIMPLE_SITE and saved_crawl["event_path"] and st.button("下一步：查看風險儀表板 →", type="primary", use_container_width=True):
            st.session_state["_go_dashboard"] = True
            st.rerun()
        if not saved_crawl["is_today"]:
            st.caption("今天尚未完成新的抓取；目前提供的是最近一次成功結果。")
    else:
        st.info("尚無成功抓取紀錄。完成第一次抓取後，結果會保留在這裡，重新整理或重新開啟網站也能下載。")
if page == "風險儀表板":
    page_heading(page)
    files = output_files()
    if not files:
        st.info("目前還沒有可顯示的負面新聞結果。請先到「今日任務」執行任一擷取方法，系統會自動去重並分類。")
    else:
        selected = st.selectbox("查看哪一次結果", files, format_func=lambda path: path.name)
        data = pd.read_excel(selected)
        data = data[pd.to_numeric(data.get("Level"), errors="coerce").between(3, 5, inclusive="both")].copy()
        if data.empty:
            st.info("這次分類結果為 0 筆：沒有新聞命中負面事件規則。可返回今日任務下載完整新聞，或改看其他日期結果。")
            st.stop()
        data["Level"] = pd.to_numeric(data["Level"], errors="coerce").fillna(0).astype(int)
        match = re.search(r"20\d{6}", selected.name)
        data_day = datetime.strptime(match.group(0), "%Y%m%d").strftime("%Y-%m-%d") if match else "未知"
        st.caption(f"資料日期：{data_day}｜檔案更新：{datetime.fromtimestamp(selected.stat().st_mtime):%Y-%m-%d %H:%M}")
        metrics = st.columns(4)
        metrics[0].metric("負面新聞（Level 3～5）", len(data))
        metrics[1].metric("涉及公司", data["Ticker"].replace("", pd.NA).nunique())
        metrics[2].metric("重大事件（Level 4～5）", int((data["Level"] >= 4).sum()))
        metrics[3].metric("持續追蹤", int(data["Action"].fillna("").str.contains("持續追蹤").sum()))
        chart_left, chart_right = st.columns([1.45, 1], gap="large")
        active_rules = load_rules(RULE_PATH.stat().st_mtime_ns)
        event_name_map = {
            str(rule.get("Event_Code") or ""): str(rule.get("事件中文") or rule.get("Event_Code") or "")
            for rule in active_rules
        }
        event_names = data["Event_Code"].fillna("").astype(str).map(event_name_map)
        event_names = event_names.where(event_names.fillna("").str.strip() != "", data["Event_Code"].fillna("未命名事件").astype(str))
        event_chart_source = pd.DataFrame({"事件類型": event_names, "Level": data["Level"]})
        event_chart = (
            event_chart_source[event_chart_source["Level"].isin([4, 5])]
            .groupby(["Level", "事件類型"], dropna=False).size()
            .reset_index(name="新聞筆數")
            .sort_values(["Level", "新聞筆數", "事件類型"], ascending=[False, False, True])
        )
        chart_left.markdown("#### 當日 Level 4～5 負面事件")
        chart_left.caption("先依 Level 5、Level 4 分組，再於各組內依新聞筆數排序；完整列出當日出現的所有重大事件種類。")
        for shown_level, group_label, color, background in (
            (5, "Level 5｜立即關注", "#B91C1C", "#FEF2F2"),
            (4, "Level 4｜高度關注", "#C2410C", "#FFF7ED"),
        ):
            level_chart = event_chart[event_chart["Level"] == shown_level].copy()
            if level_chart.empty:
                chart_left.markdown(
                    f"<div style='border-left:5px solid {color};background:{background};padding:12px 16px;border-radius:10px;margin:10px 0'>"
                    f"<div style='font-weight:750;color:{color}'>{group_label}</div>"
                    f"<div style='color:#64748B;font-size:13px;margin-top:8px'>本日沒有此等級事件</div></div>",
                    unsafe_allow_html=True,
                )
                continue
            maximum = max(int(level_chart["新聞筆數"].max()), 1)
            rows_html = []
            for event_name, count in level_chart[["事件類型", "新聞筆數"]].itertuples(index=False, name=None):
                width = max(7.0, float(count) / maximum * 100)
                rows_html.append(
                    "<div style='display:grid;grid-template-columns:125px minmax(120px,1fr) 54px;align-items:center;gap:12px;margin-top:10px'>"
                    f"<div style='color:#334155;font-size:14px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis' title='{html.escape(str(event_name))}'>{html.escape(str(event_name))}</div>"
                    f"<div style='height:12px;background:#E2E8F0;border-radius:999px;overflow:hidden'><div style='height:100%;width:{width:.1f}%;background:{color};border-radius:999px'></div></div>"
                    f"<div style='text-align:right;font-size:14px;font-weight:750;color:#1E293B'>{int(count):,} 筆</div></div>"
                )
            chart_left.markdown(
                f"<div style='border:1px solid #E2E8F0;border-left:5px solid {color};background:white;padding:13px 16px 15px;border-radius:11px;margin:10px 0 14px'>"
                f"<div style='display:flex;justify-content:space-between;align-items:center'><div style='font-weight:750;color:{color};font-size:16px'>{group_label}</div>"
                f"<div style='color:#64748B;font-size:13px'>{len(level_chart)} 種事件</div></div>{''.join(rows_html)}</div>",
                unsafe_allow_html=True,
            )

        chart_right.markdown("#### 風險等級分布")
        total_events = max(len(data), 1)
        level_cards = [
            (5, "立即關注", "#B91C1C", "#FEF2F2"),
            (4, "高度關注", "#C2410C", "#FFF7ED"),
            (3, "一般關注", "#A16207", "#FEFCE8"),
        ]
        for level, label, color, background in level_cards:
            count = int((data["Level"] == level).sum()) if level > 3 else int((data["Level"] <= 3).sum())
            company_count = int(data.loc[data["Level"] == level if level > 3 else data["Level"] <= 3, "Ticker"].replace("", pd.NA).nunique())
            percentage = count / total_events * 100
            shown_level = f"Level {level}"
            chart_right.markdown(
                f"""
                <div style="background:{background};border-left:6px solid {color};border-radius:10px;
                            padding:14px 16px;margin:0 0 12px 0;display:flex;align-items:center;justify-content:space-between;">
                  <div><div style="font-weight:700;color:{color};font-size:17px;">{shown_level}｜{label}</div>
                       <div style="color:#64748B;font-size:13px;margin-top:3px;">占全部新聞 {percentage:.1f}%</div></div>
                  <div style="font-size:30px;font-weight:750;color:#1E293B;">{count}<span style="font-size:14px;font-weight:500;color:#64748B;"> 筆（{company_count} 家公司）</span></div>
                </div>
                """,
                unsafe_allow_html=True,
            )
        st.markdown("### 新聞明細與篩選")
        filter_columns = st.columns([2, 1, 1])
        search = filter_columns[0].text_input("搜尋公司、Ticker、標題或事件")
        level_filter = filter_columns[1].selectbox("事件等級", ["全部", "Level 5", "Level 4", "Level 3"])
        event_filter_options = ["全部"] + event_names.value_counts().index.astype(str).tolist()
        event_filter = filter_columns[2].selectbox("事件類別", event_filter_options)
        filtered = data.copy()
        if search:
            mask = filtered[["Ticker", "Company", "Title", "Event_type", "Event_Code"]].fillna("").astype(str).apply(lambda column: column.str.contains(search, case=False, regex=False)).any(axis=1)
            filtered = filtered[mask]
        if level_filter == "Level 5":
            filtered = filtered[filtered["Level"] == 5]
        elif level_filter == "Level 4":
            filtered = filtered[filtered["Level"] == 4]
        elif level_filter == "Level 3":
            filtered = filtered[filtered["Level"] == 3]
        if event_filter != "全部":
            filtered = filtered[event_names.reindex(filtered.index).fillna("") == event_filter]
        filtered = filtered.copy()
        filtered["事件名稱"] = event_names.reindex(filtered.index).fillna("")
        dashboard_13f, dashboard_13f_quarter = sec13f_lookup()
        filtered["13F 機構動向"] = filtered["Ticker"].fillna("").astype(str).str.upper().map(dashboard_13f).fillna("— 尚無 13F 資料")
        detail_column_order = [
            "Ticker", "Company", "Level", "事件名稱", "Title_ZH", "Action", "13F 機構動向", "URL", "Published Time",
            "Title", "FinBERT", "Event_Code", "Event_type", "Keyword", "Source",
        ]
        filtered = filtered[[column for column in detail_column_order if column in filtered.columns]]
        st.caption(f"篩選結果：顯示 {len(filtered):,}／{len(data):,} 筆；左側先呈現公司、等級、事件與中文標題，可點「開啟」查看原始新聞。")
        st.dataframe(
            filtered,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Ticker": st.column_config.TextColumn("Ticker", width="small"),
                "Company": st.column_config.TextColumn("公司", width="medium"),
                "Level": st.column_config.NumberColumn("等級", width="small", format="%d"),
                "事件名稱": st.column_config.TextColumn("事件", width="small"),
                "Title_ZH": st.column_config.TextColumn("中文標題", width="large"),
                "Published Time": st.column_config.DatetimeColumn("發布時間", width="medium", format="YYYY-MM-DD HH:mm"),
                "URL": st.column_config.LinkColumn("新聞", display_text="開啟", width="small"),
                "Title": st.column_config.TextColumn("英文標題", width="large"),
                "FinBERT": st.column_config.NumberColumn("金融情緒分析 (FinBERT)", width="small", format="%.3f"),
                "Action": st.column_config.TextColumn("處理建議", width="medium"),
                "13F 機構動向": st.column_config.TextColumn(
                    f"13F 機構動向（{dashboard_13f_quarter}）" if dashboard_13f_quarter else "13F 機構動向",
                    width="medium",
                    help="季度申報的機構持股變化，只作輔助參考，不影響新聞 Level。",
                ),
                "Event_Code": st.column_config.TextColumn("英文事件代碼", width="medium"),
                "Event_type": st.column_config.TextColumn("英文事件類別", width="medium"),
                "Keyword": st.column_config.TextColumn("命中關鍵字", width="medium"),
                "Source": st.column_config.TextColumn("來源", width="small"),
            },
        )
if page == "持續追蹤":
    with sqlite3.connect(DB_PATH) as connection:
        saved_gap_days = connection.execute(
            "SELECT value FROM settings WHERE key='event_group_gap_days'"
        ).fetchone()
    try:
        event_gap_days = max(1, min(3650, int(saved_gap_days[0]))) if saved_gap_days else 14
    except (TypeError, ValueError):
        event_gap_days = 14
    st.markdown(
        f"<div class='method-card'><div class='method-title'>持續追蹤事件分組方式</div>"
        f"<div class='method-desc'>本頁只顯示關鍵字規則中 Action 包含「持續追蹤」的新聞；"
        f"相同公司與事件種類的新聞若相隔不超過 {event_gap_days} 天，就歸在同一事件，"
        f"超過 {event_gap_days} 天則建立新事件。分組天數可在第六頁「設定與說明」調整。</div></div>",
        unsafe_allow_html=True,
    )
    all_data = load_negative_files()
    if all_data.empty:
        st.info("尚無負面新聞資料。")
    else:
        tracked = all_data[all_data["Action"].fillna("").str.contains("持續追蹤")].copy()
        tracked = assign_event_batches(tracked)
        period_column, search_column = st.columns([1, 2])
        period = period_column.selectbox(
            "查看期間", ["最近 7 天", "最近 14 天", "最近 30 天", "全部日期"]
        )
        search = search_column.text_input("搜尋公司或 Ticker", placeholder="例如：NVIDIA、NVDA")
        tracked["日期值"] = pd.to_datetime(tracked["事件日期"], errors="coerce")
        tracking_days = {"最近 7 天": 7, "最近 14 天": 14, "最近 30 天": 30}.get(period)
        if tracking_days:
            latest_day = tracked["日期值"].max()
            tracked = tracked[tracked["日期值"] >= latest_day - pd.Timedelta(days=tracking_days - 1)]
        if search:
            mask = tracked[["Ticker", "Company"]].fillna("").astype(str).apply(lambda column: column.str.contains(search, case=False, regex=False)).any(axis=1)
            tracked = tracked[mask]
        if tracked.empty:
            st.warning("此期間或搜尋條件沒有持續追蹤公司。")
            st.stop()
        else:
            tracked["Level"] = pd.to_numeric(tracked["Level"], errors="coerce").fillna(0).astype(int)
            tracked["發布時間值"] = pd.to_datetime(tracked["Published Time"], errors="coerce")
            tracked["顯示標題"] = tracked["Title_ZH"].fillna("").astype(str)
            tracked["顯示標題"] = tracked["顯示標題"].where(
                tracked["顯示標題"].str.strip() != "",
                tracked["Title"].fillna("").astype(str),
            )
            tracked = tracked.sort_values(["日期值", "Published Time"])
            group_keys = ["Ticker", "Company", "Event_ID", "Event_Code", "Event_type"]
            grouped = tracked.groupby(group_keys, dropna=False)
            summary = grouped.agg(
                首次列入=("事件日期", "min"),
                最近列入=("事件日期", "max"),
                出現天數=("事件日期", "nunique"),
                新聞篇數=("Title", "count"),
                最高等級=("Level", "max"),
                最近發布時間值=("發布時間值", "max"),
                事件名稱=("事件中文", "first"),
            ).reset_index()
            latest_news = tracked.groupby(group_keys, as_index=False).tail(1)[group_keys + ["顯示標題", "URL"]].rename(columns={"顯示標題": "最新標題", "URL": "最新新聞"})
            summary = summary.merge(latest_news, on=group_keys, how="left").rename(columns={"Event_ID": "事件編號", "Event_Code": "事件代碼", "Event_type": "事件類別"})
            summary["事件名稱"] = summary["事件名稱"].fillna("").astype(str)
            summary["事件名稱"] = summary["事件名稱"].where(summary["事件名稱"].str.strip() != "", summary["事件代碼"])
            first_dates = pd.to_datetime(summary["首次列入"], errors="coerce")
            last_dates = pd.to_datetime(summary["最近列入"], errors="coerce")
            today_value = pd.Timestamp(datetime.now(TAIPEI).date())

            def event_status_labels(row_index: int) -> str:
                labels = []
                if pd.notna(first_dates.iloc[row_index]) and first_dates.iloc[row_index].normalize() == today_value:
                    labels.append("🔵 今日新增")
                if int(summary.at[row_index, "出現天數"]) >= 2:
                    labels.append("🟠 持續發展")
                return "｜".join(labels) if labels else "—"

            summary.insert(3, "事件狀態", [event_status_labels(index) for index in range(len(summary))])
            summary["今日新增排序"] = (first_dates.dt.normalize() == today_value).astype(int)
            summary = summary.sort_values(
                ["最高等級", "今日新增排序", "出現天數", "最近發布時間值"],
                ascending=[False, False, False, False],
            ).reset_index(drop=True)
            summary["首次列入"] = pd.to_datetime(summary["首次列入"], errors="coerce").dt.strftime("%Y-%m-%d")
            summary["最近列入"] = pd.to_datetime(summary["最近列入"], errors="coerce").dt.strftime("%Y-%m-%d")
            summary["最近發布時間"] = pd.to_datetime(summary.pop("最近發布時間值"), errors="coerce").dt.strftime("%Y-%m-%d %H:%M")
            summary = summary.drop(columns="今日新增排序")
            tracking_column_order = [
                "Ticker", "Company", "事件狀態", "最高等級", "事件代碼", "事件類別",
                "最近列入", "首次列入", "出現天數", "新聞篇數", "最近發布時間",
                "事件名稱", "最新標題", "最新新聞", "事件編號",
            ]
            summary = summary[tracking_column_order]
            first_day, last_day = tracked["日期值"].min(), tracked["日期值"].max()
            date_range = (
                f"{first_day:%Y-%m-%d}"
                if first_day == last_day
                else f"{first_day:%Y-%m-%d} ～ {last_day:%m-%d}"
            )
            st.markdown(
                f"""
                <div class="tracking-metric-grid">
                  <div class="tracking-metric-card"><div class="tracking-metric-label">持續追蹤事件</div><div class="tracking-metric-value">{len(summary):,}</div></div>
                  <div class="tracking-metric-card"><div class="tracking-metric-label">跨日持續事件</div><div class="tracking-metric-value">{int((summary["出現天數"] >= 2).sum()):,}</div></div>
                  <div class="tracking-metric-card"><div class="tracking-metric-label">累計新聞</div><div class="tracking-metric-value">{len(tracked):,} 篇</div></div>
                  <div class="tracking-metric-card"><div class="tracking-metric-label">資料日期</div><div class="tracking-metric-value is-date">{date_range}</div></div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            st.markdown("### 追蹤公司清單")
            st.caption("本表列出關鍵字規則中 Action 包含「持續追蹤」的事件。單一查看時，點下一列會直接取代上一列；合併查看可不限數量勾選同一公司的多個事件一起閱讀，完成後可按「清除選取」一次取消所有勾選。")
            st.markdown(
                """
                <div class="status-legend">
                  <span>🔵 今日新增：首次出現日是今天</span>
                  <span>🟠 持續發展：至少在兩個不同發布日出現</span>
                </div>
                """,
                unsafe_allow_html=True,
            )
            tracking_table = summary[[
                "Ticker", "Company", "事件狀態", "最高等級", "事件名稱",
                "最近列入", "出現天數", "新聞篇數", "最新標題", "最新新聞",
                "事件編號", "事件代碼", "事件類別", "首次列入",
            ]]
            view_mode = st.radio(
                "事件查看方式",
                ["單一查看", "合併查看"],
                horizontal=True,
                key="tracking_view_mode",
            )
            is_multi_view = view_mode.startswith("合併查看")
            if "tracking_selection_version" not in st.session_state:
                st.session_state.tracking_selection_version = 0
            table_event = st.dataframe(
                tracking_table,
                use_container_width=True,
                hide_index=True,
                on_select="rerun",
                selection_mode="multi-row" if is_multi_view else "single-row",
                key=f"tracking_company_table_{'multi' if is_multi_view else 'single'}_{st.session_state.tracking_selection_version}",
                column_config={
                    "Ticker": st.column_config.TextColumn("Ticker", width="small"),
                    "Company": st.column_config.TextColumn("Company", width="medium"),
                    "事件狀態": st.column_config.TextColumn("事件狀態", width="medium"),
                    "最高等級": st.column_config.NumberColumn("最高等級", width="small", format="%d"),
                    "事件名稱": st.column_config.TextColumn("事件名稱", width="medium"),
                    "最近列入": st.column_config.TextColumn("最近列入", width="small"),
                    "出現天數": st.column_config.NumberColumn("出現天數", width="small", format="%d"),
                    "新聞篇數": st.column_config.NumberColumn("新聞篇數", width="small", format="%d"),
                    "最新標題": st.column_config.TextColumn("最新標題", width="medium"),
                    "最新新聞": st.column_config.LinkColumn("最新新聞", display_text="開啟新聞", width="small"),
                    "事件編號": st.column_config.TextColumn("事件編號", width="medium"),
                    "事件代碼": st.column_config.TextColumn("英文事件代碼", width="medium"),
                    "事件類別": st.column_config.TextColumn("英文事件類別", width="small"),
                    "首次列入": st.column_config.TextColumn("首次列入", width="small"),
                },
            )
            selected_rows = table_event.selection.rows
            action_columns = st.columns([2, 1, 1] if is_multi_view else [2, 1])
            company_options = [f"{row.Ticker}｜{row.Company}｜{row.事件編號}" for row in summary.itertuples()]
            quick_company = action_columns[0].selectbox("快速查看公司詳情", ["請選擇"] + company_options)
            action_columns[1].download_button(
                "下載目前追蹤名單 CSV",
                summary.to_csv(index=False).encode("utf-8-sig"),
                f"美股持續追蹤名單_{last_day:%Y%m%d}.csv",
                mime="text/csv",
                use_container_width=True,
            )
            if is_multi_view and action_columns[2].button("清除選取", use_container_width=True):
                st.session_state.tracking_selection_version += 1
                st.rerun()
            selected_companies = pd.DataFrame()
            if selected_rows:
                selected_companies = summary.iloc[selected_rows].copy()
            elif quick_company != "請選擇":
                quick_ticker, quick_name, quick_event_id = quick_company.split("｜", 2)
                selected_companies = summary[(summary["Ticker"].astype(str) == quick_ticker) & (summary["Company"].astype(str) == quick_name) & (summary["事件編號"].astype(str) == quick_event_id)].head(1)
            if selected_companies.empty:
                st.info("請點選表格中的事件列，或使用「快速查看公司詳情」。")
                st.stop()
            if selected_companies["Ticker"].astype(str).nunique() > 1 or selected_companies["Company"].astype(str).nunique() > 1:
                st.warning("合併查看只能選擇同一公司的事件，請調整勾選。")
                st.stop()
            ticker = str(selected_companies.iloc[0]["Ticker"])
            company = str(selected_companies.iloc[0]["Company"])
            selected_event_ids = selected_companies["事件編號"].astype(str).tolist()
            company_news = tracked[
                (tracked["Ticker"].astype(str) == ticker)
                & (tracked["Company"].astype(str) == company)
                & (tracked["Event_ID"].astype(str).isin(selected_event_ids))
            ].copy()
            event_names_for_title = selected_companies["事件名稱"].fillna("").astype(str).drop_duplicates().tolist()
            shown_event_names = "＋".join(name for name in event_names_for_title if name) or "事件明細"
            st.markdown(f"### {ticker}｜{company}｜{shown_event_names}")
            detail_columns = st.columns(4)
            selected_first = pd.to_datetime(selected_companies["首次列入"], errors="coerce").min()
            selected_last = pd.to_datetime(selected_companies["最近列入"], errors="coerce").max()
            combined_days = pd.to_datetime(company_news["事件日期"], errors="coerce").dt.date.nunique()
            detail_columns[0].metric("首次列入", selected_first.strftime("%Y-%m-%d") if pd.notna(selected_first) else "未知")
            detail_columns[1].metric("最近列入", selected_last.strftime("%Y-%m-%d") if pd.notna(selected_last) else "未知")
            detail_columns[2].metric("出現天數", int(combined_days))
            detail_columns[3].metric("相關新聞", len(company_news))
            st.dataframe(
                company_news[["顯示標題", "Action", "URL", "Published Time", "Level", "Event_type", "Event_Code", "Keyword", "Event_ID"]],
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Event_ID": st.column_config.TextColumn("事件編號"),
                    "顯示標題": st.column_config.TextColumn("新聞標題", width="large"),
                    "Action": st.column_config.TextColumn("處理建議", width="medium"),
                    "URL": st.column_config.LinkColumn("新聞", display_text="開啟"),
                },
            )
        st.markdown("### 事件後股價變化")
        event_day = pd.to_datetime(company_news["事件日期"], errors="coerce").min().date()
        try:
            prices = stock_prices(ticker, (event_day - timedelta(days=30)).isoformat(), (date.today() + timedelta(days=15)).isoformat())
            later = prices[prices["Date"] >= event_day].reset_index(drop=True)
            if later.empty:
                st.info(f"事件日期為 {event_day:%Y-%m-%d}，目前尚未取得事件當日或之後的收盤價；有資料後才會顯示走勢與報酬。")
            elif len(later) == 1:
                base_date = pd.to_datetime(later.iloc[0]["Date"]).date()
                st.info(f"事件基準交易日為 {base_date:%Y-%m-%d}，目前尚未有下一個交易日的收盤價；有新交易日資料後才會顯示走勢與報酬。")
            else:
                base = later.iloc[0]["Close"]
                base_date = pd.to_datetime(later.iloc[0]["Date"]).date()
                price_min = float(prices["Close"].min())
                price_max = float(prices["Close"].max())
                price_range = price_max - price_min
                axis_padding = price_range * 0.12 if price_range > 0 else max(abs(price_min) * 0.02, 1.0)
                price_line = (
                    alt.Chart(prices)
                    .mark_line(color="#1769d2", strokeWidth=3)
                    .encode(
                        x=alt.X(
                            "Date:T",
                            title=None,
                            axis=alt.Axis(format="%m/%d", labelAngle=0, tickCount="day"),
                        ),
                        y=alt.Y(
                            "Close:Q",
                            title="收盤價",
                            scale=alt.Scale(domain=[price_min - axis_padding, price_max + axis_padding], zero=False),
                            axis=alt.Axis(format=".2f"),
                        ),
                        tooltip=[
                            alt.Tooltip("Date:T", title="日期", format="%Y-%m-%d"),
                            alt.Tooltip("Close:Q", title="收盤價", format=".2f"),
                        ],
                    )
                )
                event_marker_data = pd.DataFrame({"Date": [pd.Timestamp(base_date)]})
                event_marker = (
                    alt.Chart(event_marker_data)
                    .mark_rule(color="#ef4444", strokeWidth=2, strokeDash=[6, 4])
                    .encode(x="Date:T")
                )
                price_chart = (price_line + event_marker).properties(height=330).interactive(bind_y=False)
                st.caption("紅色虛線為事件基準交易日；左側是事件前走勢，右側是事件後走勢。下方比較事件後第 1、3、5、10 個交易日的價格變化。")
                st.altair_chart(price_chart, use_container_width=True)
                columns = st.columns(4)
                for column, horizon in zip(columns, (1, 3, 5, 10)):
                    if len(later) <= horizon:
                        column.metric(f"{horizon} 交易日（{base_date:%m/%d}～尚無）", "尚無資料")
                    else:
                        target_date = pd.to_datetime(later.iloc[horizon]["Date"]).date()
                        value = (later.iloc[horizon]["Close"] / base - 1) * 100
                        column.metric(f"{horizon} 交易日（{base_date:%m/%d}～{target_date:%m/%d}）", f"{value:+.2f}%")
        except Exception:
            st.warning("股價暫時無法取得，請稍後再試。事件與新聞資料不受影響。")
if page == "設定與說明":
    page_heading(page)
    st.markdown("### 13F 機構持股資料")
    st.caption("上傳 SEC13F 專案產出的完整 Excel；套用後第三頁與第四頁會使用這個版本，直到下次更新。")
    current_13f = latest_sec13f_path()
    if current_13f is not None:
        try:
            current_13f_summary, _, current_13f_log = load_sec13f_file(str(current_13f), current_13f.stat().st_mtime_ns)
            current_quarter = str(current_13f_summary["Quarter"].dropna().iloc[0]) if current_13f_summary["Quarter"].notna().any() else "未知"
            current_ok = int((current_13f_log.get("Status", pd.Series(dtype=str)).astype(str).str.upper() == "OK").sum()) if not current_13f_log.empty else 0
            st.success(f"目前版本：{current_quarter}｜{current_13f_summary['Ticker'].nunique():,} 檔股票｜{current_ok:,} 家機構正常完成｜更新 {datetime.fromtimestamp(current_13f.stat().st_mtime):%Y-%m-%d %H:%M}")
        except Exception as exc:
            st.warning(f"目前 13F 檔案無法讀取：{exc}")
    else:
        st.info("目前尚未套用 13F 成果檔。")
    sec13f_upload = st.file_uploader("上傳新版 13F Excel", type=["xlsx"], key="sec13f_result_upload")
    if st.button("檢查並套用 13F 資料", type="primary", use_container_width=True, disabled=sec13f_upload is None, key="apply_sec13f_result"):
        try:
            checked_summary, _, checked_log = parse_sec13f_workbook(io.BytesIO(sec13f_upload.getvalue()))
            checked_quarter = str(checked_summary["Quarter"].dropna().iloc[0]) if checked_summary["Quarter"].notna().any() else "未知"
            SEC13F_DIR.mkdir(parents=True, exist_ok=True)
            if SEC13F_PATH.is_file():
                backup_dir = SEC13F_DIR / "history"
                backup_dir.mkdir(parents=True, exist_ok=True)
                backup_name = f"13F_{datetime.now(TAIPEI):%Y%m%d_%H%M%S}.xlsx"
                (backup_dir / backup_name).write_bytes(SEC13F_PATH.read_bytes())
            SEC13F_PATH.write_bytes(sec13f_upload.getvalue())
            load_sec13f_file.clear()
            ok_count = int((checked_log.get("Status", pd.Series(dtype=str)).astype(str).str.upper() == "OK").sum()) if not checked_log.empty else 0
            st.success(f"已套用 {checked_quarter}：{checked_summary['Ticker'].nunique():,} 檔股票、{ok_count:,} 家機構正常完成。")
            st.rerun()
        except Exception as exc:
            st.error(f"13F 檔案無法套用：{exc}")
    st.divider()
    st.markdown("### 持續追蹤事件分組設定")
    with sqlite3.connect(DB_PATH) as connection:
        saved_gap_days = connection.execute(
            "SELECT value FROM settings WHERE key='event_group_gap_days'"
        ).fetchone()
    try:
        current_gap_days = max(1, min(3650, int(saved_gap_days[0]))) if saved_gap_days else 14
    except (TypeError, ValueError):
        current_gap_days = 14
    event_gap_days_setting = st.number_input(
        "建立新事件的間隔天數",
        min_value=1,
        max_value=3650,
        value=current_gap_days,
        step=1,
        help="同公司、同事件種類的新聞相隔超過此天數後，會建立新的事件編號。預設為 14 天。",
    )
    if st.button("儲存事件分組天數", type="primary", use_container_width=True):
        with sqlite3.connect(DB_PATH) as connection:
            connection.execute(
                "INSERT OR REPLACE INTO settings(key,value) VALUES('event_group_gap_days',?)",
                (str(int(event_gap_days_setting)),),
            )
        st.success(
            f"已設定：相同公司與事件種類相隔超過 {int(event_gap_days_setting)} 天後，建立新事件。"
        )
        st.rerun()
    st.divider()
    st.markdown("### 負面新聞評分關鍵字版本")
    try:
        current_book = load_workbook(RULE_PATH, read_only=True, data_only=True)
        current_rule_count = (
            sum(1 for row in current_book["Event"].iter_rows(min_row=2, values_only=True) if any(value not in (None, "") for value in row))
            if "Event" in current_book.sheetnames else 0
        )
        current_book.close()
    except Exception:
        current_rule_count = 0
    with sqlite3.connect(DB_PATH) as connection:
        saved_rule_name = connection.execute("SELECT value FROM settings WHERE key='rule_original_name'").fetchone()
        saved_rule_time = connection.execute("SELECT value FROM settings WHERE key='rule_updated_at'").fetchone()
    current_rule_name = saved_rule_name[0] if saved_rule_name else RULE_PATH.name
    current_rule_time = saved_rule_time[0] if saved_rule_time else datetime.fromtimestamp(RULE_PATH.stat().st_mtime, TAIPEI).strftime("%Y-%m-%d %H:%M:%S")
    st.markdown(
        f"""
        <div class="settings-info-grid">
          <div class="settings-info-card">
            <div class="settings-info-label">目前版本</div>
            <div class="settings-info-value">{html.escape(str(current_rule_name))}</div>
          </div>
          <div class="settings-info-card">
            <div class="settings-info-label">規則筆數</div>
            <div class="settings-info-value is-number">{current_rule_count:,} 筆</div>
          </div>
          <div class="settings-info-card">
            <div class="settings-info-label">套用時間</div>
            <div class="settings-info-value">{html.escape(str(current_rule_time))}</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.info("目前版本會用於之後所有新聞分類；只有再次上傳並套用新版時才會更換。")
    rule_choice = st.radio(
        "關鍵字版本設定",
        ["使用目前版本", "自行上傳最新版"],
        index=0,
        horizontal=True,
    )
    st.download_button("下載目前關鍵字 Excel", RULE_PATH.read_bytes(), current_rule_name, use_container_width=True)
    replacement = st.file_uploader(
        "上傳最新版關鍵字 Excel",
        type=["xlsx"],
        key="rule_upload",
        disabled=rule_choice != "自行上傳最新版",
    )
    if rule_choice == "使用目前版本":
        st.success(f"目前持續使用：{current_rule_name}")
    if st.button(
        "檢查並套用新版",
        type="primary",
        disabled=rule_choice != "自行上傳最新版" or not replacement,
        use_container_width=True,
    ):
        try:
            test_path = UPLOAD_DIR / "Parameter_Event_pending.xlsx"
            test_path.write_bytes(replacement.getvalue())
            workbook = load_workbook(test_path, read_only=True, data_only=True)
            if "Event" not in workbook.sheetnames:
                raise ValueError("缺少 Event 工作表")
            if "主要網站" not in workbook.sheetnames:
                raise ValueError("缺少「主要網站」工作表")
            headers = [str(cell.value or "").strip() for cell in workbook["Event"][1]]
            missing_headers = [name for name in ("Event_Code", "Keyword", "Level", "Action") if name not in headers]
            if missing_headers:
                raise ValueError(f"Event 工作表缺少必要欄位：{'、'.join(missing_headers)}")
            new_rule_count = sum(
                1 for row in workbook["Event"].iter_rows(min_row=2, values_only=True)
                if any(value not in (None, "") for value in row)
            )
            if new_rule_count == 0:
                raise ValueError("Event 工作表沒有規則資料")
            source_headers = [str(cell.value or "").strip() for cell in workbook["主要網站"][1]]
            if "Source" not in source_headers:
                raise ValueError("「主要網站」工作表缺少 Source 欄位")
            event_code_index = headers.index("Event_Code")
            event_codes = {
                str(row[event_code_index] or "").strip()
                for row in workbook["Event"].iter_rows(min_row=2, values_only=True)
            }
            required_system_codes = {"UNKNOWN", "NOT_RELEVANT", "NON_CORE_SOURCE"}
            missing_system_codes = sorted(required_system_codes - event_codes)
            if missing_system_codes:
                raise ValueError(f"Event 工作表缺少系統事件：{'、'.join(missing_system_codes)}")
            workbook.close()
            backup_dir = DATA_ROOT / "backups"
            backup_dir.mkdir(parents=True, exist_ok=True)
            backup_stamp = datetime.now(TAIPEI).strftime("%Y%m%d_%H%M%S")
            (backup_dir / f"Parameter_Event_{backup_stamp}.xlsx").write_bytes(RULE_PATH.read_bytes())
            RULE_PATH.write_bytes(test_path.read_bytes())
            applied_at = datetime.now(TAIPEI).strftime("%Y-%m-%d %H:%M:%S")
            with sqlite3.connect(DB_PATH) as connection:
                connection.execute("INSERT OR REPLACE INTO settings(key,value) VALUES('rule_original_name',?)", (replacement.name,))
                connection.execute("INSERT OR REPLACE INTO settings(key,value) VALUES('rule_updated_at',?)", (applied_at,))
            load_rules.clear()
            load_core_sources.clear()
            st.success(f"新版已套用：{replacement.name}，共 {new_rule_count:,} 筆規則。之後會持續使用此版本，直到下次更新。")
            st.rerun()
        except Exception as exc:
            st.error(f"新版關鍵字檔無法套用：{exc}")
    st.divider()
    st.markdown("### 版本恢復")
    if BACKUP_PATH.is_file():
        st.caption("已保存本次需求修改前的完整 V4 主程式；若需要，可恢復到修改前狀態。新聞資料與人工紀錄不會刪除。")
        confirm_restore = st.checkbox("我確認要恢復成本次修改前版本")
        if st.button("恢復本次修改前版本", disabled=not confirm_restore):
            try:
                current_path = Path(__file__).resolve()
                safety_copy = BACKUP_PATH.parent / "美股負面新聞系統_恢復前自動備份.py"
                safety_copy.write_bytes(current_path.read_bytes())
                current_path.write_bytes(BACKUP_PATH.read_bytes())
                st.success("已恢復本次修改前版本，網站將重新載入。")
                st.rerun()
            except Exception as exc:
                st.error(f"版本恢復失敗：{exc}")
    else:
        st.warning("找不到改版前備份，為安全起見不提供恢復按鈕。")
