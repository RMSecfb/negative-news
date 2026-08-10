from __future__ import annotations

import base64
import io
import os
import re
import shutil
import sqlite3
import tempfile
import threading
import time
import unicodedata
import urllib.parse
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta, timezone
from difflib import SequenceMatcher
from email.utils import parsedate_to_datetime
from pathlib import Path
from urllib.parse import urljoin

import pandas as pd
import requests
import streamlit as st
import streamlit.components.v1 as components
from streamlit.runtime.scriptrunner import RerunData, get_script_run_ctx
from streamlit.runtime import Runtime
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# 區塊：基本設定與路徑常數
# 說明：本機雙擊執行用的啟動器（開子行程＋開瀏覽器）已移除，
#       Streamlit Cloud 部署不需要，直接用 `streamlit run` 執行本檔案。
# ============================================================
STANDALONE_DIR = Path(__file__).resolve().parent
APP_TITLE = "美股負面新聞整合中心"
TAIPEI = timezone(timedelta(hours=8))
DATA_ROOT = Path(os.environ.get(
    "FUBON_REBUILT_DATA_DIR",
    str(STANDALONE_DIR / "負面新聞整合資料"),
))
OUTPUT_DIR = DATA_ROOT / "output"
UPLOAD_DIR = DATA_ROOT / "uploads"
CONFIG_DIR = DATA_ROOT / "config"
DB_PATH = DATA_ROOT / "system.db"
RULE_PATH = CONFIG_DIR / "Parameter_Event.xlsx"
DOWJONES_PATH = CONFIG_DIR / "DowJones30.xlsx"
SP500_PATH = CONFIG_DIR / "SP500.xlsx"
CUSTOM_COMPANY_PATH = CONFIG_DIR / "Company_List_Custom.xlsx"
EXPOSURE_PATH = CONFIG_DIR / "Exposure_Custom.xlsx"


# ============================================================
# 區塊：設定檔尋找工具
# 首次啟動時，若 repo 內沒有 Parameter_Event.xlsx／DowJones30.xlsx／SP500.xlsx，
# 會嘗試從程式旁或使用者常見資料夾找一份既有的來用。
# ============================================================
def find_external_workbook(exact_name: str, patterns: tuple[str, ...]) -> Path | None:
    """從程式旁、既有 data/config 與常見個人資料夾尋找設定檔。"""
    search_roots = [
        STANDALONE_DIR,
        STANDALONE_DIR / "data" / "config",
        Path.home() / "Downloads",
        Path.home() / "Desktop",
        Path.home() / "Documents",
    ]
    candidates: list[Path] = []
    for root in search_roots:
        exact = root / exact_name
        if exact.is_file():
            candidates.append(exact)
        if root.is_dir():
            for pattern in patterns:
                candidates.extend(path for path in root.glob(pattern) if path.is_file())
    return max(candidates, key=lambda path: path.stat().st_mtime) if candidates else None

NEWS_COLUMNS = ["Ticker", "Company", "Published Time", "Title", "Source", "URL", "FinBERT"]
NEGATIVE_OUTPUT_COLUMNS = [
    "Published Time", "Ticker", "Company", "Title", "Title_ZH", "FinBERT",
    "Event_type", "Event_Code", "事件中文", "Level", "Action", "Keyword", "URL", "Source",
]

# ============================================================
# 區塊：SQLite 初始化與「執行紀錄」存取
# runs 資料表記錄每次抓取的起訖時間、結果檔案路徑；
# Dow Jones 30／S&P 500 名單直接讀取 GitHub repo 內的 Excel 檔，不再需要另外記錄版本。
# ============================================================
def initialize() -> None:
    for folder in (DATA_ROOT, OUTPUT_DIR, UPLOAD_DIR, CONFIG_DIR):
        folder.mkdir(parents=True, exist_ok=True)
    rule_is_incomplete = not RULE_PATH.exists()
    if RULE_PATH.exists():
        try:
            workbook = load_workbook(RULE_PATH, read_only=True, data_only=True)
            rule_is_incomplete = "主要網站" not in workbook.sheetnames
            workbook.close()
        except Exception:
            rule_is_incomplete = True
    if rule_is_incomplete:
        external_rule = find_external_workbook("Parameter_Event.xlsx", ("Parameter_Event*.xlsx",))
        if external_rule and external_rule.resolve() != RULE_PATH.resolve():
            shutil.copy2(external_rule, RULE_PATH)
    if not DOWJONES_PATH.exists():
        external_dj30 = find_external_workbook(
            "DowJones30.xlsx",
            ("*Dow*Jones*30*.xlsx", "*DJ30*.xlsx", "*道瓊*.xlsx"),
        )
        if external_dj30:
            shutil.copy2(external_dj30, DOWJONES_PATH)
    if not SP500_PATH.exists():
        external_sp500 = find_external_workbook(
            "SP500.xlsx",
            ("*SP500*.xlsx", "*S&P*500*.xlsx", "*標普*.xlsx"),
        )
        if external_sp500:
            shutil.copy2(external_sp500, SP500_PATH)
    with sqlite3.connect(DB_PATH) as connection:
        connection.executescript("""
            CREATE TABLE IF NOT EXISTS runs (
                id INTEGER PRIMARY KEY, started_at TEXT, finished_at TEXT, status TEXT,
                start_time TEXT, end_time TEXT, universe TEXT, rows INTEGER, output_path TEXT, message TEXT
            );
        """)
    if not RULE_PATH.exists():
        st.error(
            "找不到 Parameter_Event.xlsx：請把這份規則檔放在程式旁（GitHub repo 根目錄），"
            "或放進 Downloads / Desktop / Documents 供程式自動尋找後再重新整理網站。"
        )
        st.stop()
    missing_company_files = [
        name for name, path in (("DowJones30.xlsx", DOWJONES_PATH), ("SP500.xlsx", SP500_PATH))
        if not path.exists()
    ]
    if missing_company_files:
        st.error(
            f"找不到公司名單檔案：{'、'.join(missing_company_files)}。"
            "請把這些檔案放在程式旁（GitHub repo 根目錄），或放進 Downloads / Desktop / Documents 供程式自動尋找後再重新整理網站。"
        )
        st.stop()


def last_success_end() -> datetime | None:
    with sqlite3.connect(DB_PATH) as connection:
        row = connection.execute(
            "SELECT end_time FROM runs WHERE status='success' ORDER BY id DESC LIMIT 1"
        ).fetchone()
    if not row or not row[0]:
        return None
    try:
        return datetime.fromisoformat(row[0]).astimezone(TAIPEI)
    except ValueError:
        return None


def latest_crawl_result(today_only: bool = False) -> dict | None:
    """從正式執行紀錄找回結果，不依賴目前瀏覽器的 session。"""
    with sqlite3.connect(DB_PATH) as connection:
        rows = connection.execute(
            """SELECT started_at,finished_at,start_time,end_time,universe,rows,output_path,message
               FROM runs WHERE status='success' AND (universe LIKE '方法一｜%' OR universe LIKE '方法二｜%' OR universe LIKE '方法一＋方法二｜%')
               ORDER BY id DESC"""
        ).fetchall()
    today = datetime.now(TAIPEI).date()
    for row in rows:
        try:
            end_at = datetime.fromisoformat(row[3]).astimezone(TAIPEI)
        except (TypeError, ValueError):
            continue
        if today_only and end_at.date() != today:
            continue
        output_path = Path(str(row[6]))
        if not output_path.is_file():
            continue
        method = str(row[4]).split("｜", 1)[0]
        stamp = end_at.strftime("%Y%m%d")
        event_candidate = OUTPUT_DIR / f"美股_{stamp}_負面新聞爬蟲.xlsx"
        event_path = event_candidate if event_candidate.is_file() else None
        finbert_path = None
        if method in ("方法二", "方法一＋方法二"):
            finbert_name = output_path.name.replace("MergeNews_", "MergeNews_FinBERT_").replace("Combined_News_", "Combined_FinBERT_News_").replace("All_News_", "FinBERT_News_")
            candidate = output_path.with_name(finbert_name)
            finbert_path = candidate if candidate.is_file() else None
        return {
            "started_at": row[0], "finished_at": row[1], "start_time": row[2], "end_time": row[3],
            "universe": row[4], "rows": int(row[5] or 0), "path": output_path,
            "event_path": event_path, "finbert_path": finbert_path, "message": row[7] or "", "method": method,
            "is_today": end_at.date() == today,
        }
    return None


def create_run(start: datetime, end: datetime, universe: str) -> int:
    with sqlite3.connect(DB_PATH) as connection:
        cursor = connection.execute(
            "INSERT INTO runs(started_at,status,start_time,end_time,universe,rows,message) VALUES(?,?,?,?,?,?,?)",
            (datetime.now(TAIPEI).isoformat(timespec="seconds"), "running", start.isoformat(), end.isoformat(), universe, 0, ""),
        )
        return int(cursor.lastrowid)


def finish_run(run_id: int, status: str, rows: int = 0, output_path: str = "", message: str = "") -> None:
    with sqlite3.connect(DB_PATH) as connection:
        connection.execute(
            "UPDATE runs SET finished_at=?,status=?,rows=?,output_path=?,message=? WHERE id=?",
            (datetime.now(TAIPEI).isoformat(timespec="seconds"), status, rows, output_path, message, run_id),
        )


# ============================================================
# 區塊：背景執行緒與進度追蹤
# 讓抓取新聞的過程在背景執行緒跑，網頁可即時顯示進度、
# 各階段耗時，並支援使用者中途按「停止執行」。
# ============================================================
class CrawlCancelled(Exception):
    """使用者主動停止執行。"""


class BackgroundProgress:
    def __init__(self, job: dict):
        self.job = job

    def progress(self, value=0, text=""):
        if self.job["stop_event"].is_set():
            raise CrawlCancelled("使用者已停止執行")
        self.job["progress"] = max(0.0, min(float(value or 0), 1.0))
        if text:
            self.job["progress_text"] = str(text)
        return self


@st.cache_resource
def crawl_job_registry() -> dict:
    return {}


def format_elapsed(seconds: float) -> str:
    total = max(0, int(seconds))
    hours, remainder = divmod(total, 3600)
    minutes, secs = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"


def begin_job_stage(job: dict, label: str) -> None:
    """切換主要執行階段，並保留前一階段的花費時間。"""
    now = time.monotonic()
    previous = job.get("current_stage")
    previous_started = job.get("current_stage_started")
    if previous and previous_started is not None:
        job.setdefault("stage_times", []).append({
            "label": previous,
            "seconds": max(0.0, now - float(previous_started)),
        })
    job["current_stage"] = label
    job["current_stage_started"] = now


def finish_job_stage(job: dict) -> None:
    now = time.monotonic()
    current = job.pop("current_stage", None)
    started = job.pop("current_stage_started", None)
    if current and started is not None:
        job.setdefault("stage_times", []).append({
            "label": current,
            "seconds": max(0.0, now - float(started)),
        })


def run_crawl_job(job: dict, method_name: str, universe: str, companies: pd.DataFrame,
                  start_dt: datetime, end_dt: datetime) -> None:
    run_id = job["run_id"]
    progress = BackgroundProgress(job)
    active_method_key = ""
    try:
        stamp = end_dt.strftime("%Y%m%d")
        if method_name == "方法一":
            begin_job_stage(job, "方法一｜多來源新聞擷取")
            active_method_key = "method1_monitor"
            job[active_method_key].update(status="running", stage="建立多來源新聞池")
            progress.progress(0, text="方法一正在建立新聞池")
            frame, all_errors, source_counts = fetch_method1_news(companies, start_dt, end_dt, progress)
            job[active_method_key].update(
                status="success", stage="完成", rows=len(frame), errors=list(all_errors),
                source_counts=dict(source_counts),
            )
            path = OUTPUT_DIR / f"RawNews_{stamp}.xlsx"
            begin_job_stage(job, "產生新聞檔案")
            source_summary = pd.DataFrame([{"來源": source, "新聞則數": count} for source, count in source_counts.items()])
            progress.progress(0.94, text="正在產生新聞檔案")
            write_excel(path, {"新聞": frame, "來源摘要": source_summary})
            job["summary"] = f"方法一已完成：取得 {len(frame)} 則新聞；來源異常 {len(all_errors)} 項"
        elif method_name == "方法二":
            begin_job_stage(job, "方法二｜Google News 擷取")
            active_method_key = "method2_monitor"
            job[active_method_key].update(status="running", stage="分批抓取 Google News")
            progress.progress(0, text="方法二正在分批爬取 Google News")
            raw_frame, all_errors, batch_count = fetch_method2_raw(companies, start_dt, end_dt, progress)
            job[active_method_key].update(stage="FinBERT 評分中", raw_rows=len(raw_frame), batches=batch_count, errors=list(all_errors))
            begin_job_stage(job, "FinBERT 新聞評分")
            frame, negative_frame = score_method2_finbert(raw_frame, progress)
            job[active_method_key].update(
                status="success", stage="完成", rows=len(frame), negative_rows=len(negative_frame),
                batches=batch_count, errors=list(all_errors),
            )
            path = OUTPUT_DIR / f"All_News_{stamp}.xlsx"
            finbert_path = OUTPUT_DIR / f"FinBERT_News_{stamp}.xlsx"
            begin_job_stage(job, "產生新聞檔案")
            progress.progress(0.94, text="正在產生新聞檔案")
            write_excel(path, {"All_News": frame})
            write_excel(finbert_path, {"FinBERT_News": negative_frame})
            job["summary"] = f"方法二已完成：{batch_count} 批、爬得 {len(frame)} 則；FinBERT ≤ 0 共 {len(negative_frame)} 則"
        else:
            begin_job_stage(job, "方法一｜多來源新聞擷取")
            active_method_key = "method1_monitor"
            job[active_method_key].update(status="running", stage="建立多來源新聞池")
            progress.progress(0, text="完整整合 1/4｜執行方法一")
            method1_frame, method1_errors, source_counts = fetch_method1_news(companies, start_dt, end_dt, progress)
            job[active_method_key].update(
                status="success", stage="完成", rows=len(method1_frame), errors=list(method1_errors),
                source_counts=dict(source_counts),
            )
            active_method_key = "method2_monitor"
            begin_job_stage(job, "方法二｜Google News 擷取")
            job[active_method_key].update(status="running", stage="分批抓取 Google News")
            progress.progress(0, text="完整整合 2/4｜執行方法二 Google News")
            method2_frame, method2_errors, batch_count = fetch_method2_raw(companies, start_dt, end_dt, progress)
            job[active_method_key].update(
                stage="FinBERT 評分中", raw_rows=len(method2_frame), batches=batch_count,
                errors=list(method2_errors),
            )
            begin_job_stage(job, "合併並刪除重複")
            progress.progress(0.55, text="完整整合 3/4｜合併並去除重複新聞")
            merged_frame, duplicate_log = merge_frames(method1_frame, method2_frame)
            begin_job_stage(job, "FinBERT 新聞評分")
            progress.progress(0.58, text="完整整合 4/4｜對全部整合新聞執行 FinBERT")
            frame, negative_frame = score_method2_finbert(merged_frame, progress)
            job[active_method_key].update(
                status="success", stage="完成", rows=len(method2_frame), merged_rows=len(frame),
                negative_rows=len(negative_frame), batches=batch_count, errors=list(method2_errors),
            )
            all_errors = method1_errors + method2_errors
            path = OUTPUT_DIR / f"MergeNews_{stamp}.xlsx"
            finbert_path = OUTPUT_DIR / f"MergeNews_FinBERT_{stamp}.xlsx"
            source_summary = pd.DataFrame(
                [{"來源": source, "新聞則數": count} for source, count in source_counts.items()]
                + [{"來源": "Google News（方法二）", "新聞則數": len(method2_frame)}]
            )
            begin_job_stage(job, "產生完整整合檔案")
            progress.progress(0.94, text="正在產生完整整合檔案")
            write_excel(path, {"總新聞": frame, "重複紀錄": duplicate_log, "來源摘要": source_summary})
            write_excel(finbert_path, {"FinBERT小於等於0": negative_frame})
            job["summary"] = f"排除重複後共 {len(frame)} 則；FinBERT ≤ 0  {len(negative_frame)} 則"
        begin_job_stage(job, "負面事件分類與中文翻譯")
        progress.progress(0.97, text="正在執行負面事件分類")
        rule_modified_ns = RULE_PATH.stat().st_mtime_ns
        event_frame, unknown_frame, irrelevant_frame = classify_news_sets(
            frame, load_rules(rule_modified_ns), load_core_sources(rule_modified_ns), progress
        )
        negative_translation_failures = int(event_frame.attrs.get("translation_failures", 0))
        unknown_translation_failures = int(unknown_frame.attrs.get("translation_failures", 0))
        irrelevant_translation_failures = int(irrelevant_frame.attrs.get("translation_failures", 0))
        translation_failures = negative_translation_failures + unknown_translation_failures + irrelevant_translation_failures
        event_path = OUTPUT_DIR / f"美股_{stamp}_負面新聞爬蟲.xlsx"
        begin_job_stage(job, "產生負面新聞檔案")
        write_excel(event_path, {
            "負面新聞": event_frame,
            "待人工覆核": unknown_frame,
            "無關新聞": irrelevant_frame,
        })
        history_pushed = append_negative_history_to_github(event_frame, end_dt)
        risk_news_snapshot_pushed = save_risk_news_snapshot_to_github(
            event_frame.assign(已人工覆核=False) if not event_frame.empty else event_frame, end_dt.date()
        )
        finish_run(run_id, "success", len(frame), str(path), "；".join(all_errors))
        job.update(status="success", progress=1.0, progress_text="抓取與分類全部完成",
                   history_pushed=history_pushed,
                   risk_news_snapshot_pushed=risk_news_snapshot_pushed,
                   rows=len(frame), event_rows=len(event_frame), unknown_rows=len(unknown_frame),
                   irrelevant_rows=len(irrelevant_frame), translation_failures=translation_failures,
                   negative_translation_failures=negative_translation_failures,
                   unknown_translation_failures=unknown_translation_failures,
                   irrelevant_translation_failures=irrelevant_translation_failures,
                   path=str(path))
    except CrawlCancelled:
        if active_method_key and job.get(active_method_key, {}).get("status") == "running":
            job[active_method_key].update(status="stopped", stage="使用者停止")
        finish_run(run_id, "stopped", message="使用者主動停止")
        job.update(status="stopped", progress_text="已停止；未完成的檔案不會列為正式結果")
    except Exception as exc:
        if active_method_key:
            job[active_method_key].update(status="failed", stage="執行失敗", fatal_error=str(exc))
        finish_run(run_id, "failed", message=str(exc))
        job.update(status="failed", error=str(exc), progress_text="執行失敗")
    finally:
        finish_job_stage(job)
        job["finished_monotonic"] = time.monotonic()
        job["finished_at"] = datetime.now(TAIPEI).isoformat(timespec="seconds")
        # 背景工作只在結束瞬間通知原頁面完整更新一次；執行期間不輪詢、不刷新。
        script_context = job.get("script_context")
        try:
            runtime = Runtime.instance()
            session_info = runtime._session_mgr.get_session_info(job.get("session_id", ""))
            if session_info is not None:
                session = session_info.session
                session._event_loop.call_soon_threadsafe(session.request_rerun, None)
            elif script_context is not None and script_context.script_requests is not None:
                script_context.script_requests.request_rerun(RerunData())
        except Exception:
            pass


# ============================================================
# 區塊：新聞正規化、去重比對、翻譯、事件分類、Excel輸出
# 這段是整支程式的核心邏輯：不論方法一或方法二抓到的新聞，
# 都會經過這裡統一格式、去除重複、翻譯標題，再依
# Parameter_Event.xlsx 的規則分類成「負面新聞／待人工覆核／無關新聞」。
# ============================================================
def split_terms(value: object) -> list[str]:
    return [item.strip() for item in re.split(r"[;；\n]+", str(value or "")) if item.strip()]


@st.cache_data(ttl=60)
def load_rules(modified_ns: int) -> list[dict]:
    del modified_ns
    workbook = load_workbook(RULE_PATH, read_only=True, data_only=True)
    sheet = workbook["Event"]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    rules = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, values))
        if not str(item.get("Event_Code") or "").strip():
            continue
        item["_keywords"] = split_terms(item.get("Keyword"))
        item["_exclusions"] = split_terms(item.get("Exclude_Keyword"))
        rules.append(item)
    workbook.close()
    return rules


@st.cache_data(ttl=60)
def load_core_sources(modified_ns: int) -> set[str]:
    del modified_ns
    workbook = load_workbook(RULE_PATH, read_only=True, data_only=True)
    if "主要網站" not in workbook.sheetnames:
        workbook.close()
        raise ValueError("目前關鍵字表缺少「主要網站」工作表")
    sheet = workbook["主要網站"]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    if "Source" not in headers:
        workbook.close()
        raise ValueError("「主要網站」工作表缺少 Source 欄位")
    source_index = headers.index("Source")
    sources = {
        str(values[source_index]).strip().casefold()
        for values in sheet.iter_rows(min_row=2, values_only=True)
        if source_index < len(values) and values[source_index] not in (None, "")
    }
    workbook.close()
    return sources


def normalize_news(frame: pd.DataFrame) -> pd.DataFrame:
    # 外部工具常在欄名尾端留下不可見空白；先統一清理再做欄位對應。
    frame = frame.copy()
    frame.columns = [str(column).replace("\u3000", " ").strip() for column in frame.columns]
    aliases = {
        "Ticker": ("Ticker", "股票代號", "標的代碼"),
        "Company": ("Company", "股票名稱", "標的名稱"),
        "Published Time": ("Published Time", "發布時間", "發布日期"),
        "Title": ("Title", "新聞主旨", "主旨"),
        "Source": ("Source", "新聞來源"),
        "URL": ("URL", "新聞網址", "新聞連結"),
        "FinBERT": ("FinBERT",),
    }
    result = pd.DataFrame(index=frame.index)
    for target, names in aliases.items():
        source = next((name for name in names if name in frame.columns), None)
        result[target] = frame[source] if source else ""
    for column in NEWS_COLUMNS:
        result[column] = result[column].fillna("").astype(str).str.strip()
    result["Ticker"] = result["Ticker"].str.replace(r"\.0$", "", regex=True).str.upper()
    result = result[(result["Title"] != "") & (result["URL"] != "")]
    return result.reset_index(drop=True)


TRACKING_QUERY_NAMES = {
    "utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content",
    "from", "source", "ref", "gclid", "fbclid", "mc_cid", "mc_eid",
}


def normalize_url(value: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    try:
        parts = urllib.parse.urlsplit(raw)
        if not parts.scheme or not parts.netloc:
            return raw.rstrip("/").casefold()
        query = [
            (key, val) for key, val in urllib.parse.parse_qsl(parts.query, keep_blank_values=True)
            if key.casefold() not in TRACKING_QUERY_NAMES and not key.casefold().startswith("utm_")
        ]
        return urllib.parse.urlunsplit((
            parts.scheme.casefold(), parts.netloc.casefold(), parts.path.rstrip("/"),
            urllib.parse.urlencode(query), "",
        ))
    except ValueError:
        return raw.rstrip("/").casefold()


def url_signatures(value: str) -> set[str]:
    """EXE 同款：解開 Google 轉址、URL encoding 與常見內嵌目標網址。"""
    raw = str(value or "").strip()
    if not raw:
        return set()
    decoded_versions = [raw]
    for _ in range(3):
        decoded = urllib.parse.unquote(decoded_versions[-1])
        if decoded == decoded_versions[-1]:
            break
        decoded_versions.append(decoded)
    candidates = set(decoded_versions)
    for text in decoded_versions:
        candidates.update(re.findall(r"https?://[^\s<>'\"]+", text, flags=re.I))
        try:
            for key, target in urllib.parse.parse_qsl(urllib.parse.urlsplit(text).query, keep_blank_values=False):
                if key.casefold() in {"q", "url", "target", "dest", "destination", "continue"}:
                    candidates.add(urllib.parse.unquote(target))
        except ValueError:
            pass
    return {signature for candidate in candidates if (signature := normalize_url(candidate))}


def exact_title(value: str) -> str:
    return re.sub(r"\s+", " ", unicodedata.normalize("NFKC", str(value or ""))).strip().casefold()


def title_for_similarity(value: str) -> str:
    text = exact_title(value)
    text = re.sub(r"\b(?:19|20)\d{2}[-/]\d{1,2}[-/]\d{1,2}(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?\b", " ", text)
    text = re.sub(r"\b(?:mon|tue|wed|thu|fri|sat|sun),?\s+\d{1,2}\s+(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+(?:19|20)\d{2}(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?(?:\s+gmt)?\b", " ", text, flags=re.I)
    text = re.sub(r"\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+\d{1,2},?\s+(?:19|20)\d{2}\b", " ", text, flags=re.I)
    text = re.sub(r"\s+(?:by|via)\s+[\w.&'’-]+(?:\s+[\w.&'’-]+){0,5}\s*$", "", text, flags=re.I)
    text = re.sub(r"\s+(?:[-–—|:]\s*)?(?:www\.)?[a-z0-9][a-z0-9.-]*\.(?:com|net|org|co|io|tw|ca|uk|news)(?:\.[a-z]{2})?\s*$", "", text, flags=re.I)
    return re.sub(r"[^0-9a-z\u3400-\u9fff]+", " ", text).strip()


def normalized_title(value: str) -> str:
    """給單一方法的快速去重使用，與 EXE 的標題清理一致。"""
    return title_for_similarity(value).replace(" ", "")


def numeric_facts(value: str) -> tuple[str, ...]:
    return tuple(sorted(re.findall(r"\d+(?:\.\d+)?%?", title_for_similarity(value))))


def title_similarity(left: str, right: str) -> float:
    left_key, right_key = title_for_similarity(left), title_for_similarity(right)
    if min(len(left_key), len(right_key)) < 12 or numeric_facts(left) != numeric_facts(right):
        return 0.0
    compact_score = SequenceMatcher(None, left_key.replace(" ", ""), right_key.replace(" ", "")).ratio()
    token_score = SequenceMatcher(None, " ".join(sorted(left_key.split())), " ".join(sorted(right_key.split()))).ratio()
    return max(compact_score, token_score)


def merge_company_key(record: dict) -> tuple[str, str] | None:
    ticker = re.sub(r"[\s_\-]+", "", unicodedata.normalize("NFKC", str(record.get("Ticker", ""))).strip().casefold())
    company = re.sub(r"[\s_\-]+", "", unicodedata.normalize("NFKC", str(record.get("Company", ""))).strip().casefold())
    return (ticker, company) if ticker and company else None


def merge_frames(first: pd.DataFrame, second: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """EXE 同款去重；但保留資料較完整的一筆，並固定輸出網站七欄。"""
    input_rows = (
        first.assign(_file="檔案一").to_dict("records")
        + second.assign(_file="檔案二").to_dict("records")
    )
    kept: list[dict] = []
    duplicates: list[dict] = []
    seen_urls: dict[tuple[tuple[str, str], str], int] = {}
    seen_exact_titles: dict[tuple[tuple[str, str], str], int] = {}
    retained_by_company: dict[tuple[str, str], list[int]] = {}

    def remember(record: dict, index: int) -> None:
        company_key = merge_company_key(record)
        if company_key is None:
            return
        for signature in url_signatures(record.get("URL", "")):
            seen_urls.setdefault((company_key, signature), index)
        title_key = exact_title(record.get("Title", ""))
        if title_key:
            seen_exact_titles.setdefault((company_key, title_key), index)
        indexes = retained_by_company.setdefault(company_key, [])
        if index not in indexes:
            indexes.append(index)

    for record in input_rows:
        company_key = merge_company_key(record)
        match_index: int | None = None
        reason = ""
        if company_key is not None:
            for signature in url_signatures(record.get("URL", "")):
                match_index = seen_urls.get((company_key, signature))
                if match_index is not None:
                    reason = "AB相同且網址相同（已忽略Google轉址前綴與追蹤參數）"
                    break
            if match_index is None:
                title_key = exact_title(record.get("Title", ""))
                if title_key:
                    match_index = seen_exact_titles.get((company_key, title_key))
                    if match_index is not None:
                        reason = "AB相同且主旨完全相同"
            if match_index is None:
                best_score, best_index = 0.0, None
                for candidate_index in retained_by_company.get(company_key, []):
                    score = title_similarity(kept[candidate_index].get("Title", ""), record.get("Title", ""))
                    if score > best_score:
                        best_score, best_index = score, candidate_index
                if best_index is not None and best_score >= 0.98:
                    match_index = best_index
                    reason = f"AB相同且主旨高度相似（{best_score:.0%}，數字資訊一致）"

        if match_index is None:
            kept.append(record)
            remember(record, len(kept) - 1)
            continue

        old = kept[match_index]
        if sum(bool(str(record.get(column, "")).strip()) for column in NEWS_COLUMNS) > sum(bool(str(old.get(column, "")).strip()) for column in NEWS_COLUMNS):
            kept[match_index] = record
        remember(record, match_index)
        duplicates.append({
            "原因": reason, "Ticker": record.get("Ticker", ""), "Company": record.get("Company", ""),
            "保留標題": kept[match_index].get("Title", ""), "重複標題": record.get("Title", ""),
            "保留網址": kept[match_index].get("URL", ""), "刪除網址": record.get("URL", ""),
            "來源檔": record.get("_file", ""),
        })

    result = pd.DataFrame(kept).drop(columns=["_file"], errors="ignore")
    if result.empty:
        result = pd.DataFrame(columns=NEWS_COLUMNS)
    return result.reindex(columns=NEWS_COLUMNS), pd.DataFrame(duplicates)


def translate_text_zh(text: str) -> str:
    text = str(text or "").strip()
    if not text:
        return ""
    for attempt in range(3):
        try:
            response = requests.get(
                "https://translate.googleapis.com/translate_a/single",
                params={"client": "gtx", "sl": "auto", "tl": "zh-TW", "dt": "t", "q": text},
                headers={"User-Agent": "Mozilla/5.0"}, timeout=15,
            )
            response.raise_for_status()
            payload = response.json()
            translated = "".join(str(part[0]) for part in payload[0] if part and part[0]).strip()
            if translated:
                return translated
        except Exception:
            pass
        if attempt < 2:
            time.sleep(1.5 * (attempt + 1))
    return ""


def translate_title_zh(title: str) -> str:
    return translate_text_zh(title)


def translate_output_rows(rows: list[dict], label: str, progress=None,
                          progress_start: float = 0.97, progress_end: float = 0.995) -> int:
    """批次翻譯輸出列；批次無法正確分割時再逐篇重試。"""
    if not rows:
        return 0
    separator = "<<<FUBON_NEWS_SPLIT>>>"
    batch_size = 15
    failures = 0
    batches = [rows[start:start + batch_size] for start in range(0, len(rows), batch_size)]

    def translate_batch(batch: list[dict]) -> list[str]:
        titles = [str(row.get("Title", "") or "").strip() for row in batch]
        joined = f"\n{separator}\n".join(titles)
        translated = translate_text_zh(joined)
        parts = [part.strip() for part in translated.split(separator)] if translated else []
        if len(parts) != len(batch):
            parts = [translate_title_zh(title) for title in titles]
        return parts

    done = 0
    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = {executor.submit(translate_batch, batch): batch for batch in batches}
        for future in as_completed(futures):
            batch = futures[future]
            try:
                parts = future.result()
            except Exception:
                parts = [""] * len(batch)
            for row, title_zh in zip(batch, parts):
                row["Title_ZH"] = title_zh
                if not title_zh:
                    failures += 1
            done += len(batch)
            if progress is not None:
                value = progress_start + (progress_end - progress_start) * done / len(rows)
                progress.progress(value, text=f"翻譯{label}標題｜{done}/{len(rows)}")
    return failures


def event_output(record: dict, rule: dict, keyword: str = "", title_zh: str = "") -> dict:
    return {
        "Published Time": record.get("Published Time", ""),
        "Ticker": record.get("Ticker", ""),
        "Company": record.get("Company", ""),
        "Title": record.get("Title", ""),
        "Title_ZH": title_zh,
        "FinBERT": record.get("FinBERT", ""),
        "Event_type": rule.get("Event_type", ""),
        "Event_Code": rule.get("Event_Code", ""),
        "事件中文": rule.get("事件中文", ""),
        "Level": rule.get("Level", ""),
        "Action": rule.get("Action", ""),
        "Keyword": keyword,
        "URL": record.get("URL", ""),
        "Source": record.get("Source", ""),
    }


def classify_news_sets(frame: pd.DataFrame, rules: list[dict], core_sources: set[str], progress=None) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    negative_rows: list[dict] = []
    unknown_rows, irrelevant_rows = [], []
    unknown_rule = next((rule for rule in rules if str(rule.get("Event_Code")) == "UNKNOWN"), None)
    irrelevant_rule = next((rule for rule in rules if str(rule.get("Event_Code")) == "NOT_RELEVANT"), None)
    non_core_rule = next((rule for rule in rules if str(rule.get("Event_Code")) == "NON_CORE_SOURCE"), None)
    if unknown_rule is None or irrelevant_rule is None or non_core_rule is None:
        raise ValueError("目前關鍵字表必須包含 UNKNOWN、NOT_RELEVANT 與 NON_CORE_SOURCE 事件")
    for record in frame.to_dict("records"):
        title = record["Title"].lower()
        matches = []
        for rule in rules:
            if not rule["_keywords"]:
                continue
            if any(term.lower() in title for term in rule["_exclusions"]):
                continue
            hits = [term for term in rule["_keywords"] if term.lower() in title]
            if hits:
                matches.append((rule, hits))
        if not matches:
            source = str(record.get("Source") or "").strip().casefold()
            if source not in core_sources:
                irrelevant_rows.append(event_output(record, non_core_rule))
            else:
                try:
                    finbert = float(record.get("FinBERT"))
                except (TypeError, ValueError):
                    finbert = None
                fallback = irrelevant_rule if finbert is not None and finbert > 0 else unknown_rule
                target = irrelevant_rows if fallback is irrelevant_rule else unknown_rows
                target.append(event_output(record, fallback))
            continue
        # 所有判定值皆取自目前 Parameter_Event.xlsx；不再另建新聞分數。
        matches.sort(key=lambda pair: (-int(pair[0].get("Primary_Priority") or 0), str(pair[0].get("Event_ID") or "")))
        primary, hits = matches[0]
        output = event_output(record, primary, "；".join(hits))
        code = str(primary.get("Event_Code") or "")
        event_type = str(primary.get("Event_type") or "").upper()
        if code == "UNKNOWN":
            unknown_rows.append(output)
        elif event_type == "SYSTEM":
            irrelevant_rows.append(output)
        else:
            negative_rows.append(output)
    negative_translation_failures = translate_output_rows(negative_rows, "負面新聞", progress, 0.970, 0.980)
    unknown_translation_failures = translate_output_rows(unknown_rows, "待人工覆核", progress, 0.980, 0.990)
    irrelevant_translation_failures = translate_output_rows(irrelevant_rows, "無關新聞", progress, 0.990, 0.998)

    def make_frame(rows: list[dict]) -> pd.DataFrame:
        result = pd.DataFrame(rows, columns=NEGATIVE_OUTPUT_COLUMNS)
        if not result.empty:
            result["Level"] = pd.to_numeric(result["Level"], errors="coerce")
            result = result.sort_values(["Level", "Published Time"], ascending=[False, False]).reset_index(drop=True)
        return result

    negative_frame = make_frame(negative_rows)
    unknown_frame = make_frame(unknown_rows)
    irrelevant_frame = make_frame(irrelevant_rows)
    negative_frame.attrs["translation_failures"] = negative_translation_failures
    unknown_frame.attrs["translation_failures"] = unknown_translation_failures
    irrelevant_frame.attrs["translation_failures"] = irrelevant_translation_failures
    return negative_frame, unknown_frame, irrelevant_frame


def write_excel(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name[:31], index=False)
            sheet = writer.sheets[name[:31]]
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            is_negative_output = list(frame.columns) == NEGATIVE_OUTPUT_COLUMNS
            for index, cell in enumerate(sheet[1], 1):
                if is_negative_output:
                    cell.fill = PatternFill("solid", fgColor="FFF200" if 6 <= index <= 12 else "E2F0D9")
                    cell.font = Font(color="000000", bold=True)
                else:
                    cell.fill = PatternFill("solid", fgColor="1F4E78")
                    cell.font = Font(color="FFFFFF", bold=True)
            negative_widths = {
                "Published Time": 21, "Ticker": 12, "Company": 24, "Title": 60,
                "Title_ZH": 45, "FinBERT": 12, "Event_type": 18, "Event_Code": 27,
                "事件中文": 22, "Level": 10, "Action": 45, "Keyword": 30, "URL": 65, "Source": 18,
            }
            for index, column in enumerate(frame.columns, 1):
                width = negative_widths.get(column, 22) if is_negative_output else (80 if column in {"URL", "Definition"} else 60 if column == "Title" else 22)
                sheet.column_dimensions[get_column_letter(index)].width = width
            sheet.row_dimensions[1].height = 24
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)


# ============================================================
# 區塊：方法一｜多來源新聞擷取
# 彙整 MoneyDJ、經濟日報、鉅亨網、CNBC 四個新聞池，
# 再用 TradingView 逐家公司補抓；不使用 Nasdaq。
# ============================================================
METHOD1_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/126 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9,zh-TW;q=0.8",
}
METHOD1_POOL_LIMIT = 300
METHOD1_EXCHANGES = {"AAPL", "AMZN", "CSCO", "GOOGL", "MSFT", "NVDA", "AMGN", "HON"}
METHOD1_THREAD_LOCAL = threading.local()


def method1_http_session() -> requests.Session:
    session = getattr(METHOD1_THREAD_LOCAL, "session", None)
    if session is None:
        session = requests.Session()
        session.headers.update(METHOD1_HEADERS)
        METHOD1_THREAD_LOCAL.session = session
    return session


def parse_method1_time(value: object, reference: datetime | None = None) -> datetime | None:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if not text:
        return None
    now = reference or datetime.now(TAIPEI)
    relative = re.search(r"(\d+)\s+(minute|minutes|hour|hours|day|days)\s+ago", text, re.I)
    if relative:
        amount = int(relative.group(1))
        unit = relative.group(2).lower()
        delta = timedelta(minutes=amount) if unit.startswith("minute") else timedelta(hours=amount) if unit.startswith("hour") else timedelta(days=amount)
        return now - delta
    try:
        return parsedate_to_datetime(text).astimezone(TAIPEI)
    except Exception:
        pass
    cleaned = re.sub(r"\b(?:ET|EST|EDT|CT|CST|CDT)\b", "", text, flags=re.I).strip()
    parsed = pd.to_datetime(cleaned, errors="coerce")
    if pd.isna(parsed):
        return None
    result = parsed.to_pydatetime()
    if result.tzinfo is None:
        result = result.replace(tzinfo=TAIPEI)
    return result.astimezone(TAIPEI)


def method1_in_window(published: datetime | None, start: datetime, end: datetime) -> bool:
    return published is not None and start <= published.astimezone(TAIPEI) <= end


def method1_company_terms(ticker: str, company: str) -> list[str]:
    name = re.sub(r"\s+", " ", company).strip()
    simplified = re.sub(
        r"\s+(?:incorporated|inc\.?|corporation|corp\.?|company|co\.?|limited|ltd\.?|plc|group|holdings?)$",
        "", name, flags=re.I,
    ).strip()
    terms = [name]
    if simplified and simplified.lower() != name.lower() and len(simplified) >= 4:
        terms.append(simplified)
    if ticker == "JNJ":
        terms.extend(["Johnson & Johnson", "J&J", "嬌生"])
    return list(dict.fromkeys(term for term in terms if len(term) >= 4))


def method1_match_company(title: str, summary: str, companies: pd.DataFrame) -> tuple[str, str] | None:
    text = re.sub(r"\s+", " ", f"{title} {summary}").strip()
    candidates: list[tuple[int, str, str]] = []
    for row in companies.itertuples(index=False):
        ticker, company = str(row.Ticker).upper(), str(row.Company)
        structured = re.search(rf"(?:\${re.escape(ticker)}\b|\({re.escape(ticker)}\)|(?:NASDAQ|NYSE|AMEX)\s*:\s*{re.escape(ticker)}\b)", text, re.I)
        matched_terms = [term for term in method1_company_terms(ticker, company) if re.search(rf"(?<![A-Za-z0-9]){re.escape(term)}(?![A-Za-z0-9])", text, re.I)]
        if structured or matched_terms:
            strength = 1000 if structured else max(len(term) for term in matched_terms)
            candidates.append((strength, ticker, company))
    if not candidates:
        return None
    candidates.sort(key=lambda item: (-item[0], -len(item[2]), item[1]))
    return candidates[0][1], candidates[0][2]


def method1_rss(source: str, url: str, start: datetime, end: datetime) -> list[dict]:
    response = requests.get(url, headers=METHOD1_HEADERS, timeout=25)
    response.raise_for_status()
    root = ET.fromstring(response.content)
    rows = []
    for node in root.findall(".//item") + root.findall(".//{*}entry"):
        title = (node.findtext("title") or node.findtext("{*}title") or "").strip()
        link = (node.findtext("link") or "").strip()
        if not link:
            link_node = node.find("{*}link")
            link = (link_node.get("href", "") if link_node is not None else "").strip()
        description = node.findtext("description") or node.findtext("{*}summary") or ""
        published_text = node.findtext("pubDate") or node.findtext("{*}published") or node.findtext("{*}updated") or ""
        published = parse_method1_time(published_text)
        if title and link and method1_in_window(published, start, end):
            rows.append({"Title": title, "URL": link, "Source": source, "Published": published, "Summary": BeautifulSoup(description, "html.parser").get_text(" ", strip=True)})
    return rows


def method1_moneydj(start: datetime, end: datetime) -> list[dict]:
    rows = []
    for url in ("https://www.moneydj.com/kmdj/news/newsreal_list.aspx?a=MB010000", "https://www.moneydj.com/kmdj/news/newsreal_list.aspx?a=MB020000"):
        response = requests.get(url, headers=METHOD1_HEADERS, timeout=25)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")
        for link in soup.select('a[href*="newsviewer.aspx"]'):
            title = link.get_text(" ", strip=True)
            container = link.find_parent(["li", "tr", "div"]) or link
            published = parse_method1_time(container.get_text(" ", strip=True))
            if title and method1_in_window(published, start, end):
                rows.append({"Title": title, "URL": urljoin("https://www.moneydj.com", link.get("href", "")), "Source": "MoneyDJ", "Published": published, "Summary": container.get_text(" ", strip=True)})
    return rows[:METHOD1_POOL_LIMIT]


def method1_udn(start: datetime, end: datetime) -> list[dict]:
    rows = []
    for url in ("https://money.udn.com/rssfeed/news/1001/5591", "https://money.udn.com/rssfeed/news/1001/5590"):
        rows.extend(method1_rss("經濟日報", url, start, end))
    return rows[:METHOD1_POOL_LIMIT]


def method1_cnbc(start: datetime, end: datetime) -> list[dict]:
    rows = []
    for url in ("https://www.cnbc.com/id/100003114/device/rss/rss.html", "https://www.cnbc.com/id/10001147/device/rss/rss.html"):
        rows.extend(method1_rss("CNBC", url, start, end))
    return rows[:METHOD1_POOL_LIMIT]


def method1_cnyes(start: datetime, end: datetime) -> list[dict]:
    rows = []
    for category in ("us_stock", "wd_stock", "headline"):
        url = f"https://api.cnyes.com/media/api/v1/newslist/category/{category}"
        for page in range(1, 7):
            response = requests.get(url, params={"limit": 50, "page": page}, headers=METHOD1_HEADERS, timeout=25)
            if response.status_code == 422:
                break
            response.raise_for_status()
            payload = response.json().get("items") or response.json().get("data", {}).get("items") or []
            records = payload.get("data", []) if isinstance(payload, dict) else payload
            if not records:
                break
            stop_category = False
            for record in records:
                timestamp = record.get("publishAt") or record.get("createdAt") or record.get("created_at")
                try:
                    stamp = int(timestamp)
                    if stamp > 10_000_000_000:
                        stamp //= 1000
                    published = datetime.fromtimestamp(stamp, TAIPEI)
                except Exception:
                    published = None
                if published and published < start:
                    stop_category = True
                    continue
                if not method1_in_window(published, start, end):
                    continue
                raw_title = str(record.get("title") or record.get("name") or "")
                title = BeautifulSoup(raw_title, "html.parser").get_text(" ", strip=True) if "<" in raw_title and ">" in raw_title else raw_title.strip()
                news_id = record.get("newsId") or record.get("id")
                link = record.get("url") or (f"https://news.cnyes.com/news/id/{news_id}" if news_id else "")
                if title and link:
                    rows.append({"Title": title, "URL": link, "Source": "鉅亨網", "Published": published, "Summary": ""})
            if stop_category or len(rows) >= METHOD1_POOL_LIMIT:
                break
    return rows[:METHOD1_POOL_LIMIT]


def method1_tradingview(ticker: str, company: str, start: datetime, end: datetime,
                        session: requests.Session | None = None) -> list[dict]:
    primary = "NASDAQ" if ticker in METHOD1_EXCHANGES else "NYSE"
    response = None
    used_symbol = ""
    for exchange in (primary, "NYSE" if primary == "NASDAQ" else "NASDAQ"):
        used_symbol = f"{exchange}:{ticker}"
        response = (session or requests).get(
            "https://news-mediator.tradingview.com/public/news-flow/v2/news",
            params=[("filter", "lang:zh-Hant"), ("filter", f"symbol:{used_symbol}"), ("client", "landing"), ("streaming", "false"), ("user_prostatus", "non_pro")],
            headers={**METHOD1_HEADERS, "Referer": f"https://tw.tradingview.com/symbols/{used_symbol.replace(':', '-')}/news/"},
            timeout=25,
        )
        if response.status_code != 422:
            response.raise_for_status()
            break
    if response is None or response.status_code == 422:
        return []
    rows = []
    for record in response.json().get("items", []):
        try:
            published = datetime.fromtimestamp(int(record.get("published")), TAIPEI)
        except Exception:
            continue
        title = str(record.get("title") or "").strip()
        link = str(record.get("link") or "").strip()
        if link.startswith("/"):
            link = urljoin("https://tw.tradingview.com", link)
        if not link:
            link = f"https://tw.tradingview.com/symbols/{used_symbol.replace(':', '-')}/news/"
        if title and method1_in_window(published, start, end):
            rows.append({"Ticker": ticker, "Company": company, "Published Time": published.strftime("%Y-%m-%d %H:%M:%S"), "Title": title, "Source": "TradingView", "URL": link, "FinBERT": ""})
    return rows


def fetch_method1_news(companies: pd.DataFrame, start: datetime, end: datetime, progress=None) -> tuple[pd.DataFrame, list[str], dict[str, int]]:
    """方法一：四個新聞池，加 TradingView 逐公司補抓；不使用 Nasdaq。"""
    errors: list[str] = []
    source_counts: dict[str, int] = {name: 0 for name in ("MoneyDJ", "經濟日報", "鉅亨網", "CNBC", "TradingView")}
    rows: list[dict] = []
    pool_fetchers = (("MoneyDJ", method1_moneydj), ("經濟日報", method1_udn), ("鉅亨網", method1_cnyes), ("CNBC", method1_cnbc))
    pool_items: list[dict] = []
    for source, fetcher in pool_fetchers:
        try:
            fetched = fetcher(start, end)
            pool_items.extend(fetched)
        except Exception as exc:
            errors.append(f"{source} 新聞池：{exc}")
        if isinstance(progress, BackgroundProgress):
            progress.job["method1_monitor"].update(
                stage=f"新聞池處理中｜{source}", errors=list(errors), source_counts=dict(source_counts)
            )
    for item in pool_items:
        matched = method1_match_company(item["Title"], item.get("Summary", ""), companies)
        if not matched:
            continue
        ticker, company = matched
        rows.append({"Ticker": ticker, "Company": company, "Published Time": item["Published"].strftime("%Y-%m-%d %H:%M:%S"), "Title": item["Title"], "Source": item["Source"], "URL": item["URL"], "FinBERT": ""})
        source_counts[item["Source"]] += 1

    def fetch_company(record: dict) -> tuple[list[dict], list[str]]:
        ticker, company = record["Ticker"], record["Company"]
        company_rows, company_errors = [], []
        session = method1_http_session()
        for source, fetcher in (("TradingView", method1_tradingview),):
            try:
                company_rows.extend(fetcher(ticker, company, start, end, session=session))
            except Exception as exc:
                company_errors.append(f"{source} {ticker}：{exc}")
        return company_rows, company_errors

    records = companies[["Ticker", "Company"]].to_dict("records")
    # TradingView 是方法一最主要的耗時階段；連線重用加上 12 個工作執行緒，
    # 可降低 S&P 500 逐家請求的總等待時間，同時避免過高並行導致限流。
    executor = ThreadPoolExecutor(max_workers=12)
    futures = {executor.submit(fetch_company, record): record for record in records}
    try:
        for completed, future in enumerate(as_completed(futures), 1):
            if isinstance(progress, BackgroundProgress) and progress.job["stop_event"].is_set():
                raise CrawlCancelled("使用者已停止執行")
            company_rows, company_errors = future.result()
            rows.extend(company_rows)
            errors.extend(company_errors)
            for row in company_rows:
                source_counts[row["Source"]] += 1
            if progress is not None:
                progress.progress(completed / len(records), text=f"方法一逐公司補抓｜{completed}/{len(records)}")
            if isinstance(progress, BackgroundProgress):
                progress.job["method1_monitor"].update(
                    stage=f"逐公司補抓｜{completed}/{len(records)}", rows=len(rows),
                    errors=list(errors), source_counts=dict(source_counts),
                )
    finally:
        stopping = isinstance(progress, BackgroundProgress) and progress.job["stop_event"].is_set()
        if stopping:
            for future in futures:
                future.cancel()
        executor.shutdown(wait=not stopping, cancel_futures=stopping)
    frame = normalize_news(pd.DataFrame(rows, columns=NEWS_COLUMNS))
    if frame.empty:
        return frame, errors, source_counts
    frame["_url_key"] = frame["URL"].map(normalize_url)
    frame["_title_key"] = frame["Title"].map(normalized_title)
    frame = frame.drop_duplicates(["Company", "_url_key"]).drop_duplicates(["Company", "_title_key"]).drop(columns=["_url_key", "_title_key"])
    frame = frame.sort_values("Published Time", ascending=False).reset_index(drop=True)
    return frame[NEWS_COLUMNS], errors, source_counts


# ============================================================
# 區塊：方法二｜Google News 擷取 ＋ FinBERT 情緒評分
# 分批查詢英文 Google News RSS，再用 ProsusAI/finbert 模型
# 對新聞標題評分（正面機率－負面機率），輸出 FinBERT 分數。
# ============================================================
def method2_matches(title: str, companies: pd.DataFrame) -> list[tuple[str, str]]:
    """比對 Google News 標題；同一則新聞可對應多家公司。"""
    text = str(title).strip()
    lowered = text.lower()
    matches: list[tuple[str, str]] = []
    ambiguous = {"A", "AI", "ALL", "AN", "ARE", "AT", "BE", "BY", "FOR", "IT", "ON", "OR", "SO"}
    for row in companies.itertuples(index=False):
        ticker, company = str(row.Ticker).strip().upper(), str(row.Company).strip()
        ticker_hit = bool(ticker and ticker not in ambiguous and re.search(rf"(?<![A-Za-z0-9]){re.escape(ticker)}(?![A-Za-z0-9])", text, re.I))
        company_hit = bool(company and len(company) >= 3 and company.lower() in lowered)
        if ticker_hit or company_hit:
            matches.append((ticker, company))
    return matches


def fetch_method2_raw(companies: pd.DataFrame, start: datetime, end: datetime, progress=None) -> tuple[pd.DataFrame, list[str], int]:
    """方法二第一階段：每五家公司一批查詢 Google News RSS。"""
    records = companies[["Ticker", "Company"]].drop_duplicates("Ticker").to_dict("records")
    batches = [records[index:index + 5] for index in range(0, len(records), 5)]
    rows: list[dict] = []
    errors: list[str] = []
    before_day = (end.astimezone(TAIPEI).date() + timedelta(days=1)).isoformat()
    after_day = start.astimezone(TAIPEI).date().isoformat()
    for batch_index, batch in enumerate(batches, 1):
        terms: list[str] = []
        batch_frame = pd.DataFrame(batch)
        for item in batch:
            terms.extend([f'"{item["Ticker"]}"', f'"{item["Company"]}"'])
        query = f'intitle:({" OR ".join(terms)}) after:{after_day} before:{before_day}'
        rss_url = "https://news.google.com/rss/search?" + urllib.parse.urlencode({"q": query, "hl": "en-US", "gl": "US", "ceid": "US:en"})
        try:
            response = requests.get(rss_url, headers=METHOD1_HEADERS, timeout=12)
            response.raise_for_status()
            root = ET.fromstring(response.content)
            for item in root.findall(".//item"):
                full_title = (item.findtext("title") or "").strip()
                title, source = full_title, "未知媒體"
                if " - " in full_title:
                    title, source = full_title.rsplit(" - ", 1)
                source = (item.findtext("source") or source).strip()
                published = parse_method1_time(item.findtext("pubDate"))
                if not method1_in_window(published, start, end):
                    continue
                for ticker, company in method2_matches(title, batch_frame):
                    rows.append({"Ticker": ticker, "Company": company, "Published Time": published.astimezone(TAIPEI).strftime("%Y-%m-%d %H:%M:%S"), "Title": title, "Source": source, "URL": (item.findtext("link") or "").strip(), "FinBERT": ""})
        except Exception as exc:
            tickers = ", ".join(str(item["Ticker"]) for item in batch)
            errors.append(f"Google News 批次 {tickers}：{exc}")
        if progress is not None:
            progress.progress(batch_index / max(len(batches), 1) * 0.55, text=f"方法二爬取新聞｜{batch_index}/{len(batches)} 批")
        if isinstance(progress, BackgroundProgress):
            progress.job["method2_monitor"].update(
                stage=f"Google News 批次｜{batch_index}/{len(batches)}", raw_rows=len(rows),
                batches=batch_index, errors=list(errors),
            )
        time.sleep(1)
    result = pd.DataFrame(rows, columns=NEWS_COLUMNS)
    if not result.empty:
        result["_url"] = result["URL"].map(normalize_url)
        result["_title"] = result["Title"].map(normalized_title)
        result = result.drop_duplicates(["Ticker", "_url"]).drop_duplicates(["Ticker", "_title"]).drop(columns=["_url", "_title"])
        result = result.sort_values("Published Time", ascending=False).reset_index(drop=True)
    return result, errors, len(batches)


CHINESE_CHAR_PATTERN = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff]")


def is_chinese_title(text: str) -> bool:
    """FinBERT 僅支援英文，原始標題含中文字元者視為中文新聞，評分改標記 N/A。"""
    return bool(CHINESE_CHAR_PATTERN.search(str(text or "")))


@st.cache_resource(show_spinner=False)
def load_finbert_model():
    try:
        import torch
        from transformers import AutoModelForSequenceClassification, AutoTokenizer
    except ImportError as exc:
        raise RuntimeError("方法二需要 transformers 與 torch，請先執行 pip install -r requirements.txt") from exc
    model_name = "ProsusAI/finbert"
    device = "cuda" if torch.cuda.is_available() else "cpu"
    if device == "cpu":
        # 本機為 8 個邏輯核心；保留少數核心給網站與作業系統，實測 6 執行緒最快。
        cpu_threads = min(6, max(1, int(os.cpu_count() or 4) - 2))
        torch.set_num_threads(cpu_threads)
    try:
        tokenizer = AutoTokenizer.from_pretrained(model_name, local_files_only=True)
        model = AutoModelForSequenceClassification.from_pretrained(model_name, local_files_only=True).to(device)
    except OSError:
        tokenizer = AutoTokenizer.from_pretrained(model_name)
        model = AutoModelForSequenceClassification.from_pretrained(model_name).to(device)
    model.eval()
    return torch, tokenizer, model, device


def score_method2_finbert(frame: pd.DataFrame, progress=None) -> tuple[pd.DataFrame, pd.DataFrame]:
    """方法二第二階段：FinBERT = positive probability - negative probability。
    FinBERT 模型只認得英文，原始標題含中文字元的新聞不送進模型評分，
    一律將 FinBERT 標記為「N/A」，避免出現不準確的中文評分。"""
    scored = frame.copy()
    if scored.empty:
        scored["FinBERT"] = pd.Series(dtype=object)
        return scored, scored.copy()
    titles = scored["Title"].fillna("").astype(str).tolist()
    is_chinese_flags = [is_chinese_title(title) for title in titles]
    scoreable_indexes = [index for index, is_zh in enumerate(is_chinese_flags) if not is_zh]
    skipped_count = len(titles) - len(scoreable_indexes)
    scores: list[float] = [float("nan")] * len(titles)
    if scoreable_indexes:
        torch, tokenizer, model, device = load_finbert_model()
        # 本機 CPU 用較大批次降低模型呼叫開銷；依標題長度分組可減少 padding 的無效運算。
        batch_size = 128
        ordered_indexes = sorted(scoreable_indexes, key=lambda index: len(titles[index].split()))
        labels = {int(key): str(value).lower() for key, value in model.config.id2label.items()}
        with torch.inference_mode():
            for index in range(0, len(ordered_indexes), batch_size):
                batch_indexes = ordered_indexes[index:index + batch_size]
                batch = [titles[i] for i in batch_indexes]
                inputs = tokenizer(batch, padding=True, truncation=True, max_length=64, return_tensors="pt").to(device)
                probabilities = torch.nn.functional.softmax(model(**inputs).logits, dim=-1).cpu().numpy()
                for original_index, values in zip(batch_indexes, probabilities):
                    mapped = {labels[position]: float(values[position]) for position in range(len(values))}
                    scores[original_index] = round(mapped.get("positive", 0.0) - mapped.get("negative", 0.0), 3)
                if progress is not None:
                    done = min(index + batch_size, len(ordered_indexes))
                    progress.progress(
                        0.55 + 0.45 * done / len(ordered_indexes),
                        text=f"方法二 FinBERT 評分｜{done}/{len(ordered_indexes)}（另有 {skipped_count} 則中文新聞標記為 N/A）",
                    )
    elif progress is not None:
        progress.progress(1.0, text=f"方法二 FinBERT 評分｜全數 {skipped_count} 則為中文新聞，已標記為 N/A")
    scored["_finbert_numeric"] = scores
    scored = scored.sort_values("_finbert_numeric", na_position="last").reset_index(drop=True)
    filtered = scored[(scored["_finbert_numeric"] >= -1) & (scored["_finbert_numeric"] <= 0)].reset_index(drop=True)
    filtered["FinBERT"] = filtered["_finbert_numeric"]
    scored["FinBERT"] = scored["_finbert_numeric"].apply(lambda value: "N/A" if pd.isna(value) else value)
    scored = scored.drop(columns=["_finbert_numeric"])
    filtered = filtered.drop(columns=["_finbert_numeric"])
    return scored, filtered


# ============================================================
# 區塊：公司名單（Ticker/Company）管理
# Dow Jones 30／S&P 500 直接讀取 GitHub repo 內的 Excel 檔
# （DowJones30.xlsx／SP500.xlsx），檔案內容更新後重新整理網站即可套用。
# 另外保留「上傳公司列表」功能，可另外套用一份自訂 Excel 名單。
# ============================================================
def _parse_company_frame(frame: pd.DataFrame) -> pd.DataFrame:
    columns = {str(column).strip().lower(): column for column in frame.columns}
    ticker_col = next((columns[key] for key in ("ticker", "symbol", "股票代號") if key in columns), None)
    company_col = next((columns[key] for key in ("company", "name", "股票名稱") if key in columns), None)
    if not ticker_col or not company_col:
        raise ValueError("名單需要 Ticker/Symbol 與 Company/Name 欄位")
    result = frame[[ticker_col, company_col]].rename(columns={ticker_col: "Ticker", company_col: "Company"}).dropna()
    result["Ticker"] = result["Ticker"].astype(str).str.strip().str.upper()
    result["Company"] = result["Company"].astype(str).str.strip()
    return result[(result["Ticker"] != "") & (result["Company"] != "")].drop_duplicates("Ticker")


@st.cache_data(ttl=60)
def load_company_list(path_str: str, modified_ns: int) -> pd.DataFrame:
    del modified_ns  # 僅用來讓檔案異動時使快取失效
    frame = pd.read_excel(path_str, dtype=str)
    return _parse_company_frame(frame)


def load_company_list_from_upload(upload) -> pd.DataFrame:
    """給「上傳公司列表」用：直接解析使用者上傳的 Excel，不經過快取。"""
    frame = pd.read_excel(io.BytesIO(upload.getvalue()), dtype=str)
    return _parse_company_frame(frame)


# ============================================================
# ============================================================
# 區塊：曝險／部位清單管理（給風管部加權排序、判斷影響力用）
# 選填功能：上傳一份「Ticker + 曝險金額」的 Excel，之後負面新聞總覽
# 就能依曝險金額排序、加總，找出「曝險大 + 新聞嚴重」的公司優先處理。
# 支援選填「業務別」欄位（財管／經紀／自營），有這欄的話可以在氣泡圖與
# 新聞明細那邊切換只看某一個業務線的曝險；沒有這欄的話全部視為單一整體曝險。
# ============================================================
BUSINESS_LINE_ALIASES = {
    "財管": "財管", "wm": "財管", "wealth": "財管", "wealth management": "財管",
    "private banking": "財管", "財富管理": "財管", "理財": "財管",
    "經紀": "經紀", "brokerage": "經紀", "broker": "經紀", "受託買賣": "經紀", "經紀業務": "經紀",
    "自營": "自營", "proprietary": "自營", "prop": "自營", "自營部": "自營", "自營業務": "自營",
}
BUSINESS_LINE_ORDER = ["財管", "經紀", "自營"]
BUSINESS_LINE_ALL = "全部（加總）"


def _parse_exposure_frame(frame: pd.DataFrame) -> pd.DataFrame:
    columns = {str(column).strip().lower(): column for column in frame.columns}
    ticker_col = next((columns[key] for key in ("ticker", "symbol", "股票代號") if key in columns), None)
    exposure_col = next(
        (columns[key] for key in ("exposure", "amount", "曝險金額", "曝險", "部位金額", "部位") if key in columns),
        None,
    )
    business_col = next(
        (columns[key] for key in ("業務別", "業務", "部門", "business", "businessline", "business_line", "desk") if key in columns),
        None,
    )
    if not ticker_col or not exposure_col:
        raise ValueError("曝險清單需要 Ticker/Symbol 與 Exposure/Amount（曝險金額）欄位")
    use_cols = [ticker_col, exposure_col] + ([business_col] if business_col else [])
    rename_map = {ticker_col: "Ticker", exposure_col: "Exposure"}
    if business_col:
        rename_map[business_col] = "業務別"
    result = frame[use_cols].rename(columns=rename_map)
    result["Ticker"] = result["Ticker"].astype(str).str.strip().str.upper()
    result["Exposure"] = pd.to_numeric(result["Exposure"], errors="coerce").fillna(0)
    if business_col:
        raw_business = result["業務別"].fillna("").astype(str).str.strip()
        result["業務別"] = raw_business.map(lambda v: BUSINESS_LINE_ALIASES.get(v.lower(), v)).replace("", "未分類")
    else:
        result["業務別"] = BUSINESS_LINE_ALL
    result = result[result["Ticker"] != ""]
    return result.groupby(["Ticker", "業務別"], as_index=False)["Exposure"].sum()


def exposure_business_options(exposure_df: pd.DataFrame) -> list[str]:
    """依資料中實際出現的業務別，排出下拉選單選項（財管／經紀／自營優先，其餘按字母排在後面）。"""
    if exposure_df.empty:
        return [BUSINESS_LINE_ALL]
    present = [b for b in exposure_df["業務別"].unique() if b not in (BUSINESS_LINE_ALL,)]
    if not present:
        return [BUSINESS_LINE_ALL]
    ordered = [b for b in BUSINESS_LINE_ORDER if b in present]
    extra = sorted(b for b in present if b not in BUSINESS_LINE_ORDER)
    return [BUSINESS_LINE_ALL] + ordered + extra


def exposure_map_for_business(exposure_df: pd.DataFrame, business_filter: str) -> dict:
    if exposure_df.empty:
        return {}
    if business_filter == BUSINESS_LINE_ALL:
        return exposure_df.groupby("Ticker")["Exposure"].sum().to_dict()
    return exposure_df[exposure_df["業務別"] == business_filter].set_index("Ticker")["Exposure"].to_dict()


@st.cache_data(ttl=60)
def load_exposure_list(path_str: str, modified_ns: int) -> pd.DataFrame:
    del modified_ns  # 僅用來讓檔案異動時使快取失效
    frame = pd.read_excel(path_str, dtype=str)
    return _parse_exposure_frame(frame)


def load_exposure_list_from_upload(upload) -> pd.DataFrame:
    frame = pd.read_excel(io.BytesIO(upload.getvalue()), dtype=str)
    return _parse_exposure_frame(frame)


# ============================================================
# 區塊：歷史負面新聞彙整 → 寫回 GitHub repo，避免容器重開資料歸零
# 需要在 Streamlit Secrets 設定：
#   GITHUB_TOKEN = "ghp_xxxx"                 （需要 repo 寫入權限的 Personal Access Token）
#   GITHUB_REPO  = "your-account/your-repo"   （例如 "chengyu1212/negative-news"）
# 選填：GITHUB_BRANCH（預設 main）、GITHUB_HISTORY_PATH（預設 history/negative_news_history.csv）
# 沒有設定這兩個 Secrets 時，這個功能會靜默停用，自動退回讀本機 output 資料夾
# （效果跟原本一樣：容器重開資料就會不見，但完全不影響其他功能）。
# ============================================================
GITHUB_HISTORY_PATH_DEFAULT = "history/negative_news_history.csv"


def _github_config() -> dict | None:
    try:
        token = st.secrets.get("GITHUB_TOKEN", "")
        repo = st.secrets.get("GITHUB_REPO", "")
    except Exception:
        return None
    if not token or not repo:
        return None
    return {
        "token": token,
        "repo": repo,
        "branch": st.secrets.get("GITHUB_BRANCH", "main"),
        "path": st.secrets.get("GITHUB_HISTORY_PATH", GITHUB_HISTORY_PATH_DEFAULT),
    }


def _github_headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}", "Accept": "application/vnd.github+json"}


def _github_fetch_history() -> tuple[pd.DataFrame, str | None]:
    """讀回 GitHub repo 裡的歷史統計 CSV；檔案不存在時回傳空表＋sha=None（代表之後要新建檔案）。"""
    empty = pd.DataFrame(columns=["日期", "Ticker", "Company", "則數", "最高Level"])
    config = _github_config()
    if not config:
        return empty, None
    url = f"https://api.github.com/repos/{config['repo']}/contents/{config['path']}"
    try:
        response = requests.get(url, headers=_github_headers(config["token"]), params={"ref": config["branch"]}, timeout=20)
    except requests.RequestException:
        return empty, None
    if response.status_code != 200:
        return empty, None
    payload = response.json()
    sha = payload.get("sha")
    try:
        content = base64.b64decode(payload["content"]).decode("utf-8-sig")
        frame = pd.read_csv(io.StringIO(content))
        frame["日期"] = pd.to_datetime(frame["日期"])
        return frame, sha
    except Exception:
        return empty, sha


def _github_push_history(frame: pd.DataFrame, sha: str | None) -> bool:
    config = _github_config()
    if not config:
        return False
    csv_text = frame.sort_values("日期").to_csv(index=False)
    content_b64 = base64.b64encode(csv_text.encode("utf-8")).decode("ascii")
    url = f"https://api.github.com/repos/{config['repo']}/contents/{config['path']}"
    body = {
        "message": f"更新負面新聞歷史統計（{datetime.now(TAIPEI).strftime('%Y-%m-%d %H:%M')}）",
        "content": content_b64,
        "branch": config["branch"],
    }
    if sha:
        body["sha"] = sha
    try:
        response = requests.put(url, headers=_github_headers(config["token"]), json=body, timeout=20)
    except requests.RequestException:
        return False
    return response.status_code in (200, 201)


def append_negative_history_to_github(event_frame: pd.DataFrame, event_date: datetime) -> bool:
    """把今天的負面新聞彙整成「日期 x 公司 x 則數」寫回 GitHub repo。
    沒設定 GITHUB_TOKEN/GITHUB_REPO 就直接跳過（不影響抓取流程本身）。"""
    if not _github_config() or event_frame is None or event_frame.empty:
        return False
    daily = event_frame.copy()
    daily["Level"] = pd.to_numeric(daily.get("Level"), errors="coerce").fillna(0).astype(int)
    daily["Ticker"] = daily.get("Ticker", "").fillna("").astype(str).str.strip()
    daily["Company"] = daily.get("Company", "").fillna("").astype(str).str.strip()
    daily = daily[daily["Ticker"] != ""]
    if daily.empty:
        return False
    grouped = daily.groupby(["Ticker", "Company"]).agg(則數=("Ticker", "size"), 最高Level=("Level", "max")).reset_index()
    grouped["日期"] = pd.to_datetime(event_date.date())

    history, sha = _github_fetch_history()
    if not history.empty:
        history = history[history["日期"].dt.date != event_date.date()]  # 同一天重跑就整批覆蓋，不會重複累加
    merged = pd.concat([history, grouped], ignore_index=True)
    pushed = _github_push_history(merged, sha)
    if pushed:
        load_negative_history.clear()
    return pushed


# ============================================================
# 區塊：每天的風險新聞明細（負面新聞，含全部欄位）自動回存 GitHub
# 跟上面「歷史統計」（只有 日期/Ticker/則數/最高Level 的彙總）不同，
# 這裡存的是當天完整的風險新聞明細（每一則新聞、每個欄位都在），存成 Excel（.xlsx）。
# 一天一個檔案，路徑用日期命名，同一天重新爬蟲或人工覆核修改都會整批覆蓋（覆蓋舊的、留新的），
# 確保 GitHub 上留著的永遠是「當天最新（含人工修正後）」的版本。
# 沒設定 GITHUB_TOKEN/GITHUB_REPO 就直接跳過，不影響其他功能。
# ============================================================
GITHUB_RISK_NEWS_DIR_DEFAULT = "history/risk_news_daily"


def _github_risk_news_path(snapshot_date: date) -> str:
    try:
        base_dir = st.secrets.get("GITHUB_RISK_NEWS_DIR", GITHUB_RISK_NEWS_DIR_DEFAULT)
    except Exception:
        base_dir = GITHUB_RISK_NEWS_DIR_DEFAULT
    return f"{base_dir}/{snapshot_date.strftime('%Y-%m-%d')}.xlsx"


def _github_get_file_sha(path: str) -> str | None:
    """只取某個路徑目前在 GitHub 上的 sha（用來判斷檔案存不存在、覆蓋時要不要帶 sha）。"""
    config = _github_config()
    if not config:
        return None
    url = f"https://api.github.com/repos/{config['repo']}/contents/{path}"
    try:
        response = requests.get(url, headers=_github_headers(config["token"]), params={"ref": config["branch"]}, timeout=20)
    except requests.RequestException:
        return None
    if response.status_code != 200:
        return None
    return response.json().get("sha")


def _dataframe_to_excel_bytes(frame: pd.DataFrame, sheet_name: str = "風險新聞") -> bytes:
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir) / "snapshot.xlsx"
        write_excel(tmp_path, {sheet_name: frame})
        return tmp_path.read_bytes()


def _github_push_risk_news_snapshot(file_bytes: bytes, snapshot_date: date, sha: str | None) -> bool:
    config = _github_config()
    if not config:
        return False
    path = _github_risk_news_path(snapshot_date)
    content_b64 = base64.b64encode(file_bytes).decode("ascii")
    url = f"https://api.github.com/repos/{config['repo']}/contents/{path}"
    body = {
        "message": f"更新 {snapshot_date.strftime('%Y-%m-%d')} 風險新聞明細（{datetime.now(TAIPEI).strftime('%Y-%m-%d %H:%M')}）",
        "content": content_b64,
        "branch": config["branch"],
    }
    if sha:
        body["sha"] = sha
    try:
        response = requests.put(url, headers=_github_headers(config["token"]), json=body, timeout=20)
    except requests.RequestException:
        return False
    return response.status_code in (200, 201)


def save_risk_news_snapshot_to_github(frame: pd.DataFrame, snapshot_date: date) -> bool:
    """把某一天的風險新聞明細整份存成 Excel 檔回存 GitHub，每次都整批覆蓋（不是累加）。
    每天爬蟲跑完會自動呼叫一次；之後若有人工覆核修改，也會呼叫這個函式重新整批覆蓋，
    確保存的版本永遠反映最新的人工修正結果（新的蓋掉舊的）。沒設定 GitHub Secrets 就直接跳過。"""
    if not _github_config() or frame is None or frame.empty:
        return False
    path = _github_risk_news_path(snapshot_date)
    sha = _github_get_file_sha(path)
    file_bytes = _dataframe_to_excel_bytes(frame)
    return _github_push_risk_news_snapshot(file_bytes, snapshot_date, sha)


def _load_negative_history_from_local(output_dir_str: str) -> pd.DataFrame:
    """備用方案：沒設定 GitHub 時，退回掃描本機 output 資料夾（容器重開就會不見）。"""
    output_dir = Path(output_dir_str)
    records = []
    for file in sorted(output_dir.glob("美股_*_負面新聞爬蟲.xlsx")):
        match = re.search(r"美股_(\d{8})_負面新聞爬蟲\.xlsx", file.name)
        if not match:
            continue
        try:
            frame = pd.read_excel(file, sheet_name="負面新聞")
        except Exception:
            continue
        if frame.empty:
            continue
        frame["Level"] = pd.to_numeric(frame.get("Level"), errors="coerce").fillna(0).astype(int)
        frame["Ticker"] = frame.get("Ticker", "").fillna("").astype(str).str.strip()
        frame["Company"] = frame.get("Company", "").fillna("").astype(str).str.strip()
        frame = frame[frame["Ticker"] != ""]
        if frame.empty:
            continue
        grouped = frame.groupby(["Ticker", "Company"]).agg(則數=("Ticker", "size"), 最高Level=("Level", "max")).reset_index()
        grouped["日期"] = pd.to_datetime(match.group(1), format="%Y%m%d")
        records.append(grouped)
    if not records:
        return pd.DataFrame(columns=["Ticker", "Company", "則數", "最高Level", "日期"])
    return pd.concat(records, ignore_index=True)


@st.cache_data(ttl=120)
def load_negative_history(output_dir_str: str) -> pd.DataFrame:
    if _github_config():
        history, _ = _github_fetch_history()
        return history
    return _load_negative_history_from_local(output_dir_str)


# ============================================================
# 區塊：人工覆核 → 修改後的「等級／事件中文／Action」覆寫紀錄
# 說明：負面新聞明細表（event_path 這份 Excel）存在本機 OUTPUT_DIR，
# Streamlit Cloud 容器重開就會消失，所以人工覆核的結果不能只存回那份本機檔案，
# 必須另外寫一份「覆寫紀錄」到 GitHub repo（跟歷史統計用同一組 Secrets），
# 之後每次讀 negative_df 時，都會用這份覆寫紀錄蓋掉原始規則跑出來的值。
# 用 URL 當作每一則新聞的唯一鍵（同一鍵重覆覆核時，以最後一次為準）。
# 沒設定 GitHub Secrets 時，退回存在本機 CONFIG_DIR 的 CSV（容器重開一樣會歸零）。
# ============================================================
OVERRIDE_COLUMNS = ["URL", "Ticker", "Company", "事件中文", "Level", "Action", "覆核時間", "覆核人"]
GITHUB_OVERRIDE_PATH_DEFAULT = "history/negative_news_manual_overrides.csv"
LOCAL_OVERRIDE_PATH = CONFIG_DIR / "Manual_Review_Overrides.csv"


def _github_override_config() -> dict | None:
    base = _github_config()
    if not base:
        return None
    try:
        override_path = st.secrets.get("GITHUB_OVERRIDE_PATH", GITHUB_OVERRIDE_PATH_DEFAULT)
    except Exception:
        override_path = GITHUB_OVERRIDE_PATH_DEFAULT
    return {**base, "path": override_path}


def _github_fetch_overrides() -> tuple[pd.DataFrame, str | None]:
    empty = pd.DataFrame(columns=OVERRIDE_COLUMNS)
    config = _github_override_config()
    if not config:
        return empty, None
    url = f"https://api.github.com/repos/{config['repo']}/contents/{config['path']}"
    try:
        response = requests.get(url, headers=_github_headers(config["token"]), params={"ref": config["branch"]}, timeout=20)
    except requests.RequestException:
        return empty, None
    if response.status_code != 200:
        return empty, None
    payload = response.json()
    sha = payload.get("sha")
    try:
        content = base64.b64decode(payload["content"]).decode("utf-8-sig")
        frame = pd.read_csv(io.StringIO(content), dtype=str)
        for column in OVERRIDE_COLUMNS:
            if column not in frame.columns:
                frame[column] = ""
        return frame[OVERRIDE_COLUMNS], sha
    except Exception:
        return empty, sha


def _github_push_overrides(frame: pd.DataFrame, sha: str | None) -> bool:
    config = _github_override_config()
    if not config:
        return False
    csv_text = frame[OVERRIDE_COLUMNS].to_csv(index=False)
    content_b64 = base64.b64encode(csv_text.encode("utf-8")).decode("ascii")
    url = f"https://api.github.com/repos/{config['repo']}/contents/{config['path']}"
    body = {
        "message": f"人工覆核更新負面新聞等級／事件／Action（{datetime.now(TAIPEI).strftime('%Y-%m-%d %H:%M')}）",
        "content": content_b64,
        "branch": config["branch"],
    }
    if sha:
        body["sha"] = sha
    try:
        response = requests.put(url, headers=_github_headers(config["token"]), json=body, timeout=20)
    except requests.RequestException:
        return False
    return response.status_code in (200, 201)


def _load_overrides_from_local() -> pd.DataFrame:
    if not LOCAL_OVERRIDE_PATH.is_file():
        return pd.DataFrame(columns=OVERRIDE_COLUMNS)
    try:
        frame = pd.read_csv(LOCAL_OVERRIDE_PATH, dtype=str)
    except Exception:
        return pd.DataFrame(columns=OVERRIDE_COLUMNS)
    for column in OVERRIDE_COLUMNS:
        if column not in frame.columns:
            frame[column] = ""
    return frame[OVERRIDE_COLUMNS]


@st.cache_data(ttl=60)
def load_manual_overrides() -> pd.DataFrame:
    if _github_override_config():
        overrides, _ = _github_fetch_overrides()
        return overrides
    return _load_overrides_from_local()


def save_manual_overrides(edited_rows: pd.DataFrame, reviewer: str = "") -> tuple[bool, str]:
    """把使用者在畫面上改過的列（至少含 URL / Ticker / Company / 事件中文 / Level / Action）併入覆寫紀錄並存回。
    回傳 (是否成功, 訊息)。"""
    if edited_rows.empty:
        return False, "沒有偵測到任何變更。"
    edited_rows = edited_rows.copy()
    edited_rows["URL"] = edited_rows["URL"].fillna("").astype(str).str.strip()
    edited_rows = edited_rows[edited_rows["URL"] != ""]
    if edited_rows.empty:
        return False, "變更的列缺少 URL，無法辨識是哪一則新聞。"
    edited_rows["Level"] = pd.to_numeric(edited_rows["Level"], errors="coerce").fillna(0).astype(int).astype(str)
    edited_rows["覆核時間"] = datetime.now(TAIPEI).strftime("%Y-%m-%d %H:%M:%S")
    edited_rows["覆核人"] = reviewer or ""
    edited_rows = edited_rows[OVERRIDE_COLUMNS]

    use_github = bool(_github_override_config())
    if use_github:
        existing, sha = _github_fetch_overrides()
    else:
        existing, sha = _load_overrides_from_local(), None

    if not existing.empty:
        existing = existing[~existing["URL"].isin(edited_rows["URL"])]  # 同一則新聞重覆覆核，以最新一次為準
    merged = pd.concat([existing, edited_rows], ignore_index=True)

    if use_github:
        pushed = _github_push_overrides(merged, sha)
        if not pushed:
            return False, "寫回 GitHub 失敗，請確認 GITHUB_TOKEN／GITHUB_REPO 設定是否正確、Token 是否有 repo 寫入權限。"
    else:
        CONFIG_DIR.mkdir(parents=True, exist_ok=True)
        merged.to_csv(LOCAL_OVERRIDE_PATH, index=False)
    load_manual_overrides.clear()
    return True, f"已儲存 {len(edited_rows):,} 則新聞的人工覆核結果" + ("（GitHub repo，長期保留）。" if use_github else "（本機暫存，重開網站可能會歸零；建議設定 GITHUB_TOKEN／GITHUB_REPO 以永久保存）。")


def apply_manual_overrides(negative_df: pd.DataFrame, overrides: pd.DataFrame) -> pd.DataFrame:
    """用覆寫紀錄（依 URL 對應）蓋掉原始規則跑出來的 事件中文／Level／Action。"""
    if overrides is None or overrides.empty:
        negative_df["已人工覆核"] = False
        return negative_df
    frame = negative_df.copy()
    frame["URL"] = frame["URL"].fillna("").astype(str).str.strip()
    override_map = overrides.drop_duplicates("URL", keep="last").set_index("URL")
    matched = frame["URL"].isin(override_map.index)
    for column in ("事件中文", "Level", "Action"):
        override_values = frame.loc[matched, "URL"].map(override_map[column])
        if column == "Level":
            frame.loc[matched, column] = pd.to_numeric(override_values, errors="coerce").fillna(frame.loc[matched, column]).astype(int)
        else:
            frame.loc[matched, column] = override_values.where(override_values.notna(), frame.loc[matched, column])
    frame["已人工覆核"] = matched
    return frame


def _style_by_level(row: pd.Series) -> list[str]:
    """Level 5 整列標紅、Level 4 整列標橘，其餘不上色。"""
    level = row.get("Level")
    if level == 5:
        return ["background-color: #FFCDD2"] * len(row)
    if level == 4:
        return ["background-color: #FFE0B2"] * len(row)
    return [""] * len(row)


def _format_decimals(value, decimals: int):
    """數字就固定小數位數；非數字（例如 FinBERT 的 N/A）原樣顯示，不出錯。"""
    try:
        return f"{float(value):.{decimals}f}"
    except (TypeError, ValueError):
        return value


def render_manual_review_table(
    source_df: pd.DataFrame,
    section_key: str,
    search_placeholder: str = "搜尋公司名稱、股票代號、標題",
    extra_display_columns: list[str] | None = None,
    auto_backup_date: date | None = None,
) -> None:
    """通用的『可篩選＋人工覆核可編輯』新聞明細表，負面新聞與待人工覆核新聞共用同一套邏輯。
    source_df 必須已經套用過 apply_manual_overrides（含「已人工覆核」欄位）。
    auto_backup_date：若有帶入日期，儲存人工覆核結果時會一併把當天完整風險新聞快照回存 GitHub
    （覆蓋掉當天原本存的版本），只有「風險新聞明細」會傳入這個參數。"""
    filter_cols = st.columns([2, 1, 1])
    search = filter_cols[0].text_input(search_placeholder, key=f"{section_key}_search")

    event_options = ["全部"] + sorted([str(e) for e in source_df["事件中文"].dropna().unique() if str(e).strip()])
    event_filter = filter_cols[1].selectbox("事件種類", event_options, key=f"{section_key}_event_filter")

    level_options = ["全部", "Level 5", "Level 4", "Level 3", "Level 2", "Level 1"]
    level_filter = filter_cols[2].selectbox("事件等級", level_options, key=f"{section_key}_level_filter")

    filtered = source_df.copy()
    if search:
        mask = filtered[["Ticker", "Company", "Title", "事件中文"]].fillna("").astype(str).apply(
            lambda column: column.str.contains(search, case=False, regex=False)
        ).any(axis=1)
        filtered = filtered[mask]
    if event_filter != "全部":
        filtered = filtered[filtered["事件中文"] == event_filter]
    if level_filter.startswith("Level "):
        filtered = filtered[filtered["Level"] == int(level_filter.replace("Level ", ""))]

    st.caption(f"篩選結果：顯示 {len(filtered):,}／{len(source_df):,} 則　🔴 Level 5　🟠 Level 4")

    display_columns = ["Published Time", "Ticker", "Company", "事件中文", "Level", "Title_ZH", "FinBERT", "Action", "URL"]
    if "已人工覆核" in filtered.columns:
        display_columns.insert(4, "已人工覆核")
    if extra_display_columns:
        insert_at = display_columns.index("事件中文")
        for offset, column in enumerate(extra_display_columns):
            if column in filtered.columns:
                display_columns.insert(insert_at + offset, column)

    review_mode = st.toggle("✏️ 人工覆核模式（可修改等級／事件中文／Action）", key=f"{section_key}_review_mode")

    if not review_mode:
        styled = filtered[display_columns].style.apply(_style_by_level, axis=1)
        format_dict = {}
        if "FinBERT" in display_columns:
            format_dict["FinBERT"] = lambda v: _format_decimals(v, 3)
        if "曝險金額" in display_columns:
            format_dict["曝險金額"] = lambda v: _format_decimals(v, 2)
        if format_dict:
            styled = styled.format(format_dict)
        st.dataframe(
            styled,
            use_container_width=True, hide_index=True,
            column_config={
                "URL": st.column_config.LinkColumn("新聞", display_text="開啟"),
                "曝險金額": st.column_config.NumberColumn("曝險金額", format="%.2f"),
                "FinBERT": st.column_config.NumberColumn("FinBERT", format="%.3f"),
            },
            height=560,
        )
        return

    st.caption("直接在下表修改「事件中文」「Level」「Action」欄位，改完按下方「儲存人工覆核結果」即可回存；下次讀取新聞時會自動套用。")
    reviewer_name = st.text_input("覆核人（選填，會一併記錄）", key=f"{section_key}_reviewer_name")
    editor_columns = ["Published Time", "Ticker", "Company", "Title_ZH", "FinBERT", "事件中文", "Level", "Action", "URL"]
    editable_frame = filtered[editor_columns].copy().reset_index(drop=True)
    edited_frame = st.data_editor(
        editable_frame,
        use_container_width=True, hide_index=True, height=560,
        key=f"{section_key}_editor",
        disabled=["Published Time", "Ticker", "Company", "Title_ZH", "FinBERT", "URL"],
        column_config={
            "URL": st.column_config.LinkColumn("新聞", display_text="開啟"),
            "事件中文": st.column_config.TextColumn("事件中文"),
            "Level": st.column_config.SelectboxColumn("Level", options=[1, 2, 3, 4, 5]),
            "Action": st.column_config.TextColumn("Action", width="large"),
            "FinBERT": st.column_config.NumberColumn("FinBERT", format="%.3f"),
        },
    )
    if st.button("💾 儲存人工覆核結果", use_container_width=True, key=f"{section_key}_save_manual_review"):
        comparable_cols = ["事件中文", "Level", "Action"]
        original_indexed = editable_frame.set_index("URL")
        edited_indexed = edited_frame.set_index("URL")
        changed_urls = [
            url for url in edited_indexed.index
            if url in original_indexed.index
            and not original_indexed.loc[url, comparable_cols].equals(edited_indexed.loc[url, comparable_cols])
        ]
        if not changed_urls:
            st.info("沒有偵測到任何欄位變更。")
        else:
            changed_rows = edited_frame[edited_frame["URL"].isin(changed_urls)][["URL", "Ticker", "Company", "事件中文", "Level", "Action"]]
            success, save_message = save_manual_overrides(changed_rows, reviewer=reviewer_name)
            if success:
                if auto_backup_date is not None:
                    # 把這次修改併回完整風險新聞明細，整批覆蓋回存到 GitHub 當天的快照
                    updated_snapshot = source_df.copy()
                    updated_snapshot["URL"] = updated_snapshot["URL"].fillna("").astype(str).str.strip()
                    for _, changed_row in changed_rows.iterrows():
                        row_mask = updated_snapshot["URL"] == changed_row["URL"]
                        updated_snapshot.loc[row_mask, "事件中文"] = changed_row["事件中文"]
                        updated_snapshot.loc[row_mask, "Level"] = changed_row["Level"]
                        updated_snapshot.loc[row_mask, "Action"] = changed_row["Action"]
                        updated_snapshot.loc[row_mask, "已人工覆核"] = True
                    snapshot_pushed = save_risk_news_snapshot_to_github(updated_snapshot, auto_backup_date)
                    if _github_config() and not snapshot_pushed:
                        save_message += "（提醒：當天風險新聞快照回存 GitHub 失敗，人工覆核結果本身已存好，不受影響。）"
                st.success(save_message)
                st.rerun()
            else:
                st.error(save_message)


# ============================================================
# 區塊：頁面標題文字
# 想改頁面上方的標題／說明文字，改這裡即可。
# ============================================================
def page_heading() -> None:
    st.markdown(
        "<div class='page-kicker'>今日任務</div>"
        "<div class='page-title'>開始今日新聞擷取</div>",
        
        unsafe_allow_html=True,
    )


# ============================================================
# 區塊：網站初始化 ＋ 版面樣式（CSS）
# 想調整字級、顏色、卡片外觀，改下面 <style> 區塊即可。
# ============================================================
initialize()
st.set_page_config(page_title=APP_TITLE, page_icon="🛡️", layout="wide")
st.markdown("""
<style>
  .stApp { background:#F7F9FC; }
  .block-container { max-width:1500px; padding-top:1.4rem; padding-bottom:4rem; }
  .app-hero { background:linear-gradient(120deg,#102A43,#1D4ED8); color:white; border-radius:16px; padding:20px 26px; margin-bottom:14px; box-shadow:0 10px 30px rgba(15,42,67,.12); }
  .app-hero-title { font-size:26px; font-weight:750; letter-spacing:.02em; }
  .app-hero-sub { opacity:.78; font-size:15px; margin-top:1px; margin-left:44px; }
  .page-kicker { color:#2563EB; font-size:13px; font-weight:700; letter-spacing:.08em; margin-top:24px; }
  .page-title { color:#172B4D; font-size:31px; line-height:1.25; font-weight:780; margin-top:4px; }
  .page-description { color:#64748B; font-size:16px; margin:7px 0 20px; }
  div[data-testid='stRadio'] > div { background:white; border:1px solid #E2E8F0; border-radius:13px; padding:10px; box-shadow:0 3px 12px rgba(15,23,42,.04); }
  div[data-testid='stRadio'] div[role='radiogroup'] { display:flex; gap:8px; width:100%; }
  div[data-testid='stRadio'] label[data-baseweb='radio'] { flex:1; min-height:50px; display:flex; align-items:center; justify-content:center; border-radius:9px; padding:8px 10px; color:#64748B; font-weight:700; cursor:pointer; transition:all .16s ease; }
  div[data-testid='stRadio'] label[data-baseweb='radio']:hover { background:#F1F5F9; color:#1D4ED8; }
  div[data-testid='stRadio'] label[data-baseweb='radio']:has(input:checked) { background:#EAF2FF; color:#1D4ED8; box-shadow:inset 0 0 0 1px #BFDBFE; }
  div[data-testid='stRadio'] label[data-baseweb='radio'] > div:first-child { display:none; }
  .method-card { background:white; border:1px solid #DBEAFE; border-left:6px solid #2563EB; border-radius:12px; padding:16px 18px; margin:4px 0 20px; }
  .method-title { color:#1E3A8A; font-weight:750; font-size:17px; }
  .method-desc { color:#64748B; font-size:14px; margin-top:5px; }
  .status-title { color:#172B4D; font-size:27px; line-height:1.15; font-weight:750; margin:0 0 4px; }
  .previous-result-grid { display:grid; grid-template-columns:repeat(5,minmax(0,1fr)); gap:16px; margin:4px 0 16px; }
  .previous-result-card { background:white; border:1px solid #E2E8F0; border-radius:12px; padding:16px 18px; min-height:108px; box-shadow:0 2px 8px rgba(15,23,42,.035); }
  .previous-result-label { color:#334155; font-size:15px; margin-bottom:10px; }
  .previous-result-value { color:#1E293B; font-size:24px; line-height:1.3; font-weight:520; white-space:normal; overflow-wrap:anywhere; }
  div[data-testid='stMetric'] { background:white; border:1px solid #E2E8F0; border-radius:12px; padding:14px 16px; box-shadow:0 2px 8px rgba(15,23,42,.035); }
  div[data-testid='stFileUploader'] { background:white; border-radius:12px; }
  div[data-testid='stDataFrame'] { border:1px solid #E2E8F0; border-radius:12px; overflow:hidden; }
  div.stButton > button[kind='primary'] { min-height:46px; border-radius:10px; font-weight:700; }
  div.stDownloadButton > button { min-height:43px; border-radius:10px; font-weight:650; }
  /* 三個下載按鈕分開設定顏色，目前先暫訂統一藍色，之後要改各自顏色只要調整這三段即可 */
  .st-key-saved_crawl_all button { background:#1D4ED8; color:white; border-color:#1D4ED8; }
  .st-key-saved_crawl_all button:hover { background:#1E40AF; border-color:#1E40AF; color:white; }
  .st-key-saved_crawl_event button { background:#1D4ED8; color:white; border-color:#1D4ED8; }
  .st-key-saved_crawl_event button:hover { background:#1E40AF; border-color:#1E40AF; color:white; }
  .st-key-saved_crawl_finbert button { background:#1D4ED8; color:white; border-color:#1D4ED8; }
  .st-key-saved_crawl_finbert button:hover { background:#1E40AF; border-color:#1E40AF; color:white; }
  h3 { color:#172B4D !important; margin-top:1.4rem !important; }
  @media (max-width:1000px) { .previous-result-grid { grid-template-columns:repeat(2,minmax(0,1fr)); } }
  @media (max-width:800px) { div[data-testid='stRadio'] div[role='radiogroup'] { display:block; } div[data-testid='stRadio'] label[data-baseweb='radio'] { justify-content:flex-start; } .page-title { font-size:25px; } .previous-result-grid { grid-template-columns:1fr; } }
</style>
""", unsafe_allow_html=True)
hero_subtitle = f"US Stock Market Negative News Aggregator"
st.markdown(f"<div class='app-hero'><div class='app-hero-title'>🛡️ {APP_TITLE}</div><div class='app-hero-sub'>{hero_subtitle}</div></div>", unsafe_allow_html=True)

page_heading()
st.markdown("<div class='method-card'><div class='method-title'>運作流程說明</div><div class='method-desc'>系統自動抓取多個新聞來源，比對公司列表並去除重複新聞，接著以金融情緒分析 (FinBERT) 模型分析新聞情緒，篩出可能影響公司之負面消息，並依規則分類事件等級。<br>若有上傳曝險部位，會進一步標出曝險高、事件嚴重的公司，同時比對歷史資料新聞頻率異常，最後彙整成圖表與 Excel 檔。</div></div>", unsafe_allow_html=True)

# ============================================================
# 區塊：今日任務頁面（唯一頁面）
# 1. 選公司範圍 → 2. 設定期間 → 3. 選執行方法並執行 → 顯示進度與下載結果
# ============================================================
if True:
    st.markdown("### 1. 選取公司列表")
    universe = st.radio("選擇欲執行的公司列表", ["Dow Jones 30", "S&P 500", "上傳公司列表"], index=1, horizontal=True, label_visibility="collapsed")
    st.caption("Dow Jones 30／S&P 500 讀取自 GitHub repo 內 Excel 檔。")
    if universe == "上傳公司列表":
        company_upload = st.file_uploader("上傳自訂列表 Excel 檔，需含 Ticker (Symbol) 與 Company (Name) 欄位。", type=["xlsx"], key="company_list_upload")
        if st.button("檢查並套用", use_container_width=True, disabled=company_upload is None):
            try:
                checked_companies = load_company_list_from_upload(company_upload)
                if checked_companies.empty:
                    raise ValueError("公司名單沒有可用資料")
                CUSTOM_COMPANY_PATH.write_bytes(company_upload.getvalue())
                load_company_list.clear()
                st.success(f"自訂公司名單已套用：{company_upload.name}，共 {len(checked_companies):,} 家。")
            except Exception as exc:
                st.error(f"自訂公司名單無法套用：{exc}")
        elif CUSTOM_COMPANY_PATH.is_file():
            st.caption("已套用過自訂公司名單，重新上傳並按「檢查並套用」可覆蓋。")
    now = datetime.now(TAIPEI).replace(hour=9, minute=0, second=0, microsecond=0)
    previous_end = last_success_end()
    default_start = (previous_end or (now - timedelta(days=3))).replace(hour=9, minute=0, second=0, microsecond=0)
    if previous_end:
        st.caption(f"最近執行日期：{previous_end:%Y-%m-%d}")
    st.markdown("### 2. 設定新聞期間")
    start_col, end_col = st.columns(2)
    start_date = start_col.date_input("開始日期", value=default_start.date())
    start_hour = start_col.selectbox("開始時間", list(range(24)), index=default_start.hour, format_func=lambda hour: f"{hour:02d}:00")
    end_date = end_col.date_input("結束日期", value=now.date())
    end_hour = end_col.selectbox("結束時間", list(range(24)), index=now.hour, format_func=lambda hour: f"{hour:02d}:00")
    start_dt = datetime.combine(start_date, datetime.min.time(), TAIPEI) + timedelta(hours=start_hour)
    end_dt = datetime.combine(end_date, datetime.min.time(), TAIPEI) + timedelta(hours=end_hour)
    if start_dt > end_dt:
        st.error("開始時間不可晚於結束時間")
    st.caption(f"設定時間：{start_dt:%Y-%m-%d %H:%M} ～ {end_dt:%Y-%m-%d %H:%M}（台北時間）")
    st.markdown("### 3. 選擇執行方法")
    crawl_method = st.radio("選擇擷取方法", ["方法一｜多網站擷取", "方法二｜Google News", "方法一＋方法二｜完整整合"], index=2, horizontal=True, label_visibility="collapsed")
    if crawl_method == "方法一｜多網站擷取":
        st.markdown("<div class='method-card'><div class='method-title'>方法一｜多網站擷取</div><div class='method-desc'>彙整 MoneyDJ、經濟日報、鉅亨網、CNBC 與 TradingView。</div></div>", unsafe_allow_html=True)
    elif crawl_method == "方法二｜Google News":
        st.markdown("<div class='method-card'><div class='method-title'>方法二｜Google News</div><div class='method-desc'>彙整 Google News RSS 英文新聞，並執行 FinBERT 分析。</div></div>", unsafe_allow_html=True)
    else:
        st.markdown("<div class='method-card'><div class='method-title'>方法一＋方法二｜完整整合</div><div class='method-desc'>依序完成兩種方法、合併排除重複，並執行 FinBERT 分析。</div></div>", unsafe_allow_html=True)
    method_name = {"方法一｜多網站擷取": "方法一", "方法二｜Google News": "方法二", "方法一＋方法二｜完整整合": "方法一＋方法二"}[crawl_method]
    registry = crawl_job_registry()
    current_job = registry.get("current")
    is_running = bool(current_job and current_job.get("status") in ("running", "stopping"))
    if st.button(
        f"開始執行 {method_name}", type="primary",
        disabled=start_dt > end_dt or is_running or (universe == "上傳公司列表" and not CUSTOM_COMPANY_PATH.is_file()),
        use_container_width=True,
    ):
        try:
            if universe.startswith("Dow"):
                company_path = DOWJONES_PATH
            elif universe.startswith("S&P"):
                company_path = SP500_PATH
            else:
                company_path = CUSTOM_COMPANY_PATH
            companies = load_company_list(str(company_path), company_path.stat().st_mtime_ns)
            if companies.empty:
                raise ValueError(f"{company_path.name} 沒有可用的公司資料")
            run_id = create_run(start_dt, end_dt, f"{method_name}｜{universe}")
            script_context = get_script_run_ctx()
            use_method1 = method_name in ("方法一", "方法一＋方法二")
            use_method2 = method_name in ("方法二", "方法一＋方法二")
            job = {
                "run_id": run_id, "status": "running", "method": method_name,
                "started_monotonic": time.monotonic(), "started_at": datetime.now(TAIPEI).isoformat(timespec="seconds"),
                "started_epoch_ms": int(time.time() * 1000),
                "progress": 0.0, "progress_text": "準備開始", "stop_event": threading.Event(),
                "stage_times": [], "current_stage": None, "current_stage_started": None,
                "script_context": script_context,
                "session_id": script_context.session_id if script_context is not None else "",
                "method1_monitor": {
                    "status": "pending" if use_method1 else "skipped", "stage": "等待執行" if use_method1 else "本次未執行",
                    "rows": 0, "errors": [], "source_counts": {},
                },
                "method2_monitor": {
                    "status": "pending" if use_method2 else "skipped", "stage": "等待執行" if use_method2 else "本次未執行",
                    "rows": 0, "raw_rows": 0, "negative_rows": 0, "batches": 0, "errors": [],
                },
            }
            registry["current"] = job
            worker = threading.Thread(
                target=run_crawl_job,
                args=(job, method_name, universe, companies.copy(), start_dt, end_dt),
                name=f"news-crawl-{run_id}", daemon=True,
            )
            job["thread"] = worker
            worker.start()
            st.rerun()
        except Exception as exc:
            st.error(f"無法開始抓取：{exc}")

    # 執行狀態卡片獨立成一個 fragment，每 2 秒自動局部刷新一次，
    # 不需要使用者手動按「更新目前進度」；只有這張卡片會重繪，
    # 上面公司範圍／期間等其他設定不會被打斷或重置。
    @st.fragment(run_every=2)
    def render_crawl_status():
        job = crawl_job_registry().get("current")
        if not job:
            return False
        end_tick = job.get("finished_monotonic", time.monotonic())
        elapsed_text = format_elapsed(end_tick - job["started_monotonic"])
        status = job.get("status", "running")
        with st.container(border=True):
            left, right = st.columns([3, 1])
            with left:
                st.markdown("<div class='status-title'>執行狀態</div>", unsafe_allow_html=True)
                st.progress(job.get("progress", 0), text=job.get("progress_text", "執行中"))
            if status in ("running", "stopping"):
                started_epoch_ms = int(job.get("started_epoch_ms", time.time() * 1000))
                with right:
                    components.html(
                        f"""
                        <div style="font-family:Arial,sans-serif;border:1px solid #e2e8f0;border-radius:12px;
                                    padding:10px 16px;background:white;height:78px;box-sizing:border-box;">
                          <div style="font-size:14px;color:#334155;margin-bottom:3px;">累計執行時間</div>
                          <div id="crawl-elapsed" style="font-size:26px;color:#1e293b;line-height:1.08;">{elapsed_text}</div>
                        </div>
                        <script>
                          const started = {started_epoch_ms};
                          const pad = n => String(n).padStart(2, '0');
                          const update = () => {{
                            const total = Math.max(0, Math.floor((Date.now() - started) / 1000));
                            const h = Math.floor(total / 3600);
                            const m = Math.floor((total % 3600) / 60);
                            const s = total % 60;
                            document.getElementById('crawl-elapsed').textContent = `${{pad(h)}}:${{pad(m)}}:${{pad(s)}}`;
                          }};
                          update(); setInterval(update, 1000);
                        </script>
                        """,
                        height=82,
                    )
            else:
                right.metric("累計執行時間", elapsed_text)

            stage_rows = [
                {"執行階段": item["label"], "所用時間": format_elapsed(item["seconds"]), "狀態": "已完成"}
                for item in job.get("stage_times", [])
            ]
            current_stage = job.get("current_stage")
            current_stage_started = job.get("current_stage_started")
            if current_stage and current_stage_started is not None:
                stage_rows.append({
                    "執行階段": current_stage,
                    "所用時間": format_elapsed(time.monotonic() - float(current_stage_started)),
                    "狀態": "執行中" if status == "running" else "停止中",
                })
            button_labels = {
                "running": "停止執行", "stopping": "停止中…", "success": "已完成",
                "stopped": "已停止", "failed": "執行失敗",
            }
            stop_clicked = st.button(
                button_labels.get(status, "目前不可用"),
                type="secondary", use_container_width=True, key="stop_crawl",
                disabled=status != "running",
            )
            if stop_clicked:
                job["status"] = "stopping"
                job["progress_text"] = "正在安全停止，請稍候…"
                job["stop_event"].set()
            history_note = ""
            if _github_config():
                history_note = "；歷史統計已同步至 GitHub" if job.get("history_pushed") else "；歷史統計同步 GitHub 失敗（請檢查 Secrets 設定）"
                history_note += "；風險新聞明細已同步至 GitHub" if job.get("risk_news_snapshot_pushed") else "；風險新聞明細同步 GitHub 失敗"
            status_messages = {
                "running": "⏳ 程式執行中，可於本頁查看最新進度。",
                "stopping": "🚫 已停止執行。",
                "success": (
                    f"{job.get('summary', '抓取完成')}；負面新聞 {job.get('event_rows', 0)} 則；"
                    f"待人工覆核 {job.get('unknown_rows', 0)} 則；無關新聞 {job.get('irrelevant_rows', 0)} 則。"
                    f"翻譯失敗 {job.get('translation_failures', 0)} 則。"
                    f"總耗時 {elapsed_text}{history_note}"
                ),
                "stopped": f"🚫 已停止執行。總耗時 {elapsed_text}；未完成的結果不會覆蓋前次紀錄。",
                "failed": f"抓取失敗：{job.get('error', '未知錯誤')}（耗時 {elapsed_text}）",
            }
            status_message = status_messages.get(status, "正在更新執行狀態…")
            if status == "success":
                st.success(f"✅ 執行完成｜{status_message}")
            elif status == "failed":
                st.error(status_message)
            elif status in ("stopping", "stopped"):
                st.warning(status_message)
            else:
                st.info(status_message)
        return False

    auto_refresh_running = render_crawl_status()
    saved_crawl = latest_crawl_result()
# ============================================================
# 區塊：負面新聞總覽（KPI、事件類型/公司排行圖表、可篩選明細表）
# 直接讀取剛才產生的「負面新聞」工作表來畫圖。
# ============================================================
if saved_crawl and saved_crawl["event_path"] and saved_crawl["event_path"].is_file():
            st.markdown("### 4. 負面新聞總覽")
            with st.expander("上傳曝險部位清單（選填）", expanded=False):
                exposure_upload = st.file_uploader(
                    "上傳曝險清單 Excel，需含 Ticker (Symbol) 與 Exposure (Amount) 欄位；"
                    "可另外加一欄「業務別」（財管／經紀／自營），之後就能依業務別分開檢視曝險",
                    type=["xlsx"], key="exposure_upload",
                )
                if st.button("套用曝險清單", use_container_width=True, disabled=exposure_upload is None, key="apply_exposure"):
                    try:
                        checked_exposure = load_exposure_list_from_upload(exposure_upload)
                        if checked_exposure.empty:
                            raise ValueError("曝險清單沒有可用資料")
                        EXPOSURE_PATH.write_bytes(exposure_upload.getvalue())
                        load_exposure_list.clear()
                        st.success(f"曝險清單已套用，共 {len(checked_exposure):,} 家公司。")
                    except Exception as exc:
                        st.error(f"曝險清單無法套用：{exc}")
                elif EXPOSURE_PATH.is_file():
                    st.caption("已套用曝險清單，重新上傳並按「套用曝險清單」可覆蓋。")
            negative_df = pd.read_excel(saved_crawl["event_path"], sheet_name="負面新聞")
            snapshot_date = datetime.fromisoformat(saved_crawl["end_time"]).astimezone(TAIPEI).date()
            if negative_df.empty:
                st.caption("這次沒有新聞命中負面事件規則。")
            else:
                negative_df["Level"] = pd.to_numeric(negative_df["Level"], errors="coerce").fillna(0).astype(int)
                negative_df["Ticker"] = negative_df["Ticker"].fillna("").astype(str).str.strip()

                # 套用人工覆核紀錄：用 URL 對應，蓋掉原始規則跑出來的 事件中文／Level／Action
                negative_df = apply_manual_overrides(negative_df, load_manual_overrides())

                # 併入曝險金額（若有上傳曝險清單；可依「業務別」切換只看財管／經紀／自營其中一種）
                if EXPOSURE_PATH.is_file():
                    exposure_df = load_exposure_list(str(EXPOSURE_PATH), EXPOSURE_PATH.stat().st_mtime_ns)
                else:
                    exposure_df = pd.DataFrame(columns=["Ticker", "業務別", "Exposure"])

                business_options = exposure_business_options(exposure_df)
                if len(business_options) > 1:
                    business_filter = st.selectbox(
                        "依業務別查看曝險（影響下方氣泡圖與新聞明細的曝險金額欄位）",
                        business_options, key="exposure_business_filter",
                    )
                else:
                    business_filter = BUSINESS_LINE_ALL

                exposure_map = exposure_map_for_business(exposure_df, business_filter)
                negative_df["曝險金額"] = negative_df["Ticker"].str.upper().map(exposure_map).fillna(0.0)

                # 最左邊第一格加上總新聞數量，並移除持續追蹤
                metric_cols = st.columns(4)
                metric_cols[0].metric("總新聞", f"{saved_crawl['rows']:,} 則")
                metric_cols[1].metric("負面新聞", f"{len(negative_df):,} 則")
                metric_cols[2].metric("重大事件（Level ≥ 4）", f"{int((negative_df['Level'] >= 4).sum()):,} 則")
                metric_cols[3].metric("涉及公司", f"{negative_df['Ticker'].replace('', pd.NA).nunique():,} 家")
                if exposure_map:
                    exposed_companies = negative_df[negative_df["曝險金額"] > 0]["Ticker"].nunique()
                    total_exposure_hit = negative_df.drop_duplicates("Ticker")["曝險金額"].sum()
                    business_label = f"（業務別：{business_filter}）" if business_filter != BUSINESS_LINE_ALL else ""
                    st.caption(f"曝險加權{business_label}：本次負面新聞涉及 {exposed_companies:,} 家有曝險之公司，合計曝險金額 {total_exposure_hit:,.0f}。")

                chart_left, chart_right = st.columns(2)
                import plotly.express as px

                # 1. 左圖：事件類型分布（加入「其他」類別以完整呈現代碼 100% 占比）
                chart_left.markdown("#### 事件類型總占比")
                all_event_counts = negative_df["事件中文"].fillna("未分類").value_counts()

                if len(all_event_counts) > 10:
                    top_9 = all_event_counts.head(9)
                    other_count = all_event_counts.iloc[9:].sum()
                    event_counts_df = pd.concat([top_9, pd.Series({"其他": other_count})]).reset_index()
                else:
                    event_counts_df = all_event_counts.reset_index()

                event_counts_df.columns = ["事件中文", "數量"]

                fig_donut = px.pie(
                    event_counts_df,
                    values="數量",
                    names="事件中文",
                    hole=0.5,
                    color_discrete_sequence=px.colors.qualitative.Pastel
                )
                fig_donut.update_traces(
                    textposition='inside',
                    textinfo='percent+label',
                    hovertemplate="<b>%{label}</b><br>數量: %{value} 則 (%{percent})"
                )
                fig_donut.update_layout(
                    showlegend=False,
                    height=320,
                    margin=dict(t=20, b=20, l=20, r=20)
                )
                chart_left.plotly_chart(fig_donut, use_container_width=True)

                # 2. 右圖：公司負面新聞排行（對齊左圖高度與邊距）
                chart_right.markdown("#### 公司負面新聞排行（前 10）")
                tickers = negative_df["Ticker"].fillna("").astype(str).str.strip()
                companies = negative_df["Company"].fillna("").astype(str).str.strip()
                company_label = pd.Series(
                    (tickers + "｜" + companies).where(tickers != "", companies).values,
                    index=negative_df.index,
                )
                company_counts = company_label[company_label != ""].value_counts().head(10).reset_index()
                company_counts.columns = ["公司", "數量"]

                fig_bar = px.bar(
                    company_counts,
                    x="公司",
                    y="數量",
                    text="數量",
                    color_discrete_sequence=["#1D4ED8"]
                )
                fig_bar.update_traces(textposition='outside')
                fig_bar.update_layout(
                    xaxis_title=None,
                    yaxis_title=None,
                    height=320,
                    margin=dict(t=20, b=20, l=20, r=20),
                    xaxis=dict(tickangle=-45)
                )
                chart_right.plotly_chart(fig_bar, use_container_width=True)

                # 3. 曝險風險象限圖（只有上傳過曝險清單才顯示）
                # X 軸＝曝險金額、Y 軸＝風險嚴重度（最高Level × 負面新聞則數），
                # 用象限（中位數切分）取代單一合成分數，避免公式設計不夠嚴謹時誤導判斷。
                if exposure_map:
                    business_heading = f"（業務別：{business_filter}）" if business_filter != BUSINESS_LINE_ALL else ""
                    st.markdown(f"#### 曝險風險象限氣泡圖{business_heading}")
                    exposure_rank = (
                        negative_df[negative_df["曝險金額"] > 0]
                        .groupby(["Ticker", "Company"])
                        .agg(曝險金額=("曝險金額", "max"), 負面新聞則數=("Ticker", "size"), 最高Level=("Level", "max"))
                        .reset_index()
                    )
                    if exposure_rank.empty:
                        st.caption("這次負面新聞沒有命中曝險清單裡的公司。")
                    else:
                        exposure_rank["風險嚴重度"] = exposure_rank["最高Level"] * exposure_rank["負面新聞則數"]
                        fig_exposure = px.scatter(
                            exposure_rank, x="曝險金額", y="風險嚴重度",
                            size="負面新聞則數", color="最高Level", text="Ticker",
                            hover_name="Company",
                            hover_data={"Ticker": True, "負面新聞則數": True, "最高Level": True, "曝險金額": ":,.0f", "風險嚴重度": True},
                            color_continuous_scale=["#BFDBFE", "#3B82F6", "#1E3A8A"],
                            size_max=36,
                        )
                        fig_exposure.update_traces(textposition="top center")
                        exposure_median = exposure_rank["曝險金額"].median()
                        severity_median = exposure_rank["風險嚴重度"].median()
                        fig_exposure.add_vline(x=exposure_median, line_dash="dash", line_color="#94A3B8")
                        fig_exposure.add_hline(y=severity_median, line_dash="dash", line_color="#94A3B8")
                        fig_exposure.update_layout(
                            xaxis_title="曝險金額", yaxis_title="風險嚴重度（Level × 則數）", height=380,
                            margin=dict(t=20, b=20, l=20, r=20),
                        )
                        st.plotly_chart(fig_exposure, use_container_width=True)
                        st.caption(
                            "氣泡大小＝負面新聞則數多寡；虛線為中位數"
                        )


                # 4. 新聞頻率異常偵測 + 5. 時間序列走勢圖
                # 歷史資料來源：已設定 GitHub Secrets 時讀 GitHub repo 裡的統計檔（不會因容器重開而消失）；
                # 沒設定時退回讀本機 output 資料夾（容器重開就會歸零）。
                history_df = load_negative_history(str(OUTPUT_DIR))
                available_days = sorted(history_df["日期"].dt.date.unique()) if not history_df.empty else []
                st.markdown("#### 新聞頻率異常與走勢")
                history_source_note = "（歷史資料來源：GitHub repo，長期保留）" if _github_config() else "（歷史資料來源：本機暫存，重開網站可能會歸零）"
                if len(available_days) < 2:
                    st.caption(
                        f"目前累積 {len(available_days)} 天歷史負面新聞資料，至少需要 2 天才能比較頻率變化。"
                        f"隨著每天執行，這裡會自動累積出趨勢圖與異常警示 {history_source_note}。"
                    )
                else:
                    st.caption(history_source_note)
                    latest_day = available_days[-1]
                    baseline_days = [day for day in available_days if day != latest_day]
                    latest_counts = history_df[history_df["日期"].dt.date == latest_day].groupby(["Ticker", "Company"])["則數"].sum()
                    baseline_avg = (
                        history_df[history_df["日期"].dt.date.isin(baseline_days)]
                        .groupby(["Ticker", "Company"])["則數"].sum() / max(len(baseline_days), 1)
                    )
                    anomaly_rows = []
                    for (ticker, company), latest_count in latest_counts.items():
                        base = float(baseline_avg.get((ticker, company), 0.0))
                        is_new_spike = base == 0 and latest_count >= 3
                        is_ratio_spike = base > 0 and latest_count >= base * 2 and latest_count >= 2
                        if is_new_spike or is_ratio_spike:
                            anomaly_rows.append({
                                "Ticker": ticker, "Company": company,
                                f"{latest_day} 則數": int(latest_count),
                                "過去平均(每日)": round(base, 1),
                                "倍數": "新增" if base == 0 else f"{latest_count / base:.1f}x",
                            })
                    if anomaly_rows:
                        st.warning(f"🔺 偵測到 {len(anomaly_rows)} 家公司負面新聞則數異常放量（{latest_day} vs. 過去 {len(baseline_days)} 天平均）：")
                        st.dataframe(pd.DataFrame(anomaly_rows), use_container_width=True, hide_index=True, height=300)
                    else:
                        st.caption(f"{latest_day} 沒有公司出現明顯的負面新聞異常放量（門檻：較過去平均高 2 倍以上，或平日無負面新聞卻突然出現 3 則以上）。")

                    trend_options = ["全市場"] + sorted(
                        (history_df["Ticker"] + "｜" + history_df["Company"]).unique().tolist()
                    )
                    trend_pick = st.selectbox("查看頻率趨勢", trend_options, key="trend_company_pick")
                    if trend_pick == "全市場":
                        trend_data = history_df.groupby("日期", as_index=False)["則數"].sum()
                    else:
                        pick_ticker = trend_pick.split("｜", 1)[0]
                        trend_data = history_df[history_df["Ticker"] == pick_ticker].groupby("日期", as_index=False)["則數"].sum()
                    fig_trend = px.line(trend_data, x="日期", y="則數", markers=True)
                    fig_trend.update_layout(xaxis_title=None, yaxis_title="負面新聞則數", height=310, margin=dict(t=20, b=20, l=20, r=20))
                    st.plotly_chart(fig_trend, use_container_width=True)

                st.markdown("#### 風險新聞明細")
                render_manual_review_table(
                    negative_df,
                    section_key="negative",
                    extra_display_columns=["曝險金額"] if exposure_map else None,
                    auto_backup_date=snapshot_date,
                )

            # ============================================================
            # 區塊：待人工覆核新聞（規則沒能自動分類的新聞）
            # 跟「負面新聞」共用同一份 event_path Excel（分頁「待人工覆核」），
            # 也共用同一套人工覆核回存機制（用 URL 當唯一鍵，寫回同一份覆寫紀錄）。
            # 獨立於上面「負面新聞」是否有資料，只要這次執行有 event_path 就顯示。
            # ============================================================
            try:
                unknown_df = pd.read_excel(saved_crawl["event_path"], sheet_name="待人工覆核")
            except Exception:
                unknown_df = pd.DataFrame(columns=NEGATIVE_OUTPUT_COLUMNS)
            st.markdown("#### 待人工覆核新聞")
            if unknown_df.empty:
                st.caption("這次沒有待人工覆核的新聞。")
            else:
                unknown_df["Level"] = pd.to_numeric(unknown_df["Level"], errors="coerce").fillna(0).astype(int)
                unknown_df["Ticker"] = unknown_df["Ticker"].fillna("").astype(str).str.strip()
                unknown_df = apply_manual_overrides(unknown_df, load_manual_overrides())
                render_manual_review_table(unknown_df, section_key="pending_review")
else:
        current_job = crawl_job_registry().get("current")
        if current_job and current_job.get("status") in ("running", "stopping"):
            st.info("⏳ 執行中，請稍後再查看「新聞總覽」圖表與明細。")
        elif saved_crawl:
            st.info("本次無可用的新聞檔案，暫無法顯示圖表。")
        else:
            st.info("尚無執行紀錄。首次執行後，結果會保留於此，重新整理或重開網頁也能下載檔案。")


# ============================================================
# 區塊：5. 下載執行結果
# ============================================================
if saved_crawl:
        current_job = crawl_job_registry().get("current")
        is_current_result = bool(
            current_job
            and current_job.get("status") == "success"
            and current_job.get("path")
            and Path(current_job["path"]) == saved_crawl["path"]
        )
        st.markdown("### 5. 下載執行結果")
        try:
            saved_start = datetime.fromisoformat(saved_crawl["start_time"]).astimezone(TAIPEI)
            saved_end = datetime.fromisoformat(saved_crawl["end_time"]).astimezone(TAIPEI)
            started = datetime.fromisoformat(saved_crawl["started_at"]).astimezone(TAIPEI)
            finished = datetime.fromisoformat(saved_crawl["finished_at"]).astimezone(TAIPEI)
            period_text = f"{saved_start:%m/%d %H:%M} ～ {saved_end:%m/%d %H:%M}"
            finished_text = f"{finished:%Y-%m-%d %H:%M}"
            elapsed_text = format_elapsed((finished - started).total_seconds())
        except (TypeError, ValueError):
            period_text, finished_text, elapsed_text = "時間紀錄無法解析", str(saved_crawl["finished_at"] or ""), "—"
        st.markdown(
            f"""
            <div class="previous-result-grid">
              <div class="previous-result-card"><div class="previous-result-label">執行方法</div><div class="previous-result-value">{saved_crawl['method']}</div></div>
              <div class="previous-result-card"><div class="previous-result-label">新聞則數</div><div class="previous-result-value">{saved_crawl['rows']:,} 筆</div></div>
              <div class="previous-result-card"><div class="previous-result-label">新聞期間</div><div class="previous-result-value">{period_text}</div></div>
              <div class="previous-result-card"><div class="previous-result-label">累計執行時間</div><div class="previous-result-value">{elapsed_text}</div></div>
              <div class="previous-result-card"><div class="previous-result-label">完成時間</div><div class="previous-result-value">{finished_text}</div></div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        file_prefix = "今日" if saved_crawl["is_today"] else "本次"
        download_count = 1 + int(saved_crawl["event_path"] is not None) + int(saved_crawl["finbert_path"] is not None)
        download_columns = st.columns(download_count)
        main_label = "總新聞" if saved_crawl["method"] == "方法一＋方法二" else "新聞"
        if saved_crawl["method"] == "方法一＋方法二":
            main_download_name = saved_crawl["path"].name
        else:
            try:
                download_stamp = datetime.fromisoformat(saved_crawl["end_time"]).astimezone(TAIPEI).strftime("%Y%m%d")
            except (TypeError, ValueError):
                download_stamp = datetime.now(TAIPEI).strftime("%Y%m%d")
            main_download_name = f"新聞爬蟲_{download_stamp}.xlsx"
        download_columns[0].download_button(f"下載{file_prefix}{main_label}", saved_crawl["path"].read_bytes(), main_download_name, use_container_width=True, key="saved_crawl_all")
        column_index = 1
        if saved_crawl["event_path"]:
            download_columns[column_index].download_button(f"下載{file_prefix}負面新聞", saved_crawl["event_path"].read_bytes(), saved_crawl["event_path"].name, use_container_width=True, key="saved_crawl_event")
            column_index += 1
        if saved_crawl["finbert_path"]:
            download_columns[column_index].download_button(f"下載{file_prefix} FinBERT ≤ 0", saved_crawl["finbert_path"].read_bytes(), saved_crawl["finbert_path"].name, use_container_width=True, key="saved_crawl_finbert")
        if not saved_crawl["is_today"]:
            st.caption("今日新聞尚未完成；目前提供最近一次執行結果")

